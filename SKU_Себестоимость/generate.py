#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Генератор SKU-дашборда из файла себестоимости.
Запускайте через ОБНОВИТЬ.bat при каждом обновлении Excel.
"""
import openpyxl, re, json, sys, os, glob, shutil

FOLDER = os.path.dirname(os.path.abspath(__file__))
PARENT = os.path.dirname(FOLDER)

# Ищем Excel-файл в этой папке
xlsx_files = glob.glob(os.path.join(FOLDER, '*.xlsx'))
if not xlsx_files:
    # Пробуем родительскую папку
    xlsx_files = [
        os.path.join(PARENT, '2025-2026год анализ себестоимости по май.xlsx'),
        os.path.join(PARENT, '2025-2026 анализ себестоимости СВОДКА.xlsx'),
    ]
    xlsx_files = [f for f in xlsx_files if os.path.exists(f)]

if not xlsx_files:
    print("ОШИБКА: не найден Excel-файл.")
    print("Положите файл '2025-2026 анализ себестоимости СВОДКА.xlsx' в эту папку.")
    input("Нажмите Enter для выхода...")
    sys.exit(1)

EXCEL = xlsx_files[0]
print(f"Читаю: {os.path.basename(EXCEL)}")

MO = ["янв'25","фев'25","мар'25","апр'25","май'25","июн'25",
      "июл'25","авг'25","сен'25","окт'25","ноя'25","дек'25",
      "янв'26","фев'26","мар'26","апр'26","май'26"]

def mo_idx(sh):
    m = re.search(r'(\d{2})\.(\d{4})', sh)
    if not m: return -1
    mo, yr = int(m.group(1)), int(m.group(2))
    if yr == 2025: return mo - 1
    if yr == 2026: return 12 + mo - 1
    return -1

wb = openpyxl.load_workbook(EXCEL, data_only=True)
skus = {}

for sh in wb.sheetnames:
    idx = mo_idx(sh)
    if idx < 0 or idx > 16: continue
    ws = wb[sh]

    # Определяем раскладку колонок по строке заголовка (строка 4)
    # Формат A: Категория в col1 → Выручка col7, ВП col8
    # Формат B: Год/Месяц + Категория в col3 → Выручка col9, ВП col10
    row4 = [ws.cell(4, c).value for c in range(1, 6)]
    if row4[0] and 'Категория' in str(row4[0]):
        C_CAT, C_NAME, C_REV, C_VP = 1, 2, 7, 8   # Формат A
    else:
        C_CAT, C_NAME, C_REV, C_VP = 3, 4, 9, 10  # Формат B

    for row in ws.iter_rows(min_row=5, values_only=True):
        cat  = row[C_CAT-1]
        name = row[C_NAME-1]
        rev  = row[C_REV-1]
        vp   = row[C_VP-1]
        if not name or not isinstance(name, str): continue
        n = name.strip()
        if not n or n.lower().startswith('итого') or n.lower().startswith('всего'): continue
        if not isinstance(rev, (int, float)): continue
        if n not in skus:
            skus[n] = {'cat': str(cat or '').strip() or 'Прочее',
                       'rev': [0]*17, 'vp': [0]*17, 'qty': 0}
        if rev and rev > 0: skus[n]['rev'][idx] += float(rev)
        if vp:              skus[n]['vp'][idx]  += float(vp)

IDX_2025 = list(range(12))
IDX_2026 = list(range(12, 17))
result = []
for name, s in skus.items():
    total_rev = sum(s['rev'])
    total_vp  = sum(s['vp'])
    if total_rev <= 0: continue
    rev25 = sum(s['rev'][i] for i in IDX_2025)
    rev26 = sum(s['rev'][i] for i in IDX_2026)
    avg25 = rev25 / 12; avg26 = rev26 / 5
    trend  = round((avg26/avg25 - 1)*100) if avg25 > 0 else 0
    margin = round(total_vp/total_rev*100) if total_rev > 0 else 0
    active = sum(1 for v in s['rev'] if v > 0)
    result.append({
        'name': name, 'cat': s['cat'],
        'total_rev': round(total_rev), 'total_vp': round(total_vp),
        'total_qty': 0, 'margin': margin,
        'active_months': active, 'trend': trend,
        'monthly_rev': [round(v) for v in s['rev']],
        'monthly_vp':  [round(v) for v in s['vp']],
    })

result.sort(key=lambda x: -x['total_rev'])
data = {'mo_labels': MO, 'skus': result}
print(f"Обработано SKU: {len(result)}")

# Вычисляем месячные итоги для AN_MO_REV / AN_MO_VP
AN_MO_REV = [0]*17
AN_MO_VP  = [0]*17
for s in result:
    for i in range(17):
        AN_MO_REV[i] += s['monthly_rev'][i]
        AN_MO_VP[i]  += s['monthly_vp'][i]
AN_MO_REV = [round(v) for v in AN_MO_REV]
AN_MO_VP  = [round(v) for v in AN_MO_VP]

# Читаем шаблон (дашборд_sku_себестоимость.html из родительской папки)
TEMPLATE = os.path.join(PARENT, 'дашборд_sku_себестоимость.html')
if not os.path.exists(TEMPLATE):
    # Запасной вариант — оригинальный SKU-дашборд
    TEMPLATE = os.path.join(PARENT, 'дашборд_sku_2025-2026.html')
if not os.path.exists(TEMPLATE):
    print(f"ОШИБКА: не найден шаблон")
    input("Нажмите Enter...")
    sys.exit(1)

src = open(TEMPLATE, encoding='utf-8').read()

# Заменяем блок SKU_DATA (сохраняем JS-функции внутри того же <script>)
data_start = src.find('<script>\nconst SKU_DATA')
data_end   = src.find('</script>', data_start) + len('</script>')
skus_end   = src.find('\n', src.find('SKUS = SKU_DATA.skus', data_start)) + 1
js_functions = src[skus_end:data_end - len('</script>')]

new_json  = json.dumps(data, ensure_ascii=False, separators=(',',':'))
new_block = f'<script>\nconst SKU_DATA={new_json};\nSKUS = SKU_DATA.skus;\nconst MO = SKU_DATA.mo_labels;\n{js_functions}</script>'

result_html = src[:data_start] + new_block + src[data_end:]

# Заменяем AN_MO_REV и AN_MO_VP на актуальные значения
result_html = re.sub(
    r'const AN_MO_REV = \[[\d, ]+\];',
    f'const AN_MO_REV = {json.dumps(AN_MO_REV)};',
    result_html)
result_html = re.sub(
    r'const AN_MO_VP = \[[\d, ]+\];',
    f'const AN_MO_VP = {json.dumps(AN_MO_VP)};',
    result_html)

result_html = result_html.replace(
    '<title>SKU Анализ — ТОО «Фуд завод»</title>',
    '<title>SKU Себестоимость — ТОО «Фуд завод»</title>'
)

OUT = os.path.join(PARENT, 'дашборд_sku_себестоимость.html')
open(OUT, 'w', encoding='utf-8').write(result_html)
print(f"Сохранено: дашборд_sku_себестоимость.html ({len(result_html)//1024} KB)")

# Копируем nav.js если нужно
nav_src = os.path.join(PARENT, 'nav.js')
nav_dst = os.path.join(FOLDER, 'nav.js')
if os.path.exists(nav_src) and not os.path.exists(nav_dst):
    shutil.copy(nav_src, nav_dst)

print(f"\nИсточник данных: {os.path.basename(EXCEL)}")
print("Готово!")
