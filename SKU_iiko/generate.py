#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
iiko SKU Dashboard — автоматический генератор.
Подключается к iiko API, скачивает отчёт о продажах за 2025-2026,
строит дашборд_sku_iiko.html с фильтрами по периоду и складу.
"""

import requests, hashlib, json, re, os, sys, warnings, glob
from datetime import datetime
from collections import defaultdict

warnings.filterwarnings("ignore")  # игнорируем SSL предупреждения

# ══════════════════════════════════════════
#   НАСТРОЙКИ  (при необходимости поменяйте)
# ══════════════════════════════════════════
IIKO_URL   = "https://fudzavod.iiko.it"
IIKO_LOGIN = "GerassimenkoO"
IIKO_PASS  = "1234"
DATE_FROM  = "2025-01-01"
DATE_TO    = datetime.today().strftime("%Y-%m-%d")
# ══════════════════════════════════════════

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)
DATA_DIR   = os.path.join(SCRIPT_DIR, 'данные')
os.makedirs(DATA_DIR, exist_ok=True)

MO_RU_SHORT = {1:"янв",2:"фев",3:"мар",4:"апр",5:"май",6:"июн",
               7:"июл",8:"авг",9:"сен",10:"окт",11:"ноя",12:"дек"}

# ──────────────────────────────────────────
#   iiko REST API
# ──────────────────────────────────────────
def iiko_auth(session):
    import hashlib as _hl
    # iiko использует SHA1 для паролей
    sha1 = _hl.sha1(IIKO_PASS.encode('utf-8')).hexdigest()
    r = session.get(
        f"{IIKO_URL}/resto/api/auth",
        params={"login": IIKO_LOGIN, "pass": sha1},
        verify=False, timeout=30
    )
    if r.status_code == 200:
        token = r.text.strip().strip('"')
        if token and len(token) >= 8:
            print(f"  ✓ Авторизован, токен: {token[:16]}...")
            return token
    raise ValueError(f"Авторизация не удалась: {r.status_code} {r.text[:200]}")

def iiko_departments(session, token):
    """Возвращает dict {uuid: name} для подразделений iiko."""
    headers = {"Cookie": f"key={token}"}
    names = {}
    endpoints = [
        f"{IIKO_URL}/resto/api/corporation/departments",
        f"{IIKO_URL}/resto/api/corporation/departments?includeDeleted=false",
        f"{IIKO_URL}/resto/api/departments",
        f"{IIKO_URL}/resto/api/v2/entities/restaurants?revisionFrom=-1",
    ]
    for url in endpoints:
        try:
            r = session.get(url, headers=headers, verify=False, timeout=30)
            if not r.ok:
                continue
            data = r.json()
            # Пробуем разные форматы ответа
            items = (data if isinstance(data, list)
                     else data.get('corporateItemDtos',
                          data.get('items',
                          data.get('restaurantDtos', []))))
            for d in (items if isinstance(items, list) else []):
                if isinstance(d, dict):
                    uid  = d.get('id', '')
                    name = (d.get('name') or d.get('departmentName')
                            or d.get('restaurantName') or uid)
                    if uid:
                        names[uid] = name
            if names:
                print(f"  Подразделений найдено: {len(names)}")
                return names
        except Exception:
            pass
    return {}

def iiko_olap_fields(session, token):
    """Получить список доступных полей OLAP."""
    headers = {"Cookie": f"key={token}"}
    for url in [
        f"{IIKO_URL}/resto/api/v2/reports/olap/fields?reportType=SALES",
        f"{IIKO_URL}/resto/api/v2/reports/olap/fields",
    ]:
        try:
            r = session.get(url, headers=headers, verify=False, timeout=30)
            if r.ok:
                print(f"  Поля OLAP ({url}): {r.text[:1000]}")
                return r.json()
        except Exception as e:
            print(f"  fields endpoint error: {e}")
    return None

def iiko_olap(session, token, date_from, date_to):
    """
    OLAP-отчёт продаж: по блюду, категории и дате.
    Поля согласно официальной документации iiko REST API v2.
    """
    headers = {
        "Cookie":       f"key={token}",
        "Content-Type": "application/json"
    }
    body = {
        "reportType": "SALES",
        "buildSummary": "false",
        "groupByRowFields": [
            "OpenDate",        # дата (для группировки по месяцу в Python)
            "DishCategory",    # категория блюда
            "DishName",        # название блюда
            "Department.Id",   # подразделение (склад/цех)
        ],
        "groupByColFields": [],
        "aggregateFields": [
            "DishDiscountSumInt",   # выручка со скидкой
            "DishAmountInt",        # количество
        ],
        "filters": {
            "OpenDate.Typed": {
                "filterType": "DateRange",
                "periodType": "CUSTOM",
                "from":  date_from,
                "to":    date_to,
                "includeLow":  "true",
                "includeHigh": "true"
            },
            "DeletedWithWriteoff": {
                "filterType": "ExcludeValues",
                "values": ["DELETED_WITHOUT_WRITEOFF"]
            }
        }
    }
    r = session.post(
        f"{IIKO_URL}/resto/api/v2/reports/olap",
        json=body, headers=headers,
        verify=False, timeout=300
    )
    if not r.ok:
        print(f"  OLAP ошибка {r.status_code}: {r.text[:400]}")
    r.raise_for_status()
    return r.json()

def parse_olap_response(resp, dept_names=None):
    """
    Разбирает ответ iiko OLAP.
    Поддерживает форматы: dict с 'data'+'columns', и список dict.
    Возвращает list of dict: {month_key, cat, name, склад, rev, qty, vp}
    """
    records = []

    if isinstance(resp, dict) and 'data' in resp:
        raw = resp['data']
        # Определяем индексы колонок
        cols = resp.get('columns', [])
        col_names = []
        for c in cols:
            if isinstance(c, dict):
                col_names.append(c.get('name', c.get('id', '')))
            else:
                col_names.append(str(c))

        def ci(name_variants):
            for n in name_variants:
                for i, cn in enumerate(col_names):
                    if n.lower() in cn.lower():
                        return i
            return -1

        # Поля iiko REST API v2 (актуальные имена)
        i_month = ci(['OpenDate', 'OpenTime', 'Date'])
        i_cat   = ci(['DishCategory', 'MenuItemCategory', 'Category', 'Категория'])
        i_name  = ci(['DishName', 'MenuItem', 'Блюдо', 'Номенклатур'])
        i_store = ci(['Department', 'StoreFrom', 'Store', 'Склад'])
        i_rev   = ci(['DishDiscountSumInt', 'DishSumInt', 'DishSum', 'Выручка'])
        i_qty   = ci(['DishAmountInt', 'DishAmount', 'Количество'])
        i_vp    = ci(['ProfitInt', 'Profit', 'Прибыль'])

        for row in raw:
            if isinstance(row, dict):
                def gv(i, key):
                    if i >= 0 and i < len(col_names):
                        return row.get(col_names[i])
                    return row.get(key)
                month = gv(i_month, 'OpenDate')
                cat   = gv(i_cat,   'DishCategory') or ''
                name  = gv(i_name,  'DishName') or ''
                store_raw = gv(i_store, 'Department.Id') or ''
                store = (dept_names or {}).get(store_raw, store_raw) or 'Все склады'
                rev   = gv(i_rev,   'DishDiscountSumInt') or 0
                qty   = gv(i_qty,   'DishAmountInt') or 0
                vp    = gv(i_vp,    'ProfitInt') or 0
            elif isinstance(row, list):
                def gi(i): return row[i] if 0 <= i < len(row) else None
                month = gi(i_month)
                cat   = gi(i_cat)   or ''
                name  = gi(i_name)  or ''
                store = gi(i_store) or 'Все склады'
                rev   = gi(i_rev)   or 0
                qty   = gi(i_qty)   or 0
                vp    = gi(i_vp)    or 0
            else:
                continue

            if not name or not month: continue
            n = str(name).strip()
            if not n or n.lower().startswith('итого'): continue

            # month_key: "2025-01"  (iiko может отдавать "2025.06.01" — нормализуем)
            mk = str(month)[:7].replace('.', '-')
            if not re.match(r'20\d\d-\d\d', mk): continue

            records.append({
                'month_key': mk,
                'cat':   str(cat).strip() or 'Прочее',
                'name':  n,
                'склад': str(store).strip() or 'Все склады',
                'rev':   abs(float(rev)) if rev else 0,
                'qty':   abs(float(qty)) if qty else 0,
                'vp':    float(vp) if vp else 0,
            })

    elif isinstance(resp, list):
        # Список dict без колонок
        for row in resp:
            if not isinstance(row, dict): continue
            def gd(keys, default=''):
                for k in keys:
                    if k in row: return row[k]
                return default
            records.append({
                'month_key': str(gd(['OpenDate.YearMonth','month']))[:7],
                'cat':   str(gd(['MenuItemCategory','category','cat'], 'Прочее')).strip(),
                'name':  str(gd(['MenuItem','item','name'], '')).strip(),
                'склад': str(gd(['StoreFrom','store','склад'], 'Все склады')).strip(),
                'rev':   abs(float(gd(['DishSumInt','revenue','rev'], 0) or 0)),
                'qty':   abs(float(gd(['DishAmountInt','qty'], 0) or 0)),
                'vp':    float(gd(['ProfitInt','profit','vp'], 0) or 0),
            })

    return [r for r in records if r['name'] and re.match(r'20\d\d-\d\d', r.get('month_key',''))]

def fetch_iiko():
    print(f"  Подключение к {IIKO_URL}...")
    sess = requests.Session()
    token = iiko_auth(sess)
    print(f"  Авторизация OK")
    print(f"  Загрузка данных {DATE_FROM} — {DATE_TO}...")
    # Получаем имена подразделений для замены UUID
    dept_names = iiko_departments(sess, token)
    raw = iiko_olap(sess, token, DATE_FROM, DATE_TO)
    records = parse_olap_response(raw, dept_names)
    if not records:
        raise ValueError("OLAP вернул пустой результат. Проверьте права доступа и наличие данных.")
    print(f"  Получено строк: {len(records)}")
    return records

# ──────────────────────────────────────────
#   Запасной вариант — Excel из папки данные/
# ──────────────────────────────────────────
MO_RU_STEMS = {
    'январ':1,'феврал':2,'март':3,'апрел':4,'май':5,'июн':6,
    'июл':7,'август':8,'сентябр':9,'октябр':10,'ноябр':11,'декабр':12
}

def detect_month(filepath):
    name = os.path.basename(filepath).lower()
    m = re.search(r'(20\d\d)[_\-\.]?(0[1-9]|1[0-2])', name)
    if m: return f"{m.group(1)}-{m.group(2)}"
    m = re.search(r'(0[1-9]|1[0-2])[_\-\.](20\d\d)', name)
    if m: return f"{m.group(2)}-{m.group(1)}"
    yr_m = re.search(r'20\d\d', name)
    yr = yr_m.group() if yr_m else None
    for stem, mo in MO_RU_STEMS.items():
        if stem in name and yr:
            return f"{yr}-{mo:02d}"
    return None

def fetch_excel():
    import openpyxl
    files = sorted(glob.glob(os.path.join(DATA_DIR, '**', '*.xlsx'), recursive=True)
                 + glob.glob(os.path.join(DATA_DIR, '*.xlsx')))
    if not files:
        print(f"\n  Папка данные/ пуста: {DATA_DIR}")
        print("  Добавьте Excel-файлы из iiko или проверьте подключение к API.")
        input("Enter..."); sys.exit(1)

    records = []
    for fpath in files:
        mk = detect_month(fpath)
        if not mk:
            print(f"  ⚠  Не определён месяц для {os.path.basename(fpath)}")
            continue
        print(f"  Читаю: {os.path.basename(fpath)} → {mk}")
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
            for sname in wb.sheetnames:
                ws = wb[sname]
                # Ищем заголовок
                hrow, headers = None, []
                for r in range(1, min(8, ws.max_row+1)):
                    row = [ws.cell(r,c).value for c in range(1, ws.max_column+1)]
                    joined = ' '.join(str(v).lower() for v in row if v)
                    if 'элемент' in joined or 'номенклатур' in joined or 'категория' in joined:
                        hrow, headers = r, row; break
                if not hrow:
                    continue
                def ci(keys):
                    for k in keys:
                        for i,h in enumerate(headers):
                            if h and k in str(h).lower(): return i
                    return -1
                i_cat  = ci(['категория'])
                i_name = ci(['элемент номенклатуры','номенклатур','блюдо'])
                i_rev  = ci(['выручка','сумма прих'])
                i_qty  = ci(['количество'])
                i_vp   = ci(['валовая прибыль','вп'])
                if i_name < 0: continue
                cur_cat = ''
                for r in range(hrow+1, ws.max_row+1):
                    row = [ws.cell(r,c).value for c in range(1, ws.max_column+1)]
                    cat_v  = row[i_cat]  if i_cat>=0  else None
                    name_v = row[i_name] if i_name>=0 else None
                    if cat_v and not name_v:
                        cur_cat = str(cat_v).strip(); continue
                    if not name_v: continue
                    n = str(name_v).strip()
                    if not n or n.lower().startswith('итого'): continue
                    def num(i):
                        if i<0: return 0.0
                        v = row[i]
                        try: return abs(float(v)) if v else 0.0
                        except: return 0.0
                    records.append({
                        'month_key': mk,
                        'cat':   (str(cat_v).strip() if cat_v else cur_cat) or 'Прочее',
                        'name':  n,
                        'склад': 'Все склады',
                        'rev':   num(i_rev),
                        'qty':   num(i_qty),
                        'vp':    num(i_vp),
                    })
        except Exception as e:
            print(f"  ⚠  Ошибка {os.path.basename(fpath)}: {e}")
    return records

# ──────────────────────────────────────────
#   Запуск
# ──────────────────────────────────────────
print("=" * 55)
print("  iiko SKU Dashboard — автообновление")
print("=" * 55)

records = None
source = ""
try:
    records = fetch_iiko()
    source = f"iiko API ({IIKO_URL})"
except Exception as e:
    print(f"\n  ⚠  iiko API недоступен: {e}")
    print("  Переключаюсь на Excel-файлы из папки данные/...")
    records = fetch_excel()
    source = "Excel файлы (папка данные/)"

if not records:
    print("ОШИБКА: нет данных."); input("Enter..."); sys.exit(1)

# ──────────────────────────────────────────
#   Агрегация по SKU / месяц / склад
# ──────────────────────────────────────────
# Собираем список месяцев
mo_keys = sorted(set(r['month_key'] for r in records
                     if re.match(r'20\d\d-\d\d', r['month_key'])))
mo_index = {k: i for i, k in enumerate(mo_keys)}
N_MO = len(mo_keys)

yr_short = {2025:"'25", 2026:"'26", 2027:"'27", 2028:"'28"}
MO_LABELS = []
for k in mo_keys:
    yr, mo = int(k[:4]), int(k[5:])
    MO_LABELS.append(f"{MO_RU_SHORT[mo]}{yr_short.get(yr, k[:4])}")

IDX_2025 = [mo_index[k] for k in mo_keys if k.startswith('2025')]
IDX_2026 = [mo_index[k] for k in mo_keys if k.startswith('2026')]

# Уникальные склады (для фильтра)
склады_all = sorted(set(r['склад'] for r in records))

skus = {}
for rec in records:
    n  = rec['name']
    mk = rec['month_key']
    if mk not in mo_index: continue
    idx = mo_index[mk]
    if n not in skus:
        skus[n] = {
            'cat':    rec['cat'],
            'склады': set(),
            'rev':  [0.0]*N_MO,
            'vp':   [0.0]*N_MO,
            'qty':  [0.0]*N_MO,
        }
    skus[n]['rev'][idx]  += rec['rev']
    skus[n]['vp'][idx]   += rec['vp']
    skus[n]['qty'][idx]  += rec['qty']
    skus[n]['склады'].add(rec['склад'])

result = []
for name, s in skus.items():
    total_rev = sum(s['rev'])
    total_vp  = sum(s['vp'])
    total_qty = sum(s['qty'])
    if total_rev <= 0: continue

    rev25 = sum(s['rev'][i] for i in IDX_2025) if IDX_2025 else 0
    rev26 = sum(s['rev'][i] for i in IDX_2026) if IDX_2026 else 0
    avg25 = rev25 / len(IDX_2025) if IDX_2025 and rev25 > 0 else 0
    avg26 = rev26 / len(IDX_2026) if IDX_2026 and rev26 > 0 else 0
    trend  = round((avg26/avg25 - 1)*100) if avg25 > 0 else 0
    margin = round(total_vp/total_rev*100, 1) if total_rev > 0 else 0
    active = sum(1 for v in s['rev'] if v > 0)

    result.append({
        'name':  name,
        'cat':   s['cat'],
        'склады': sorted(s['склады']),
        'total_rev': round(total_rev),
        'total_vp':  round(total_vp),
        'total_qty': round(total_qty),
        'margin':    margin,
        'active_months': active,
        'trend': trend,
        'monthly_rev': [round(v) for v in s['rev']],
        'monthly_vp':  [round(v) for v in s['vp']],
        'monthly_qty': [round(v) for v in s['qty']],
    })

result.sort(key=lambda x: -x['total_rev'])

AN_MO_REV = [round(sum(s['monthly_rev'][i] for s in result)) for i in range(N_MO)]
AN_MO_VP  = [round(sum(s['monthly_vp'][i]  for s in result)) for i in range(N_MO)]

data = {
    'mo_labels': MO_LABELS,
    'mo_keys':   mo_keys,
    'skus':      result,
    'склады':    склады_all,
    'updated':   datetime.now().strftime("%d.%m.%Y %H:%M"),
    'source':    source,
}
new_json = json.dumps(data, ensure_ascii=False, separators=(',',':'))

print(f"\n  SKU:         {len(result)}")
print(f"  Выручка:     {sum(AN_MO_REV)/1e9:.2f} млрд ₸")
print(f"  Месяцев:     {N_MO}  ({MO_LABELS[0]} — {MO_LABELS[-1]})")
print(f"  Склады:      {', '.join(склады_all[:5])}")
print(f"  Источник:    {source}")

# ──────────────────────────────────────────
#   HTML-генерация
# ──────────────────────────────────────────
TEMPLATE = os.path.join(PARENT_DIR, 'дашборд_sku_себестоимость.html')
if not os.path.exists(TEMPLATE):
    TEMPLATE = os.path.join(PARENT_DIR, 'дашборд_sku_2025-2026.html')
if not os.path.exists(TEMPLATE):
    print("ОШИБКА: шаблон не найден."); input("Enter..."); sys.exit(1)

src = open(TEMPLATE, encoding='utf-8').read()

# Заменяем блок данных
data_start = src.find('<script>\nconst SKU_DATA')
data_end   = src.find('</script>', data_start) + len('</script>')
skus_end   = src.find('\n', src.find('SKUS = SKU_DATA.skus', data_start)) + 1
mo_end     = src.find('\n', src.find('const MO = SKU_DATA.mo_labels', skus_end)) + 1
js_funcs   = src[mo_end:data_end - len('</script>')]

new_block = (
    f'<script>\nconst SKU_DATA={new_json};\n'
    f'SKUS = SKU_DATA.skus;\n'
    f'const MO = SKU_DATA.mo_labels;\n'
    f'{js_funcs}</script>'
)
html = src[:data_start] + new_block + src[data_end:]

# Заменяем константы аналитики
html = re.sub(r'const AN_MO_REV = \[[\d, ]+\];',
              f'const AN_MO_REV = {json.dumps(AN_MO_REV)};', html)
html = re.sub(r'const AN_MO_VP  = \[[\d, ]+\];',
              f'const AN_MO_VP  = {json.dumps(AN_MO_VP)};',  html)

# КРИТИЧНО: заменяем жёстко прошитые индексы годов на реальные из данных
# Шаблон предполагает 12 мес 2025 + 5 мес 2026, но iiko может начинаться с другого месяца
html = re.sub(
    r'const IDX_2025 = Array\.from\(\{length:\d+\},\(_,i\)=>i\);',
    f'const IDX_2025 = {json.dumps(IDX_2025)};', html
)
html = re.sub(
    r'const IDX_2026 = Array\.from\(\{length:\d+\},\(_,i\)=>i\+\d+\);',
    f'const IDX_2026 = {json.dumps(IDX_2026)};', html
)

# Заголовок
html = html.replace(
    '<title>SKU Себестоимость — ТОО «Фуд завод»</title>',
    '<title>SKU Продажи iiko — ТОО «Фуд завод»</title>'
)
html = html.replace('SKU Себестоимость', 'SKU Продажи (iiko)')

# Отметка времени обновления
updated_str = datetime.now().strftime("%d.%m.%Y %H:%M")
html = html.replace(
    'class="live-pill"',
    f'class="live-pill" title="Обновлено {updated_str}"'
)

OUT = os.path.join(PARENT_DIR, 'дашборд_sku_iiko.html')
open(OUT, 'w', encoding='utf-8').write(html)
print(f"\n  → дашборд_sku_iiko.html ({os.path.getsize(OUT)//1024} KB)")
print(f"  Обновлено: {updated_str}")
print("\nГотово!")
input("Нажмите Enter для закрытия...")
