# -*- coding: utf-8 -*-
"""
gen_contractor_items.py — строит contractor_items.js:
по каждому месяцу и контрагенту (группировка по номеру, как в балансе)
— выручка, доля и полный список позиций (наименование, кол-во, сумма).

Источник — те же выгрузки «I Отчет ПРОДАЖИ» (берётся самая свежая по каждому месяцу).

Запуск: python gen_contractor_items.py
"""
import sys, os, re, json, glob, warnings
from collections import defaultdict

warnings.filterwarnings("ignore")
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

import openpyxl

HERE   = os.path.dirname(os.path.abspath(__file__))
OUT_JS = os.path.join(HERE, "contractor_items.js")
YEAR   = 2026
REVENUE_TYPE = "Выручка расходной накладной"

RU_MONTHS = ["", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
             "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]

# Красивые названия групп по номеру контрагента
BRANDS = {
    '0':  'Частное лицо',
    '1':  'Базилик 1 (Шевченко)',
    '2':  'Базилик 2 (Желтоксан)',
    '3':  'Базилик 3 (Астана)',
    '7':  'Kaspi Банк (Склад ГП)',
    '9':  'Базилик 5 (Аксай)',
    '70': 'ТОО ТУАР (Шинлайн)',
    '84': 'ТОО Гамаус',
    '85': 'ДФЗ (дистрибьютор)',
    '86': 'Crave Cafe',
    '89': 'ИП Цой Д.Л. (торты)',
    '90': 'ТОО Май Март',
    '95': 'ТОО Фуд Пикассо',
    '96': 'ТОО DSF (дистрибьютор)',
    '97': 'ИП Ник и Ко (кейтеринг)',
    '98': 'O-live',
    '99': 'RP АЗС (все точки)',
    '100':'Crave Cafe (Сейфуллина)',
    '102':'Яндекс Лавка (все точки)',
    '103':'Бобух ИП',
    '104':'DQ12 ТОО (Тараз)',
    '105':'Алга Экспресс',
    '107':'AL Group',
    '108':'GO Market',
    '109':'ALI OIL (Астана)',
    '110':'АЗС Sinooil (все точки)',
    '111':'ГЛОВО (все точки)',
    '112':'PRO OIL SERVICE',
    '113':'Wolt',
    '115':'Даркстор',
    '117':'А-Групп Холдинг',
    '119':'NOA Coffee',
    '121':'ТОО Shagyn Sauda',
}


def log(*a):
    print(*a, flush=True)


def read_period(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    txt = ''
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 2:
            txt = str(row[0]) if row and row[0] else ''
            break
    wb.close()
    m = re.search(r'с\s*(\d{2})\.(\d{2})\.(\d{4})\s*по\s*(\d{2})\.(\d{2})\.(\d{4})', txt)
    if not m:
        return None
    return {'txt': txt, 'd1': int(m.group(1)), 'm1': int(m.group(2)), 'y1': int(m.group(3)),
            'd2': int(m.group(4)), 'm2': int(m.group(5))}


def scan_reports(folder, year):
    best = {}
    for path in glob.glob(os.path.join(folder, "I Отчет*ПРОДАЖИ*.xlsx")):
        if os.path.basename(path).startswith('~$'):
            continue
        try:
            p = read_period(path)
        except Exception:
            continue
        if not p or p['y1'] != year or p['m1'] != p['m2'] or p['d1'] != 1:
            continue
        m = p['m1']
        if m not in best or p['d2'] > best[m][1]['d2']:
            best[m] = (path, p)
    return best


def group_key(name):
    """Номер контрагента. Если номера нет — сам контрагент отдельной группой."""
    s = str(name).strip()
    m = re.match(r'^\s*(\d+)', s)
    return m.group(1) if m else s


def group_label(num, sample_name):
    if not num.isdigit():
        return str(sample_name)
    if num in BRANDS:
        return f"{num}-{BRANDS[num]}"
    s = re.sub(r'^\s*\d+\s*[-–]?\s*', '', str(sample_name)).strip()
    s = re.split(r'[,:(]', s)[0].strip()
    return f"{num}-{s}" if s else str(sample_name)


def parse_month(path):
    """-> {num: {'label':.., 'rev':.., 'items': {item:{'q':..,'r':..}}, 'points':set}}"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    groups = defaultdict(lambda: {'label': None, 'rev': 0.0,
                                  'items': defaultdict(lambda: {'q': 0.0, 'r': 0.0}),
                                  'points': set(), 'sample': ''})
    cur = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 5 or not row or len(row) < 5:
            continue
        if row[0] and str(row[0]).strip():
            cur = str(row[0]).strip()
        if str(row[2]).strip() != REVENUE_TYPE:
            continue
        item = str(row[1] or '').strip()
        if not item or not cur:
            continue
        try:
            rev = float(row[4])
        except (TypeError, ValueError):
            continue
        try:
            qty = abs(float(row[3]))
        except (TypeError, ValueError):
            qty = 0.0
        num = group_key(cur)
        g = groups[num]
        if not g['sample']:
            g['sample'] = cur
        g['rev'] += rev
        g['points'].add(cur)
        it = g['items'][item]
        it['q'] += qty
        it['r'] += rev
    wb.close()
    return groups


def main():
    log("=" * 60)
    log("  Сбор данных «контрагент → что купил»")
    log("=" * 60)

    reports = scan_reports(HERE, YEAR)
    if not reports:
        log("  Выгрузки не найдены."); sys.exit(1)

    out = {}
    for m in sorted(reports):
        path, per = reports[m]
        groups = parse_month(path)
        total = sum(g['rev'] for g in groups.values())
        lst = []
        for num, g in groups.items():
            items = [{'n': n, 'q': round(v['q']), 'r': round(v['r'])}
                     for n, v in sorted(g['items'].items(), key=lambda kv: -kv[1]['r'])
                     if round(v['r']) != 0]
            if not items:
                continue
            lst.append({
                'num': num,
                'name': group_label(num, g['sample']),
                'rev': round(g['rev']),
                'pct': round(g['rev'] / total * 100, 1) if total else 0,
                'points': len(g['points']),
                'items': items,
            })
        lst.sort(key=lambda x: -x['rev'])
        mk = f"{YEAR}-{m:02d}"
        out[mk] = lst
        n_items = sum(len(c['items']) for c in lst)
        log(f"  {RU_MONTHS[m]:10} контрагентов: {len(lst):3}   позиций: {n_items:5}   выручка: {round(total):>14,}")

    # ── итог за год: те же контрагенты и позиции, но суммарно ──
    ygr = {}
    for mk, lst in out.items():
        for c in lst:
            g = ygr.setdefault(c['num'], {'num': c['num'], 'name': c['name'], 'rev': 0,
                                          'points': 0, 'items': {}})
            g['rev'] += c['rev']
            g['points'] = max(g['points'], c['points'])
            for it in c['items']:
                t = g['items'].setdefault(it['n'], {'n': it['n'], 'q': 0, 'r': 0})
                t['q'] += it['q']; t['r'] += it['r']
    ytot = sum(g['rev'] for g in ygr.values()) or 1
    ylist = []
    for g in ygr.values():
        items = sorted(g['items'].values(), key=lambda x: -x['r'])
        ylist.append({'num': g['num'], 'name': g['name'], 'rev': round(g['rev']),
                      'pct': round(g['rev'] / ytot * 100, 1), 'points': g['points'],
                      'items': items})
    ylist.sort(key=lambda x: -x['rev'])
    out['year'] = ylist
    log(f"  {'ГОД (итого)':10} контрагентов: {len(ylist):3}   выручка: {round(ytot):>14,}")

    js = "window.CTR = " + json.dumps(out, ensure_ascii=False, separators=(',', ':')) + ";\n"
    open(OUT_JS, 'w', encoding='utf-8').write(js)
    log(f"\n  Записано: {OUT_JS}  ({len(js)/1024/1024:.2f} МБ)")


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        import traceback
        log("  ОШИБКА:", e)
        traceback.print_exc()
        sys.exit(1)
