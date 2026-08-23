# -*- coding: utf-8 -*-
"""Полная себестоимость завода: анализ по ОПиУ + продажи по контрагентам.

Строит скрытую вкладку «Полная себестоимость: за счёт чего прибыль и убыток»
в дашборде себестоимости. Считает:
  • маржинальную прибыль, постоянные затраты, точку безубыточности и запас прочности по месяцам;
  • факторное разложение изменения прибыли (объём / маржинальность / постоянные) месяц к месяцу;
  • структуру затрат по 79 статьям ОПиУ;
  • выручку по каналам продаж (контрагентам) с 2025 года;
  • фудкост по категориям продукции.

Источники (лежат в репозитории):
  • «Отчет о прибылях и убытках 2025-2026.xlsx» — управленческий ОПиУ, янв-2025 … май-2026;
  • «8. Продажи 2025-2026гг..xlsx» — выручка по контрагентам помесячно;
  • «SKU_Себестоимость/2025-2026год анализ себестоимости по май.xlsx» — себестоимость по SKU.
Обновлять файлы при закрытии месяца — вкладка пересоберётся сама на следующем прогоне CI.
"""
import io, json, os, re, datetime
from collections import defaultdict
import os as _os, sys as _sys
_sys.path.insert(0, _os.path.dirname(_os.path.abspath(__file__)))
import almaty  # время завода — Алматы (UTC+5), не UTC раннера

HERE = os.path.dirname(os.path.abspath(__file__))
PL_FILE = "Отчет о прибылях и убытках 2025-2026.xlsx"
SALES_FILE = "8. Продажи 2025-2026гг..xlsx"
SKU_FILE = os.path.join("SKU_Себестоимость", "2025-2026год анализ себестоимости по май.xlsx")
TARGET = "дашборд_себестоимость_2025-2026.html"

MS = ["", "Янв", "Фев", "Мар", "Апр", "Май", "Июн", "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"]
MN = ["", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]

# ── классификация статей ОПиУ ────────────────────────────────────────────────
VARIABLE = ["1.1.Себестоимость продуктовая", "1.31.Масло фритюрное", "1.5.Расходный материал производство",
            "1.30.Возвраты от дистрибьютеров", "1.7.Истек срок хранения (порча)", "1.28.Брак", "1.24.Бракераж",
            "1.3.Недостача инвентаризации", "1.4.Излишки инвентаризации",
            "1.13.Коррекция отрицательных остатков на складе", "1.27.Нарушение тех.процесса",
            "1.26.За счет МОЛ", "1.16.Мусор", "1.18.Электроэнергия", "Логистика доставка"]
FIXED = ["Итого 2.ФОТ Производство", "Итого 3.Арендная плата", "ИТОГО 3.1.ФОТ АУП", "Итого 3.3. РазныеАдмРасходы",
         "2.4.Маркетинг", "2.5.1.Расходы по реализации Прочие", "2.5.5.Проработка новых блюд", "2.5.9.Продвижение товара",
         "1.2.Производ.расходы прочие", "1.9.Ремонт/Обслуживание производ.оборудования", "1.14.Ремонт помещений",
         "1.12.Спецодежда", "1.11.Списание сломанных ТМЗ", "1.20.Расходный материал тех.отдела",
         "1.25.Проработка блюд (текущих)", "Расходы по вознаграждениям", "Зарплата", "3.Расходы АДМ, прочие",
         "1.8.Пробы", "1.19.Вредные условия труда"]
# крупные блоки для структуры
LAYERS = [
    ("food",  "Продуктовая себестоимость", ["1.1.Себестоимость продуктовая", "1.31.Масло фритюрное"]),
    ("povh",  "Производственные накладные", ["1.2.Производ.расходы прочие", "1.3.Недостача инвентаризации",
              "1.4.Излишки инвентаризации", "1.5.Расходный материал производство", "1.7.Истек срок хранения (порча)",
              "1.9.Ремонт/Обслуживание производ.оборудования", "1.11.Списание сломанных ТМЗ", "1.12.Спецодежда",
              "1.13.Коррекция отрицательных остатков на складе", "1.14.Ремонт помещений", "1.16.Мусор",
              "1.18.Электроэнергия", "1.20.Расходный материал тех.отдела", "1.24.Бракераж",
              "1.25.Проработка блюд (текущих)", "1.26.За счет МОЛ", "1.27.Нарушение тех.процесса", "1.28.Брак",
              "1.30.Возвраты от дистрибьютеров", "1.8.Пробы", "1.19.Вредные условия труда"]),
    ("fot",   "ФОТ производства", ["Итого 2.ФОТ Производство"]),
    ("rent",  "Аренда и коммуналка", ["Итого 3.Арендная плата"]),
    ("comm",  "Реализация, логистика, маркетинг", ["Итого 2.Расходы по реализации(папка)"]),
    ("adm",   "Администрация (АУП)", ["Итого 3.Расходы АДМ", "Расходы по вознаграждениям", "Зарплата",
              "3.Расходы АДМ, прочие"]),
]
DETAIL_LINES = ["1.1.Себестоимость продуктовая", "2.1.ЗП Производство", "2.5.Налоги Производство",
                "2.4.Питание персонала", "3.1.1.ЗП АУП", "3.1.4. Налоги АУП", "3.3.2.Налоги НДС",
                "3.1.Аренда (пр-во)", "3.2.Аренда КомУсл Пр-во", "1.18.Электроэнергия", "Логистика доставка",
                "2.4.Маркетинг", "3.3.1.Админ.расходы ПРОЧИЕ", "3.3.3.Услуги охраны", "1.30.Возвраты от дистрибьютеров",
                "1.3.Недостача инвентаризации", "1.5.Расходный материал производство", "1.28.Брак",
                "1.14.Ремонт помещений", "1.9.Ремонт/Обслуживание производ.оборудования", "3.1.2.Аренда квартир д/сотрудников (АУП)",
                "1.4.Излишки инвентаризации", "1.7.Истек срок хранения (порча)", "1.2.Производ.расходы прочие"]


def load_pl():
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(HERE, PL_FILE), data_only=True)
    ws = wb.active
    cols = {}
    for i, c in enumerate(ws[5]):
        m = re.match(r'^(\d{2})\.(\d{2})\.(\d{4})$', str(c.value or "").strip())
        if m:
            cols["%s-%s" % (m.group(3), m.group(2))] = i
    rows = {}
    for r in ws.iter_rows(min_row=6, values_only=True):
        n = r[0]
        if not n:
            continue
        n = str(n).strip()
        vals = {k: (r[i] if isinstance(r[i], (int, float)) else 0) for k, i in cols.items()}
        if any(vals.values()):
            rows[n] = vals
    return sorted(cols), rows


def load_channels():
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(HERE, SALES_FILE), data_only=True, read_only=True)
    mre = re.compile(r'^(\d{2})\s*\(')
    res = defaultdict(lambda: defaultdict(float))
    names = defaultdict(lambda: defaultdict(float))
    for sheet in wb.sheetnames:
        y = re.search(r'(\d{4})', sheet)
        if not y:
            continue
        year = y.group(1)
        ws = wb[sheet]
        h5 = h6 = None
        body = []
        for i, r in enumerate(ws.iter_rows(values_only=True), 1):
            if i == 5: h5 = r
            elif i == 6: h6 = r
            elif i > 6: body.append(r)
        colmap, cur = {}, None
        for j in range(len(h6)):
            lbl = str(h5[j] or "").strip()
            m = mre.match(lbl)
            if m: cur = m.group(1)
            elif lbl.lower().startswith("итог"): cur = None
            if str(h6[j] or "").strip().startswith("Сумма прихода") and cur:
                colmap[j] = "%s-%s" % (year, cur)
        ctr = typ = None
        for r in body:
            if r[0]: ctr = str(r[0]).strip()
            if r[1]: typ = str(r[1]).strip()
            if not ctr or typ != "Выручка расходной накладной" or ctr.lower().startswith("итог"):
                continue
            for j, mk in colmap.items():
                v = r[j] if j < len(r) else None
                if isinstance(v, (int, float)) and v:
                    res[channel_of(ctr)][mk] += v
                    names[ctr][mk] += v
    return res, names


def channel_of(n):
    s = str(n).strip()
    m = re.match(r'^(\d+)\s*[-–]', s)
    num = m.group(1) if m else None
    low = s.lower()
    if num == "85" or "дфз" in low: return "ДФЗ · дистрибьютор"
    if num == "84" or "гамаус" in low: return "Гамаус · дистрибьютор"
    if num == "95" or "пикассо" in low: return "Фуд Пикассо · дистрибьютор"
    if num == "96" or "dsf" in low: return "DSF · дистрибьютор"
    if num == "90": return "Маймарт"
    if num == "102" or "яндекс" in low: return "Яндекс Лавка"
    if num == "7" or "kaspi" in low: return "Kaspi"
    if num in ("110", "99") or "азс" in low or "sinooil" in low: return "АЗС"
    if num in ("1", "2", "9") or "базилик" in low: return "Базилик"
    if "crave" in low: return "Crave Cafe"
    if "o-live" in low or num == "98": return "O-live"
    if "глово" in low or "glovo" in low: return "Glovo"
    return "Прочие"


def load_cats():
    import openpyxl
    p = os.path.join(HERE, SKU_FILE)
    if not os.path.exists(p):
        return {}, []
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    cat = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0]))
    months = set()
    for s in wb.sheetnames:
        m = re.search(r'с (\d{2})\.(\d{2})\.(\d{4})', s)
        if not m:
            continue
        mk = "%s-%s" % (m.group(3), m.group(2))
        months.add(mk)
        for r in wb[s].iter_rows(min_row=5, values_only=True):
            if not r or not r[1]:
                continue
            try:
                qty = float(r[3] or 0); cpu = float(r[4] or 0); rev = float(r[6] or 0)
            except (TypeError, ValueError):
                continue
            c = (str(r[0] or "").strip() or "Прочее")
            cat[c][mk][0] += rev
            cat[c][mk][1] += qty * cpu
    return cat, sorted(months)



def load_iiko_clients():
    """Продажи по покупателям из iiko: contractor_items.js + себестоимость SKU из sku_live.js.

    Возвращает {месяц: {покупатель: {"rev": выручка, "cost": продуктовая себестоимость,
    "cats": {категория: [выручка, себестоимость]}}}}. Себестоимость единицы берём из
    sku_live: (выручка - валовая прибыль) / количество за тот же месяц.
    """
    def _js(fn):
        pth = os.path.join(HERE, fn)
        if not os.path.exists(pth):
            return None
        t = io.open(pth, encoding="utf-8").read()
        i = t.index("=")
        obj, _ = json.JSONDecoder().raw_decode(t[i + 1:].lstrip())
        return obj

    try:
        ctr = _js("contractor_items.js")
        sku = _js("sku_live.js")
    except Exception as e:
        print("[!] покупатели: не прочитаны contractor_items/sku_live:", e)
        return {}
    if not ctr or not sku:
        return {}

    keys = sku.get("mo_keys") or []
    idx = {m: i for i, m in enumerate(keys)}
    S = {}
    for x in sku.get("skus", []):
        S[x.get("name") or x.get("n")] = x
    out = {}
    for m, clients in ctr.items():
        if m == "year" or m not in idx:
            continue
        mi = idx[m]
        row = {}
        for c in clients:
            rev = c.get("rev") or 0
            if rev <= 0:
                continue
            cost = 0.0
            known = 0.0
            cats = {}
            for it in c.get("items", []):
                x = S.get(it["n"])
                if not x:
                    continue
                q = abs((x.get("monthly_qty") or [0] * 20)[mi] or 0)
                if not q:
                    continue
                r0 = (x.get("monthly_rev") or [0] * 20)[mi] or 0
                v0 = (x.get("monthly_vp") or [0] * 20)[mi] or 0
                unit = (r0 - v0) / q
                cst = (it.get("q") or 0) * unit
                cost += cst
                known += it.get("r") or 0
                cat = x.get("cat") or "Прочее"
                a = cats.setdefault(cat, [0.0, 0.0])
                a[0] += it.get("r") or 0
                a[1] += cst
            if known > 0 and known < rev:      # позиции без калькуляции — достраиваем по средней
                k = rev / known
                cost *= k
                for a in cats.values():
                    a[0] *= k
                    a[1] *= k
            qty = sum(abs(it.get("q") or 0) for it in c.get("items", []))
            row[c["name"]] = {"rev": rev, "cost": cost, "cats": cats, "qty": qty}
        if row:
            out[m] = row
    return out



def load_returns_by_buyer():
    """Возвраты из iiko (returns_meta.js), сгруппированные по ведущему номеру контрагента.

    В контрагентах iiko номер — это, по сути, канал: «90-…» это все точки Маймарта,
    «102-…» — все точки Яндекс Лавки. Возвраты приходят по точкам, а прибыльность
    считается по каналу, поэтому складываем по номеру."""
    p = os.path.join(HERE, "returns_meta.js")
    if not os.path.exists(p):
        return {}
    try:
        t = io.open(p, encoding="utf-8").read()
        d, _ = json.JSONDecoder().raw_decode(t[t.index("=") + 1:].lstrip())
    except Exception as e:
        print("[!] возвраты не прочитаны:", e)
        return {}
    out = {}
    for c in d.get("contractors", []):
        m = re.match(r"^\s*(\d+)", str(c.get("n") or ""))
        if not m:
            continue
        a = out.setdefault(m.group(1), {"r": 0.0, "g": 0.0, "m": {}, "pts": 0})
        a["r"] += c.get("r") or 0
        a["g"] += c.get("g") or 0
        a["pts"] += 1
        for mo, v in (c.get("m") or {}).items():
            b = a["m"].setdefault(mo, [0.0, 0.0])
            b[0] += v[0]; b[1] += v[1]
    for a in out.values():
        a["s"] = round(a["r"] / a["g"] * 100, 1) if a["g"] else 0
        a["r"] = round(a["r"]); a["g"] = round(a["g"])
        a["m"] = {k: [round(v[0]), round(v[1])] for k, v in a["m"].items()}
    return out


def buyer_items(SKUidx=None):
    """Товарный разрез по каждому покупателю за 2026 год: количество, выручка,
    себестоимость (количество × себестоимость единицы из iiko), валовая прибыль."""
    def _js(fn):
        pth = os.path.join(HERE, fn)
        if not os.path.exists(pth):
            return None
        t = io.open(pth, encoding="utf-8").read()
        obj, _ = json.JSONDecoder().raw_decode(t[t.index("=") + 1:].lstrip())
        return obj
    try:
        ctr = _js("contractor_items.js"); sku = _js("sku_live.js")
    except Exception:
        return {}
    if not ctr or not sku:
        return {}
    keys = sku.get("mo_keys") or []
    y26 = [i for i, m in enumerate(keys) if m.startswith("2026")]
    unit, cat = {}, {}
    for x in sku.get("skus", []):
        q = sum(abs(x["monthly_qty"][i] or 0) for i in y26)
        if not q:
            continue
        r = sum(x["monthly_rev"][i] or 0 for i in y26)
        v = sum(x["monthly_vp"][i] or 0 for i in y26)
        unit[x["name"]] = (r - v) / q
        cat[x["name"]] = x.get("cat") or "—"
    out = {}
    for c in ctr.get("year", []):
        rows = []
        for it in c.get("items", []):
            q = it.get("q") or 0
            r = it.get("r") or 0
            if q <= 0 or r <= 0:
                continue
            u = unit.get(it["n"])
            cost = (u * q) if u is not None else None
            rows.append({"n": it["n"], "cat": cat.get(it["n"], "—"), "q": round(q), "r": round(r),
                         "c": (round(cost) if cost is not None else None),
                         "p": round(r / q), "u": (round(u) if u is not None else None)})
        rows.sort(key=lambda z: -z["r"])
        if len(rows) > 28:
            tail = rows[28:]
            rows = rows[:28]
            rows.append({"n": "Прочие позиции (%d)" % len(tail), "cat": "—",
                         "q": sum(z["q"] for z in tail), "r": sum(z["r"] for z in tail),
                         "c": sum(z["c"] for z in tail if z["c"] is not None) or None,
                         "p": None, "u": None, "rest": True})
        out[c["name"]] = rows
    return out


def build_buyers(pl, months, factors):
    """ОПиУ, разложенный на покупателей.

    Правило разнесения (оно же написано на самой вкладке):
      • выручка — доля покупателя в выручке завода по расходным накладным iiko;
      • продуктовая себестоимость — по фактическому товарному набору покупателя
        (количество × себестоимость единицы из iiko), а не по доле выручки;
      • остальные переменные — пропорционально выручке;
      • постоянные — по производству (доля покупателя в продуктовой себестоимости),
        потому что цех тянут объёмом выпуска, а не ценой продажи; разнесение по выручке
        сохранено как альтернатива (поле fixr).
    Сумма по всем покупателям в каждом месяце равна ОПиУ завода до тенге.
    """
    cl = load_iiko_clients()
    if not cl:
        return {}, []
    ms = [m for m in sorted(cl) if m in pl and pl[m].get("rev")]
    if not ms:
        return {}, []

    B = {}
    for m in ms:
        p = pl[m]
        lay = p.get("layers") or {}
        food = lay.get("food", 0)
        rest_var = p["var"] - food
        totrev = sum(x["rev"] for x in cl[m].values())
        totcost = sum(x["cost"] for x in cl[m].values()) or 1.0
        totqty = sum(x.get("qty") or 0 for x in cl[m].values()) or 1.0
        for name, x in cl[m].items():
            sh = x["rev"] / totrev if totrev else 0
            cs = x["cost"] / totcost
            layers = {k: round(v * sh) for k, v in lay.items()}
            layers["food"] = round(food * cs)
            rev = p["rev"] * sh
            var = food * cs + rest_var * sh
            # Постоянные затраты завода тянет производство, а не цена. Базой берём
            # продуктовую себестоимость покупателя: сколько цех для него реально сделал.
            # Разнесение по выручке оставляем как альтернативу — оно даёт скидку на
            # накладные тому, кто покупает дёшево, и это искажает картину.
            fix = p["fix"] * cs
            fixr = p["fix"] * sh
            cm = rev - var
            cmr = cm / rev if rev else 0
            op = cm - fix
            bep = fix / cmr if cmr > 0 else 0
            b = B.setdefault(name, {"pl": {}, "months": [], "share": {}, "cats": {}})
            b["pl"][m] = {"rev": round(rev), "var": round(var), "fix": round(fix), "cm": round(cm),
                          "cmr": round(cmr * 100, 2), "op": round(op), "bep": round(bep),
                          "safety": round((rev - bep) / rev * 100, 1) if rev else 0,
                          "gross": 0, "net": 0, "layers": layers,
                          "fixr": round(fixr), "qty": round(x.get("qty") or 0),
                          "cost": round(x["cost"]),
                          "src": p.get("src", ""), "est": p.get("est", False)}
            b["months"].append(m)
            b["share"][m] = round(sh, 6)
            for cat, a in x["cats"].items():
                t = b["cats"].setdefault(cat, {})
                t[m] = [round(a[0]), round(a[1])]

    RET = load_returns_by_buyer()
    ITEMS = buyer_items()

    out = {}
    for name, b in B.items():
        mm = sorted(b["months"])
        for i, m in enumerate(mm):
            b["pl"][m]["fx"] = None if i == 0 else factors(b["pl"][mm[i - 1]], b["pl"][m])
            b["pl"][m]["yoy"] = None
        cats = []
        for cat, mo in b["cats"].items():
            rev = sum(v[0] for v in mo.values())
            cost = sum(v[1] for v in mo.values())
            if rev < 300000 or cost <= 0:
                continue
            cats.append({"n": cat, "rev": round(rev), "cost": round(cost),
                         "fc": round(cost / rev * 100, 1), "gp": round(rev - cost), "m": mo})
        cats.sort(key=lambda x: -x["rev"])
        num = re.match(r"^\s*(\d+)", name)
        ret = RET.get(num.group(1)) if num else None
        rev = round(sum(b["pl"][m]["rev"] for m in mm))
        cm = round(sum(b["pl"][m]["cm"] for m in mm))
        fix = round(sum(b["pl"][m]["fix"] for m in mm))
        var = round(sum(b["pl"][m]["var"] for m in mm))
        out[name] = {"months": mm, "pl": b["pl"], "share": b["share"], "cats": cats[:14],
                     "rev": rev, "cm": cm, "fix": fix, "var": var,
                     "cmr": round(cm / rev * 100, 2) if rev else 0,
                     "bep": round(fix / (cm / rev)) if rev and cm > 0 else 0,
                     "op": round(sum(b["pl"][m]["op"] for m in mm)),
                     "qty": round(sum(b["pl"][m].get("qty") or 0 for m in mm)),
                     "ret": ret, "items": ITEMS.get(name) or []}
    order = sorted(out, key=lambda n: -out[n]["rev"])
    return out, order


def build():
    months, R = load_pl()

    def v(k, m):
        return R.get(k, {}).get(m, 0)

    pl = {}
    for m in months:
        rev = v("Итого Выручка", m)
        var = sum(v(k, m) for k in VARIABLE)
        fix = sum(v(k, m) for k in FIXED)
        layers = {key: sum(v(k, m) for k in keys) for key, _t, keys in LAYERS}
        # Полная себестоимость = сумма шести слоёв (она сверена с ОПиУ построчно).
        # Деление на переменные/постоянные вспомогательное, поэтому постоянные берём
        # остатком — иначе мелкая статья, не попавшая ни в один список, ломает сходимость.
        S = sum(layers.values())
        if S > 0:
            var = min(var, S)
            fix = S - var
        cm = rev - var
        cmr = cm / rev if rev else 0
        op = cm - fix
        bep = fix / cmr if cmr > 0 else 0
        pl[m] = {
            "rev": round(rev), "var": round(var), "fix": round(fix), "cm": round(cm),
            "cmr": round(cmr * 100, 2), "op": round(op), "bep": round(bep),
            "safety": round((rev - bep) / rev * 100, 1) if rev else 0,
            "gross": round(v("Валовая прибыль", m)), "net": round(v("ИТОГО ЧИСТАЯ ПРИБЫЛЬ", m)),
            "layers": {k: round(x) for k, x in layers.items()},
        }

    # ── свежие месяцы из iiko (opiu_full.json), которых ещё нет в xlsx ──
    fresh = os.path.join(HERE, "opiu_full.json")
    if os.path.exists(fresh):
        try:
            fd = json.load(open(fresh, encoding="utf-8"))
            if fd.get("check", {}).get("ok"):
                m2l = {"food": "food", "prod": "povh", "fot": "fot", "ar": "rent", "com": "comm", "adm": "adm"}
                added = 0
                for m, v in sorted(fd.get("months", {}).items()):
                    if m in pl or not v.get("rev"):
                        continue
                    if not v.get("ok", True) and not (v.get("gaps") or []):
                        continue
                    rev = v["rev"]
                    var = v.get("var") or 0
                    fix = v.get("fix") or 0
                    if not var or not fix:
                        continue
                    gaps = v.get("gaps") or []
                    layers = {m2l[k]: round(x) for k, x in v.get("abs", {}).items() if k in m2l}
                    if gaps:
                        # начисления по этим группам ещё не проведены — заменяем их средней долей
                        base = {}
                        for key, lay in m2l.items():
                            vals = [pl[x]["layers"].get(lay, 0) / pl[x]["rev"]
                                    for x in pl if pl[x].get("rev") and not pl[x].get("est")]
                            base[key] = sum(vals) / len(vals) if vals else 0
                        for key in gaps:
                            if key not in m2l:
                                continue
                            fact = (v.get("abs", {}) or {}).get(key, 0)
                            est = rev * base.get(key, 0)
                            fix = fix - fact + est          # ФОТ и АУП — постоянные затраты
                            layers[m2l[key]] = round(est)   # слои приводим к той же оценке
                    # ЖЁСТКАЯ СВЕРКА: сумма шести групп затрат — это и есть полная себестоимость.
                    # var+fix обязаны сходиться с ней до тенге, иначе операционный результат
                    # разойдётся с «% полной себестоимости» на дашборде.
                    S = sum(layers.values())
                    if S > 0:
                        var = min(var, S)          # переменные не могут быть больше всех затрат
                        fix = S - var              # постоянные — остаток, сходимость гарантирована
                    cm = rev - var
                    cmr = cm / rev if rev else 0
                    op = cm - fix
                    pl[m] = {"rev": round(rev), "var": round(var), "fix": round(fix), "cm": round(cm),
                             "cmr": round(cmr * 100, 2), "op": round(op),
                             "bep": round(fix / cmr) if cmr > 0 else 0,
                             "safety": round((rev - (fix / cmr if cmr > 0 else 0)) / rev * 100, 1) if rev else 0,
                             "gross": 0, "net": 0, "layers": layers,
                             "src": "iiko+оценка" if gaps else "iiko", "est": bool(gaps),
                             "gaps": [g for g in gaps if g in m2l]}
                    months.append(m)
                    added += 1
                months = sorted(set(months))
                if added:
                    print("добавлено месяцев из iiko: %d" % added)
        except Exception as e:
            print("[!] opiu_full.json не прочитан:", e)

    # факторное разложение изменения операционной прибыли.
    # Маржинальность берём НЕокруглённую: cmr в pl хранится с точностью 0,01%,
    # а 0,01% от 260 млн — это 26 тыс. ₸, из-за чего сумма факторов не сходилась с дельтой.
    def _cmr(p):
        return (p["rev"] - p["var"]) / p["rev"] if p["rev"] else 0.0

    def _factors(p0, p1):
        c0, c1 = _cmr(p0), _cmr(p1)
        vol = (p1["rev"] - p0["rev"]) * c0          # эффект объёма при старой маржинальности
        mar = p1["rev"] * (c1 - c0)                 # эффект маржинальности на новом объёме
        fxd = -(p1["fix"] - p0["fix"])              # эффект постоянных затрат
        # vol+mar+fix тождественно равно (op1-op0); остаток гасим в самый крупный фактор
        d = p1["op"] - p0["op"]
        parts = {"vol": vol, "mar": mar, "fix": fxd}
        resid = d - (vol + mar + fxd)
        big = max(parts, key=lambda k: abs(parts[k]))
        parts[big] += resid
        return {"vol": round(parts["vol"]), "mar": round(parts["mar"]),
                "fix": round(parts["fix"]), "d": round(d)}

    for i, m in enumerate(months):
        if i == 0:
            pl[m]["fx"] = None
            continue
        pl[m]["fx"] = _factors(pl[months[i - 1]], pl[m])
    # год к году
    for m in months:
        prev = "%d-%s" % (int(m[:4]) - 1, m[5:])
        if prev in pl:
            f = _factors(pl[prev], pl[m])
            f["prev"] = prev
            pl[m]["yoy"] = f
        else:
            pl[m]["yoy"] = None

    lines = []
    for name, vals in R.items():
        if name.startswith("Итого") or name.startswith("ИТОГО") or name in ("Валовая прибыль", "Торговая выручка", "Выручка"):
            continue
        tot = sum(vals.values())
        if abs(tot) < 500000:
            continue
        grp = "перем." if name in VARIABLE else ("постоян." if name in FIXED else "прочее")
        lines.append({"n": name, "g": grp, "m": {k: round(x) for k, x in vals.items() if x}})
    lines.sort(key=lambda x: -abs(sum(x["m"].values())))
    LOSS_LINES = ["1.30.Возвраты от дистрибьютеров", "1.28.Брак", "1.7.Истек срок хранения (порча)",
                  "1.3.Недостача инвентаризации", "1.24.Бракераж", "1.27.Нарушение тех.процесса",
                  "1.11.Списание сломанных ТМЗ"]
    keep = lines[:40]
    have = {x["n"] for x in keep}
    for nm in LOSS_LINES:
        if nm not in have and nm in R:
            keep.append({"n": nm, "g": "перем." if nm in VARIABLE else "постоян.",
                         "m": {k: round(x) for k, x in R[nm].items() if x}})
    lines = keep

    chan, names = load_channels()
    cmonths = sorted({m for d in chan.values() for m in d})
    # отсекаем неполный последний месяц продаж (менее 40% от среднего)
    tot_by_m = {m: sum(d.get(m, 0) for d in chan.values()) for m in cmonths}
    avg = sum(tot_by_m.values()) / max(1, len(tot_by_m))
    cmonths = [m for m in cmonths if tot_by_m[m] > avg * 0.4]
    chan_out = {c: {m: round(d.get(m, 0)) for m in cmonths if d.get(m)} for c, d in chan.items()}
    chan_out = {c: d for c, d in chan_out.items() if sum(d.values()) > 3000000}

    top_ctr = []
    for n, d in names.items():
        tot = sum(x for m, x in d.items() if m in cmonths)
        if tot > 20000000:
            top_ctr.append({"n": n, "t": round(tot), "m": {m: round(d.get(m, 0)) for m in cmonths if d.get(m)}})
    top_ctr.sort(key=lambda x: -x["t"])
    top_ctr = top_ctr[:40]

    cat, cmo = load_cats()
    cats = []
    for c, d in cat.items():
        rev = sum(x[0] for x in d.values()); cost = sum(x[1] for x in d.values())
        if rev < 5000000 or cost <= 0:
            continue
        fc = cost / rev
        if not (0.15 <= fc <= 0.98):
            continue
        cats.append({"n": c, "rev": round(rev), "cost": round(cost), "fc": round(fc * 100, 1),
                     "gp": round(rev - cost),
                     "m": {m: [round(x[0]), round(x[1])] for m, x in d.items() if x[0]}})
    cats.sort(key=lambda x: -x["rev"])

    y = {}
    for yr in ("2025", "2026"):
        ms = [m for m in months if m.startswith(yr)]
        if not ms:
            continue
        y[yr] = {"months": ms,
                 "rev": sum(pl[m]["rev"] for m in ms), "var": sum(pl[m]["var"] for m in ms),
                 "fix": sum(pl[m]["fix"] for m in ms), "op": sum(pl[m]["op"] for m in ms),
                 "cmr": round(sum(pl[m]["cm"] for m in ms) / max(1, sum(pl[m]["rev"] for m in ms)) * 100, 2)}

    buyers, border = build_buyers(pl, months, _factors)

    return {
        "months": months, "pl": pl, "lines": lines,
        "buyers": buyers, "border": border,
        "layers": [{"k": k, "t": t} for k, t, _ in LAYERS],
        "chan": chan_out, "cmonths": cmonths, "ctr": top_ctr,
        "cats": cats[:16], "years": y,
        "built": almaty.now().strftime("%d.%m.%Y %H:%M"),
    }


SECTION = r'''
<div id="fullcost-analytics" style="max-width:1400px;margin:26px auto 0;padding:0 16px;font-family:Inter,-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif">
  <details id="fc-details" open style="background:#111827;border:1px solid #1f2937;border-radius:16px;overflow:hidden">
    <summary style="cursor:pointer;list-style:none;padding:16px 20px;font-size:15px;font-weight:800;color:#f1f5f9;display:flex;align-items:center;gap:10px;flex-wrap:wrap;background:linear-gradient(90deg,#111827,#0f172a)">
      <span style="color:#c9a94e"><span id="fc-caret">&#9662;</span> &#129518; Полная себестоимость: за счёт чего прибыль и убыток</span>
      <span id="fc-sum" style="font-weight:600;font-size:12px;color:#94a3b8"></span>
      <span style="font-weight:500;font-size:12px;color:#64748b;margin-left:auto">ОПиУ + продажи по контрагентам с 2025 года &middot; нажмите на заголовок, чтобы свернуть</span>
    </summary>
    <div style="padding:14px 18px 22px;background:#0b1220">

      <div style="display:flex;flex-wrap:wrap;gap:8px;align-items:center;margin-bottom:14px">
        <div id="fc-period" style="display:inline-flex;background:#0f172a;border:1px solid #334155;border-radius:10px;padding:3px"></div>
        <select id="fc-month" style="background:#0f172a;color:#e2e8f0;border:1px solid #334155;border-radius:9px;padding:7px 11px;font-size:12.5px;cursor:pointer"></select>
        <div id="fc-mode" style="display:inline-flex;background:#0f172a;border:1px solid #334155;border-radius:10px;padding:3px"></div>
        <select id="fc-buyer" style="background:#0f172a;color:#e2e8f0;border:1px solid #334155;border-radius:9px;padding:7px 11px;font-size:12.5px;cursor:pointer;max-width:280px"></select>
        <button id="fc-open" type="button" style="margin-left:auto;background:#c9a94e;color:#111827;border:0;border-radius:10px;padding:9px 16px;font-size:12.5px;font-weight:800;cursor:pointer">&#128203; Полный разбор</button>
      </div>

      <div id="fc-gapnote"></div>
      <div id="fc-bnote"></div>
      <div id="fc-kpi" style="display:grid;grid-template-columns:repeat(auto-fit,minmax(158px,1fr));gap:9px"></div>
      <div id="fc-alert" style="margin-top:12px"></div>

      <div style="background:#111827;border:1px solid #1f2937;border-radius:14px;padding:14px;margin-top:12px">
        <div style="font-size:13.5px;font-weight:700;color:#f1f5f9;margin-bottom:2px">&#128201; Выручка против точки безубыточности</div>
        <div style="font-size:11.5px;color:#64748b;margin-bottom:8px">столбики — выручка, линия — сколько нужно выручки, чтобы выйти в ноль. Разрыв между ними и есть прибыль или убыток.</div>
        <div style="height:330px"><canvas id="fc-ch1"></canvas></div>
          <div id="fc-obs1" style="margin-top:10px"></div>
      </div>

      <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(400px,1fr));gap:12px;margin-top:12px">
        <div style="background:#111827;border:1px solid #1f2937;border-radius:14px;padding:14px">
          <div style="font-size:13.5px;font-weight:700;color:#f1f5f9;margin-bottom:2px">&#129521; Структура полной себестоимости</div>
          <div id="fc-struct-sub" style="font-size:11.5px;color:#64748b;margin-bottom:8px"></div>
          <div style="height:300px"><canvas id="fc-ch2"></canvas></div>
          <div id="fc-obs2" style="margin-top:10px"></div>
        </div>
        <div style="background:#111827;border:1px solid #1f2937;border-radius:14px;padding:14px">
          <div style="font-size:13.5px;font-weight:700;color:#f1f5f9;margin-bottom:2px">&#9878;&#65039; За счёт чего изменился результат</div>
          <div id="fc-fx-sub" style="font-size:11.5px;color:#64748b;margin-bottom:8px"></div>
          <div style="height:300px"><canvas id="fc-ch3"></canvas></div>
          <div id="fc-obs3" style="margin-top:10px"></div>
        </div>
      </div>


      <div id="fc-prof-card" style="background:#111827;border:1px solid #1f2937;border-radius:14px;padding:14px;margin-top:12px">
        <div style="display:flex;flex-wrap:wrap;gap:8px;align-items:center;margin-bottom:2px">
          <div style="font-size:13.5px;font-weight:700;color:#f1f5f9">&#128176; Прибыльность контрагентов: кто окупает свою долю завода</div>
          <div id="fc-prof-per" style="margin-left:auto;display:inline-flex;background:#0f172a;border:1px solid #334155;border-radius:9px;padding:3px"></div>
          <div id="fc-prof-grp" style="display:inline-flex;background:#0f172a;border:1px solid #334155;border-radius:9px;padding:3px"></div>
          <div id="fc-prof-base" style="display:inline-flex;background:#0f172a;border:1px solid #334155;border-radius:9px;padding:3px"></div>
          <div id="fc-prof-sort" style="display:inline-flex;background:#0f172a;border:1px solid #334155;border-radius:9px;padding:3px"></div>
        </div>
        <div id="fc-prof-base-note"></div>
        <div style="font-size:11.5px;color:#64748b;margin-bottom:8px">Столбик — выручка контрагента, ромб — его точка безубыточности: сколько выручки нужно, чтобы покрыть переменные затраты и свою долю постоянных. Столбик длиннее ромба — контрагент в плюсе.</div>
        <div id="fc-prof-kpi" style="display:grid;grid-template-columns:repeat(auto-fit,minmax(158px,1fr));gap:9px;margin-bottom:10px"></div>
        <div style="font-size:12.5px;font-weight:700;color:#f1f5f9;margin:10px 0 2px">&#128178; Сколько денег приносит и сколько забирает каждый</div>
        <div style="font-size:11.5px;color:#64748b;margin-bottom:6px">Результат контрагента за период: маржинальная прибыль минус его доля постоянных затрат. Вправо — приносит, влево — забирает.</div>
        <div id="fc-prof-wrap2" style="height:760px"><canvas id="fc-ch7"></canvas></div>
        <div id="fc-prof-sum" style="margin:8px 0 4px"></div>
        <div style="font-size:12.5px;font-weight:700;color:#f1f5f9;margin:14px 0 2px">&#128201; Выручка против собственного порога безубыточности</div>
        <div id="fc-prof-wrap" style="height:760px"><canvas id="fc-ch6"></canvas></div>
        <div id="fc-prof-tbl" style="overflow-x:auto;margin-top:12px"></div>
        <div style="font-size:12.5px;font-weight:700;color:#f1f5f9;margin:16px 0 2px">&#9878;&#65039; Экономика одной единицы отгрузки</div>
        <div style="font-size:11.5px;color:#64748b;margin-bottom:6px">Здесь видно, почему у разных покупателей разный результат при одинаковом товаре: цена за единицу против того, что эта единица стоит заводу. Единицы у позиций разные (штуки, порции, упаковки), поэтому сравнивать корректно похожие каналы между собой, а не сеть с дистрибьютором.</div>
        <div id="fc-prof-unit" style="overflow-x:auto"></div>
        <div id="fc-prof-note" style="margin-top:10px"></div>
        <div style="font-size:13.5px;font-weight:700;color:#f1f5f9;margin:16px 0 2px">&#129534; Разбор по каждому</div>
        <div style="font-size:11.5px;color:#64748b;margin-bottom:8px">Формулировки посчитаны из его собственных цифр: что именно делает его прибыльным или убыточным и на сколько нужно сдвинуть цену, объём или маржинальность.</div>
        <div id="fc-prof-cards"></div>
      </div>

      <div id="fc-cut-card" style="background:#111827;border:1px solid #1f2937;border-radius:14px;padding:14px;margin-top:12px">
        <div style="font-size:13.5px;font-weight:700;color:#f1f5f9;margin-bottom:2px">&#9986;&#65039; Кого можно отключить — и что будет с заводом</div>
        <div style="font-size:11.5px;color:#64748b;margin-bottom:8px">Решение «отключать или держать» принимается не по строке «результат», а по маржинальной прибыли. Ниже — правило, ранжирование и симулятор: отметьте, кого хотите отключить, и посмотрите, что станет с результатом завода.</div>
        <div id="fc-cut-rule"></div>
        <div id="fc-cut-tbl" style="overflow-x:auto;margin-top:10px"></div>
        <div style="font-size:12.5px;font-weight:700;color:#f1f5f9;margin:16px 0 6px">&#129518; Симулятор отключения</div>
        <div id="fc-cut-sim"></div>
        <div style="font-size:12.5px;font-weight:700;color:#f1f5f9;margin:18px 0 2px">&#128465;&#65039; Что убрать из продаж, не теряя клиента</div>
        <div style="font-size:11.5px;color:#64748b;margin-bottom:8px">Убирать надо не покупателя, а конкретные позиции в его матрице. Здесь связки «покупатель × товар», где цена не покрывает даже переменные затраты: каждая отгруженная штука уменьшает результат завода. Расчёт по 2026 году целиком — на одном месяце слишком мало штук, чтобы принимать решение.</div>
        <div id="fc-kill"></div>
      </div>
      <div style="background:#111827;border:1px solid #1f2937;border-radius:14px;padding:14px;margin-top:12px">
        <div style="display:flex;flex-wrap:wrap;gap:8px;align-items:center;margin-bottom:8px">
          <div style="font-size:13.5px;font-weight:700;color:#f1f5f9">&#128197; Месяц к месяцу</div>
          <span style="font-size:11.5px;color:#64748b">клик по строке — выбрать месяц</span>
          <div id="fc-cmp" style="margin-left:auto;display:inline-flex;background:#0f172a;border:1px solid #334155;border-radius:9px;padding:3px"></div>
        </div>
        <div id="fc-mom" style="overflow-x:auto"></div>
      </div>

      <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(400px,1fr));gap:12px;margin-top:12px">
        <div style="background:#111827;border:1px solid #1f2937;border-radius:14px;padding:14px">
          <div style="font-size:13.5px;font-weight:700;color:#f1f5f9;margin-bottom:8px">&#128200; Статьи затрат: что выросло и что упало</div>
          <div id="fc-lines" style="overflow-x:auto"></div>
        </div>
        <div style="background:#111827;border:1px solid #1f2937;border-radius:14px;padding:14px">
          <div id="fc-chan-title" style="font-size:13.5px;font-weight:700;color:#f1f5f9;margin-bottom:8px">&#127978; Каналы продаж</div>
          <div style="height:250px;margin-bottom:8px"><canvas id="fc-ch4"></canvas></div>
          <div id="fc-obs4" style="margin-top:10px"></div>
          <div id="fc-chan" style="overflow-x:auto"></div>
        </div>
      </div>

      <div style="background:#111827;border:1px solid #1f2937;border-radius:14px;padding:14px;margin-top:12px">
        <div style="font-size:13.5px;font-weight:700;color:#f1f5f9;margin-bottom:2px">&#127859; Фудкост по категориям продукции</div>
        <div style="font-size:11.5px;color:#64748b;margin-bottom:8px">по позициям с заполненной себестоимостью в отчёте о продажах iiko</div>
        <div style="height:300px"><canvas id="fc-ch5"></canvas></div>
          <div id="fc-obs5" style="margin-top:10px"></div>
      </div>

    </div>
  </details>

  <div id="fc-modal" style="display:none;position:fixed;inset:0;z-index:9999;background:rgba(2,6,23,.82);backdrop-filter:blur(3px);overflow-y:auto;padding:26px 14px">
    <div style="max-width:940px;margin:0 auto;background:#0f172a;border:1px solid #334155;border-radius:16px;box-shadow:0 24px 70px rgba(0,0,0,.6)">
      <div style="display:flex;align-items:center;gap:12px;padding:18px 24px;border-bottom:1px solid #1f2937;position:sticky;top:0;background:#0f172a;border-radius:16px 16px 0 0">
        <div>
          <div style="font-size:16px;font-weight:800;color:#f1f5f9;letter-spacing:-.01em">Полный разбор себестоимости и результата</div>
          <div id="fc-modal-sub" style="font-size:11.5px;color:#64748b;margin-top:2px"></div>
        </div>
        <button id="fc-close" type="button" style="margin-left:auto;background:#1e293b;color:#cbd5e1;border:1px solid #334155;border-radius:9px;padding:7px 13px;font-size:12.5px;font-weight:700;cursor:pointer">Закрыть</button>
      </div>
      <div id="fc-modal-body" style="padding:20px 26px 30px"></div>
    </div>
  </div>

  <script>window.FULLCOST = __FCDATA__;</script>
  <script>
  (function(){
    var D=window.FULLCOST; if(!D) return;
    var MS=["","Янв","Фев","Мар","Апр","Май","Июн","Июл","Авг","Сен","Окт","Ноя","Дек"];
    var MN=["","январь","февраль","март","апрель","май","июнь","июль","август","сентябрь","октябрь","ноябрь","декабрь"];
    var LT={}; D.layers.forEach(function(l){ LT[l.k]=l.t; });
    var LC={food:"#ef4444",povh:"#f97316",fot:"#eab308",rent:"#84cc16",comm:"#22d3ee",adm:"#a78bfa"};
    function mln(v){ var a=Math.abs(v)/1e6; var s=(a>=100?a.toFixed(0):a.toFixed(1)).replace(".",","); return (v<0?"−":"")+s+" млн"; }
    function pc(v,d){ d=(d==null?1:d); return (v<0?"−":"")+Math.abs(v).toFixed(d).replace(".",",")+"%"; }
    function sg(v){ return (v>0?"+":"")+mln(v).replace("−","−"); }
    function lbl(m){ return MS[+m.slice(5)]+" "+m.slice(2,4); }
    function esc(s){ return String(s).replace(/[&<>]/g,function(c){return {"&":"&amp;","<":"&lt;",">":"&gt;"}[c];}); }
    function num(v){ return String(Math.round(v)).replace(/\B(?=(\d{3})+(?!\d))/g," "); }
    function mlnS(v){ return Math.abs(v)<100000 ? ((v<0?"−":"")+num(Math.abs(v)/1000)+" тыс") : mln(v); }

    var BUYER="";
    function BO(){ return (BUYER&&D.buyers&&D.buyers[BUYER])?D.buyers[BUYER]:null; }
    function PL(){ var b=BO(); return b?b.pl:D.pl; }
    function MONTHS(){ var b=BO(); return b?b.months:D.months; }
    function CATS(){ var b=BO(); return b?(b.cats||[]):D.cats; }
    function SHARE(m){ var b=BO(); return b?(b.share[m]||0):1; }
    function LINES(){ var b=BO(); if(!b) return D.lines;
      return D.lines.map(function(l){ var mm={}; for(var k in l.m){ mm[k]=Math.round(l.m[k]*(b.share[k]||0)); } return {n:l.n,g:l.g,m:mm}; }); }
    function YRS(){ var b=BO(); if(!b) return D.years;
      var y={}; b.months.forEach(function(m){ var k=m.slice(0,4), p=b.pl[m];
        if(!y[k]) y[k]={months:[],rev:0,"var":0,fix:0,op:0,cm:0};
        y[k].months.push(m); y[k].rev+=p.rev; y[k]["var"]+=p.var; y[k].fix+=p.fix; y[k].op+=p.op; y[k].cm+=p.cm; });
      for(var k2 in y) y[k2].cmr=y[k2].rev?+(y[k2].cm/y[k2].rev*100).toFixed(2):0;
      return y; }

    var st={period:"all",month:D.months[D.months.length-1],mode:"abs",cmp:"mom"};
    function months(){ return st.period==="all"?MONTHS():MONTHS().filter(function(m){return m.indexOf(st.period)===0;}); }
    function agg(ms){
      var o={rev:0,var_:0,fix:0,op:0,layers:{}};
      D.layers.forEach(function(l){ o.layers[l.k]=0; });
      ms.forEach(function(m){ var p=PL()[m]; o.rev+=p.rev; o.var_+=p.var; o.fix+=p.fix; o.op+=p.op;
        D.layers.forEach(function(l){ o.layers[l.k]+=p.layers[l.k]||0; }); });
      o.cm=o.rev-o.var_; o.cmr=o.rev?o.cm/o.rev*100:0; o.full=o.rev-o.op;
      o.bep=o.cmr>0?o.fix/(o.cmr/100):0; o.safety=o.rev?(o.rev-o.bep)/o.rev*100:0;
      return o;
    }
    function seg(id,items,cur,cb){
      var el=document.getElementById(id); if(!el) return;
      el.innerHTML=items.map(function(it){ var on=it[0]===cur;
        return '<button type="button" data-v="'+it[0]+'" style="border:0;background:'+(on?"#c9a94e":"transparent")+';color:'+(on?"#111827":"#cbd5e1")+';font-size:12px;font-weight:700;padding:6px 12px;border-radius:8px;cursor:pointer">'+it[1]+'</button>';
      }).join("");
      el.onclick=function(e){ var b=e.target.closest("button"); if(b) cb(b.getAttribute("data-v")); };
    }


    var GAPRU={fot:"ФОТ производства",adm:"АУП",food:"продуктовая себестоимость",prod:"производственные",ar:"аренда",com:"реализация"};
    function gapNames(p){
      var g=(p&&p.gaps)||[];
      if(!g.length) return "часть затрат";
      return g.map(function(k){ return GAPRU[k]||k; }).join(" и ");
    }
    function gapNote(){
      var el=document.getElementById("fc-gapnote"); if(!el) return;
      var p=PL()[st.month];
      if(!p||!p.est){ el.innerHTML=""; return; }
      el.innerHTML='<div style="background:rgba(245,158,11,.10);border:1px solid rgba(245,158,11,.35);border-radius:11px;'
        +'padding:9px 13px;margin:0 0 10px;font-size:12px;color:#fde68a;line-height:1.6">'
        +'<b>'+MN[+st.month.slice(5)]+' '+st.month.slice(0,4)+' — из iiko, кроме одной статьи.</b> Выручка и все затраты взяты из iiko, '
        +'но <b>'+gapNames(p)+'</b> там ещё не начислен: бухгалтерия проводит его после закрытия месяца. '
        +'Если оставить как есть, месяц покажет прибыль, которой нет, поэтому эта статья подставлена по средней доле '
        +'закрытых месяцев. Как только начисление пройдёт в iiko, цифра станет фактической сама — пересобирать ничего не нужно.</div>';
    }

    function bnote(){
      var el=document.getElementById("fc-bnote"); if(!el) return;
      var b=BO();
      if(!b){ el.innerHTML=""; return; }
      var a=agg(months()), ms=months();
      var all=0, allop=0;
      ms.forEach(function(m){ all+=D.pl[m].rev; allop+=D.pl[m].op; });
      var sh=all?a.rev/all*100:0;
      var facCmr=0; (function(){ var r=0,v=0; ms.forEach(function(m){ r+=D.pl[m].rev; v+=D.pl[m].var; }); facCmr=r?(r-v)/r*100:0; })();
      var dm=a.cmr-facCmr;
      el.innerHTML='<div style="background:rgba(56,189,248,.08);border:1px solid rgba(56,189,248,.32);border-radius:12px;padding:11px 14px;margin-bottom:12px;font-size:12.5px;color:#cbd5e1;line-height:1.6">'
        +'<b style="color:#f1f5f9">'+esc(BUYER)+'</b> — доля в выручке завода <b style="color:#f1f5f9">'+pc(sh)+'</b>, '
        +'маржинальность <b style="color:'+(dm>=0?"#22c55e":"#ef4444")+'">'+pc(a.cmr)+'</b> против '+pc(facCmr)+' по заводу ('+(dm>=0?"+":"−")+Math.abs(dm).toFixed(1).replace(".",",")+' пункта).'
        +'<div style="color:#94a3b8;font-size:11.5px;margin-top:5px">Как разнесено: выручка — по расходным накладным iiko; продуктовая себестоимость — по фактическому товарному набору покупателя (количество × себестоимость единицы из iiko), поэтому маржинальность у всех разная; прочие переменные — пропорционально доле в выручке, а постоянные — по производству, то есть по доле покупателя в продуктовой себестоимости (цех тянет объём выпуска, а не цена). Сумма по всем покупателям равна ОПиУ завода до тенге. Постоянные затраты покупателя — не «его» расходы, а его доля общезаводских: при его уходе они никуда не денутся. Маржинальность здесь ниже, чем на вкладке «Себестоимость · маржа»: там из выручки вычитается только продуктовая себестоимость, а тут — все переменные затраты ОПиУ, включая логистику, электроэнергию, потери и возвраты.</div></div>';
    }

    function kpi(){
      var a=agg(months()), p=PL()[st.month];
      var neg=a.op<0;
      var cards=[
        ["Выручка",mln(a.rev),months().length+" мес.","#e2e8f0"],
        ["Маржинальная прибыль",pc(a.cmr),"выручка минус переменные","#22d3ee"],
        ["Постоянные затраты",mln(a.fix),(a.fix/months().length/1e6).toFixed(0)+" млн в месяц","#f59e0b"],
        ["Полная себестоимость",mln(a.full),Math.round(a.full/a.rev*100)+"₸ на 100₸ выручки","#fb923c"],
        [neg?"Операционный убыток":"Операционная прибыль",mln(a.op),pc(a.op/a.rev*100)+" к выручке",neg?"#ef4444":"#22c55e"],
        ["Точка безубыточности",mln(a.bep/months().length),"выручки в месяц","#a78bfa"],
        ["Запас прочности",pc(a.safety),a.safety<0?"выручки не хватает":"есть подушка",a.safety<0?"#ef4444":"#22c55e"],
        [MN[+st.month.slice(5)]+" "+st.month.slice(0,4),mln(p.op),"результат месяца",p.op<0?"#ef4444":"#22c55e"]
      ];
      document.getElementById("fc-kpi").innerHTML=cards.map(function(c){
        return '<div style="background:#111827;border:1px solid #1f2937;border-radius:12px;padding:11px 13px">'
          +'<div style="font-size:10px;color:#94a3b8;font-weight:700;letter-spacing:.04em;text-transform:uppercase;line-height:1.3">'+c[0]+'</div>'
          +'<div style="font-size:19px;font-weight:800;color:'+c[3]+';margin:5px 0 2px">'+c[1]+'</div>'
          +'<div style="font-size:10.5px;color:#64748b;line-height:1.35">'+c[2]+'</div></div>';
      }).join("");
    }

    function alertBox(){
      var a=agg(months()), gap=a.bep-a.rev, mo=months().length;
      var need=a.rev?((a.bep/a.rev-1)*100):0;
      var html;
      if(a.op<0){
        html='<b>Убыток '+mln(a.op)+'.</b> Чтобы выйти в ноль, при нынешней маржинальности '+pc(a.cmr)+' нужно либо поднять выручку на '+pc(need)+' ('+mln(gap/mo)+' в месяц), либо срезать постоянные затраты на '+mln(-a.op/mo)+' в месяц, либо поднять маржинальность на '+(-a.op/a.rev*100).toFixed(1).replace(".",",")+' пункта.';
      } else {
        html='<b>Прибыль '+mln(a.op)+'.</b> Запас прочности '+pc(a.safety)+': выручка может упасть на '+mln(a.rev-a.bep)+' до точки безубыточности.';
      }
      document.getElementById("fc-alert").innerHTML='<div style="background:'+(a.op<0?"rgba(239,68,68,.1)":"rgba(34,197,94,.1)")+';border:1px solid '+(a.op<0?"rgba(239,68,68,.32)":"rgba(34,197,94,.32)")+';border-radius:12px;padding:12px 15px;font-size:13px;color:#e2e8f0;line-height:1.65">'+html+'</div>';
    }

    function destroy(id){ var cv=document.getElementById(id); if(!cv||!window.Chart) return null; try{var e=Chart.getChart?Chart.getChart(cv):null; if(e)e.destroy();}catch(x){} return cv; }
    var AX={ticks:{color:"#64748b",font:{size:10}},grid:{color:"rgba(51,65,85,.35)"}};

    function ch1(){
      var cv=destroy("fc-ch1"); if(!cv) return; var ms=months();
      new Chart(cv.getContext("2d"),{data:{labels:ms.map(lbl),datasets:[
        {type:"bar",label:"Выручка",data:ms.map(function(m){return +(PL()[m].rev/1e6).toFixed(1);}),backgroundColor:ms.map(function(m){return PL()[m].op<0?"rgba(239,68,68,.55)":"rgba(34,197,94,.55)";}),borderRadius:5,order:3},
        {type:"line",label:"Точка безубыточности",data:ms.map(function(m){return +(PL()[m].bep/1e6).toFixed(1);}),borderColor:"#c9a94e",backgroundColor:"#c9a94e",borderWidth:2.5,tension:.25,pointRadius:3,order:1},
        {type:"line",label:"Операционная прибыль",data:ms.map(function(m){return +(PL()[m].op/1e6).toFixed(1);}),borderColor:"#38bdf8",backgroundColor:"#38bdf8",borderWidth:2,borderDash:[5,4],tension:.25,pointRadius:2,yAxisID:"y1",order:2}
      ]},options:{responsive:true,maintainAspectRatio:false,interaction:{mode:"index",intersect:false},
        plugins:{legend:{labels:{color:"#cbd5e1",font:{size:11},boxWidth:12}},datalabels:{display:false},
          tooltip:{callbacks:{label:function(c){return c.dataset.label+": "+String(c.parsed.y).replace(".",",")+" млн";}}}},
        scales:{x:{ticks:{color:"#94a3b8",font:{size:11,weight:"600"}},grid:{display:false}},
          y:Object.assign({},AX,{title:{display:true,text:"млн ₸",color:"#475569",font:{size:10}}}),
          y1:{position:"right",ticks:{color:"#38bdf8",font:{size:10}},grid:{display:false}}}}});
    }

    function ch2(){
      var cv=destroy("fc-ch2"); if(!cv) return; var ms=months(), pctMode=st.mode==="pct";
      document.getElementById("fc-struct-sub").textContent=pctMode?"доли от выручки, %":"абсолютные суммы, млн ₸";
      var ds=D.layers.map(function(l){ return {label:l.t,backgroundColor:LC[l.k],borderRadius:3,
        data:ms.map(function(m){ var x=PL()[m].layers[l.k]||0; return pctMode?+(x/PL()[m].rev*100).toFixed(1):+(x/1e6).toFixed(1); })}; });
      if(pctMode) ds.push({label:"Результат",type:"line",borderColor:"#f8fafc",backgroundColor:"#f8fafc",borderWidth:2,pointRadius:2,tension:.25,
        data:ms.map(function(m){ return +(PL()[m].op/PL()[m].rev*100).toFixed(1); })});
      new Chart(cv.getContext("2d"),{type:"bar",data:{labels:ms.map(lbl),datasets:ds},
        options:{responsive:true,maintainAspectRatio:false,
          plugins:{legend:{labels:{color:"#cbd5e1",font:{size:10},boxWidth:10}},datalabels:{display:false},
            tooltip:{callbacks:{label:function(c){return c.dataset.label+": "+String(c.parsed.y).replace(".",",")+(pctMode?"%":" млн");}}}},
          scales:{x:{stacked:true,ticks:{color:"#94a3b8",font:{size:10}},grid:{display:false}},
            y:Object.assign({stacked:true},AX)}}});
    }

    function ch3(){
      var cv=destroy("fc-ch3"); if(!cv) return;
      var p=PL()[st.month], f=(st.cmp==="yoy"?p.yoy:p.fx);
      var sub=document.getElementById("fc-fx-sub");
      if(!f){ sub.textContent="для этого месяца нет базы сравнения"; return; }
      var base=st.cmp==="yoy"?f.prev:MONTHS()[MONTHS().indexOf(st.month)-1];
      sub.textContent=MN[+st.month.slice(5)]+" "+st.month.slice(0,4)+" против "+MN[+base.slice(5)]+" "+base.slice(0,4)+" · изменение прибыли "+sg(f.d);
      var steps=[["Было",PL()[base].op,"#64748b",true],["Объём продаж",f.vol,f.vol>=0?"#22c55e":"#ef4444",false],
                 ["Маржинальность",f.mar,f.mar>=0?"#22c55e":"#ef4444",false],["Постоянные затраты",f.fix,f.fix>=0?"#22c55e":"#ef4444",false]];
      var labels=[],data=[],colors=[],cur=0;
      steps.forEach(function(s){ if(s[3]){ cur=s[1]; labels.push(s[0]); data.push([0,cur/1e6]); colors.push(s[2]); }
        else { var nx=cur+s[1]; labels.push(s[0]); data.push([cur/1e6,nx/1e6]); colors.push(s[2]); cur=nx; } });
      labels.push("Стало"); data.push([0,cur/1e6]); colors.push(cur<0?"#ef4444":"#22c55e");
      new Chart(cv.getContext("2d"),{type:"bar",data:{labels:labels,datasets:[{data:data,backgroundColor:colors,borderRadius:4,barPercentage:.7}]},
        options:{indexAxis:"y",responsive:true,maintainAspectRatio:false,
          plugins:{legend:{display:false},datalabels:{display:false},
            tooltip:{callbacks:{label:function(c){var v=c.raw;return " "+mln((v[1]-v[0])*1e6);}}}},
          scales:{x:Object.assign({},AX),y:{ticks:{color:"#cbd5e1",font:{size:11}},grid:{display:false}}}}});
    }

    function ch4(){
      var cv=destroy("fc-ch4"); if(!cv) return;
      if(BO()){
        var bm=months(), bp=PL();
        new Chart(cv.getContext("2d"),{data:{labels:bm.map(lbl),datasets:[
          {type:"bar",label:"Выручка, млн",data:bm.map(function(m){return +(bp[m].rev/1e6).toFixed(1);}),backgroundColor:"#60a5fa",borderRadius:4,order:3},
          {type:"bar",label:"Маржинальная прибыль, млн",data:bm.map(function(m){return +(bp[m].cm/1e6).toFixed(1);}),backgroundColor:"#34d399",borderRadius:4,order:2},
          {type:"line",label:"Маржинальность, %",data:bm.map(function(m){return bp[m].cmr;}),borderColor:"#c9a94e",backgroundColor:"#c9a94e",borderWidth:2,pointRadius:3,tension:.25,yAxisID:"y1",order:1}
        ]},options:{responsive:true,maintainAspectRatio:false,
          plugins:{legend:{labels:{color:"#cbd5e1",font:{size:9.5},boxWidth:9}},datalabels:{display:false}},
          scales:{x:{ticks:{color:"#94a3b8",font:{size:9.5}},grid:{display:false}},y:Object.assign({},AX),
            y1:{position:"right",ticks:{color:"#c9a94e",font:{size:10},callback:function(v){return v+"%";}},grid:{display:false}}}}});
        return;
      }
      var ms=D.cmonths.filter(function(m){ return st.period==="all"||m.indexOf(st.period)===0; });
      var names=Object.keys(D.chan).sort(function(a,b){ return sum(D.chan[b])-sum(D.chan[a]); }).slice(0,8);
      function sum(o){ var s=0; for(var k in o) s+=o[k]; return s; }
      var PAL=["#60a5fa","#f59e0b","#34d399","#a78bfa","#f472b6","#22d3ee","#fb923c","#94a3b8"];
      new Chart(cv.getContext("2d"),{type:"bar",data:{labels:ms.map(lbl),datasets:names.map(function(n,i){
        return {label:n.split(" ·")[0],backgroundColor:PAL[i%PAL.length],borderRadius:3,data:ms.map(function(m){return +((D.chan[n][m]||0)/1e6).toFixed(1);})};
      })},options:{responsive:true,maintainAspectRatio:false,
        plugins:{legend:{labels:{color:"#cbd5e1",font:{size:9.5},boxWidth:9}},datalabels:{display:false},
          tooltip:{callbacks:{label:function(c){return c.dataset.label+": "+String(c.parsed.y).replace(".",",")+" млн";}}}},
        scales:{x:{stacked:true,ticks:{color:"#94a3b8",font:{size:9.5}},grid:{display:false}},y:Object.assign({stacked:true},AX)}}});
    }

    function ch5(){
      var cv=destroy("fc-ch5"); if(!cv) return;
      var cs=CATS().slice(0,14);
      new Chart(cv.getContext("2d"),{data:{labels:cs.map(function(c){return c.n;}),datasets:[
        {type:"bar",label:"Выручка, млн",data:cs.map(function(c){return +(c.rev/1e6).toFixed(1);}),backgroundColor:"#334155",borderRadius:4,yAxisID:"y",order:3},
        {type:"bar",label:"Валовая прибыль, млн",data:cs.map(function(c){return +(c.gp/1e6).toFixed(1);}),backgroundColor:"#22c55e",borderRadius:4,yAxisID:"y",order:2},
        {type:"line",label:"Фудкост, %",data:cs.map(function(c){return c.fc;}),borderColor:"#c9a94e",backgroundColor:"#c9a94e",borderWidth:2,pointRadius:3,tension:.25,yAxisID:"y1",order:1}
      ]},options:{responsive:true,maintainAspectRatio:false,
        plugins:{legend:{labels:{color:"#cbd5e1",font:{size:10},boxWidth:10}},datalabels:{display:false}},
        scales:{x:{ticks:{color:"#94a3b8",font:{size:10},maxRotation:40,minRotation:0},grid:{display:false}},
          y:Object.assign({},AX),y1:{position:"right",ticks:{color:"#c9a94e",font:{size:10},callback:function(v){return v+"%";}},grid:{display:false}}}}});
    }

    var TONE={bad:["#ef4444","rgba(239,68,68,.10)"],warn:["#f59e0b","rgba(245,158,11,.10)"],
              good:["#22c55e","rgba(34,197,94,.10)"],info:["#38bdf8","rgba(56,189,248,.08)"],
              tip:["#c9a94e","rgba(201,169,78,.10)"]};
    var ICONS={
      "fc-obs1":[["\uD83C\uDFAF","tip"],["\uD83D\uDCC5","bad"],["\uD83C\uDFC6","info"],["\u2195\uFE0F","info"],["\uD83D\uDCC8","warn"],
                 ["\uD83E\uDDEE","info"],["\uD83D\uDCB0","good"],["\uD83D\uDE80","warn"],["\uD83D\uDD3B","bad"],["\u2728","good"]],
      "fc-obs2":[["\uD83E\uDD69","bad"],["\uD83D\uDCCA","info"],["\uD83D\uDC77","warn"],["\uD83C\uDFE2","warn"],["\uD83C\uDFED","info"],
                 ["\uD83D\uDD04","tip"],["\uD83D\uDD04","tip"],["\uD83D\uDCAF","bad"],["\uD83D\uDD0D","tip"],["\uD83C\uDFAF","good"]],
      "fc-obs3":[["\u2696\uFE0F","info"],["\uD83E\uDD47","warn"],["\uD83E\uDD48","info"],["\uD83D\uDCE6","info"],["\uD83D\uDCC9","warn"],
                 ["\uD83C\uDFE6","warn"],["\u2705","good"],["\u26A0\uFE0F","bad"],["\uD83E\uDDF0","tip"],["\uD83D\uDD01","tip"]],
      "fc-obs4":[["\uD83C\uDFEA","info"],["\uD83E\uDD47","info"],["\u26A0\uFE0F","bad"],["\uD83D\uDEAA","bad"],["\uD83D\uDCC8","good"],
                 ["\uD83D\uDCC8","good"],["\u2696\uFE0F","info"],["\uD83D\uDCB5","info"],["\uD83D\uDD22","info"],["\uD83E\uDDED","tip"]],
      "fc-obs5":[["\uD83C\uDF7D\uFE0F","info"],["\uD83E\uDD47","good"],["\uD83D\uDCB0","good"],["\uD83C\uDFC6","info"],["\uD83D\uDD34","bad"],
                 ["\uD83D\uDFE2","good"],["\u2195\uFE0F","info"],["\uD83D\uDCA1","tip"],["\uD83D\uDCCA","info"],["\u2139\uFE0F","tip"]]
    };
    function obsBlock(id, title, items){
      var el=document.getElementById(id); if(!el) return;
      var meta=ICONS[id]||[];
      var rows=items.map(function(t,i){
        var m=meta[i]||["\u2022","info"], tone=TONE[m[1]]||TONE.info;
        return '<div style="display:flex;gap:10px;align-items:flex-start;background:'+tone[1]+';border-left:3px solid '+tone[0]
          +';border-radius:9px;padding:8px 11px;margin-bottom:6px">'
          +'<div style="font-size:16px;line-height:1.3;flex:0 0 auto">'+m[0]+'</div>'
          +'<div style="font-size:12.5px;line-height:1.7;color:#cbd5e1"><b style="color:'+tone[0]+'">'+(i+1)+'.</b> '+t+'</div></div>';
      }).join("");
      el.innerHTML='<details style="background:#0b1220;border:1px solid #1f2937;border-radius:12px;padding:10px 14px">'
        +'<summary style="cursor:pointer;font-size:12.5px;font-weight:800;color:#c9a94e;list-style:none">&#128161; '+title
        +' <span style="color:#64748b;font-weight:600">— '+items.length+' наблюдений, нажмите</span></summary>'
        +'<div style="margin-top:10px">'+rows+'</div></details>';
    }
    function bb(t){ return '<b style="color:#f1f5f9">'+t+'</b>'; }
    function pp(v){ return (v>0?"+":"−")+Math.abs(v).toFixed(1).replace(".",",")+" пункта"; }

    function observations(){
      var ms=months(), a=agg(ms), n=ms.length;
      var over=ms.filter(function(m){ return PL()[m].op>0; });
      var byRev=ms.slice().sort(function(x,y){ return PL()[y].rev-PL()[x].rev; });
      var byOp=ms.slice().sort(function(x,y){ return PL()[y].op-PL()[x].op; });
      var byCmr=ms.slice().sort(function(x,y){ return PL()[y].cmr-PL()[x].cmr; });
      var byFix=ms.slice().sort(function(x,y){ return PL()[y].fix-PL()[x].fix; });
      var f=ms[0], l=ms[ms.length-1];
      var lbl=function(m){ return MN[+m.slice(5)]+" "+m.slice(0,4); };

      // 1 — выручка против точки безубыточности
      obsBlock("fc-obs1","Выручка против точки безубыточности",[
        "Порог безубыточности в среднем "+bb(mln(a.bep/n)+" в месяц")+", фактическая выручка "+bb(mln(a.rev/n))+" — разрыв "+bb(mln(Math.abs(a.bep-a.rev)/n))+" ежемесячно.",
        "Прибыльных месяцев "+bb(over.length+" из "+n)+"; в остальных выручка не дотянула до порога.",
        "Лучший месяц по результату — "+bb(lbl(byOp[0])+": "+mln(PL()[byOp[0]].op))+", худший — "+bb(lbl(byOp[byOp.length-1])+": "+mln(PL()[byOp[byOp.length-1]].op))+".",
        "Максимальная выручка была в "+lbl(byRev[0])+" ("+mln(PL()[byRev[0]].rev)+"), минимальная в "+lbl(byRev[byRev.length-1])+" ("+mln(PL()[byRev[byRev.length-1]].rev)+") — разница "+bb(mln(PL()[byRev[0]].rev-PL()[byRev[byRev.length-1]].rev))+".",
        "Сам порог не постоянный: он гуляет от "+mln(Math.min.apply(null,ms.map(function(m){return PL()[m].bep;})))+" до "+mln(Math.max.apply(null,ms.map(function(m){return PL()[m].bep;})))+" — растёт вместе с постоянными затратами и падает при росте маржинальности.",
        "В "+lbl(l)+" порог составил "+bb(mln(PL()[l].bep))+" при выручке "+mln(PL()[l].rev)+" — запас прочности "+bb(pc(PL()[l].safety))+".",
        "Каждый миллион выручки сверх порога приносит "+bb(mln(1e6*a.cmr/100))+" прибыли — это и есть эффект масштаба при нынешней структуре.",
        "Чтобы закрыть разрыв только объёмом, нужно "+bb(pc((a.bep/a.rev-1)*100))+" к нынешней выручке — примерно "+mln((a.bep-a.rev)/n)+" в месяц.",
        "Линия порога выше столбиков — визуальный признак убытка: площадь между ними за период и есть "+bb(mln(a.op))+".",
        "Если удержать выручку на уровне лучшего месяца ("+mln(PL()[byRev[0]].rev)+"), при нынешней марже результат был бы около "+bb(mln(PL()[byRev[0]].rev*a.cmr/100-a.fix/n))+" в месяц."
      ]);

      // 2 — структура затрат
      var L=D.layers.map(function(x){ return {k:x.k,t:x.t,v:a.layers[x.k]}; }).sort(function(x,y){ return y.v-x.v; });
      var fFirst=PL()[f].layers, fLast=PL()[l].layers;
      var shift=D.layers.map(function(x){
        return {t:x.t, d:(fLast[x.k]/PL()[l].rev-fFirst[x.k]/PL()[f].rev)*100};
      }).sort(function(x,y){ return Math.abs(y.d)-Math.abs(x.d); });
      obsBlock("fc-obs2","Структура полной себестоимости",[
        "Крупнейшая статья — "+bb(L[0].t+": "+mln(L[0].v))+", это "+bb(pc(L[0].v/a.rev*100))+" выручки и "+pc(L[0].v/(a.rev-a.op)*100)+" всех затрат.",
        "Вторая — "+bb(L[1].t)+" ("+pc(L[1].v/a.rev*100)+"), третья — "+L[2].t+" ("+pc(L[2].v/a.rev*100)+"). Вместе первые три дают "+bb(pc((L[0].v+L[1].v+L[2].v)/a.rev*100))+" выручки.",
        "Продукты и ФОТ производства вдвоём — "+bb(pc((a.layers.food+a.layers.fot)/a.rev*100))+": именно здесь решается судьба маржи.",
        "Администрация "+bb(pc(a.layers.adm/a.rev*100))+" — она почти не зависит от объёма, поэтому при падении выручки её доля растёт автоматически.",
        "Аренда всего "+pc(a.layers.rent/a.rev*100)+" — на фоне остальных статей это не рычаг: даже полное её обнуление не закрывает разрыв.",
        "Самое заметное изменение доли с "+lbl(f)+" по "+lbl(l)+" — "+bb(shift[0].t+" "+pp(shift[0].d))+".",
        "Второе по величине изменение — "+shift[1].t+" "+pp(shift[1].d)+", третье — "+shift[2].t+" "+pp(shift[2].d)+".",
        "Сумма всех долей за период — "+bb(pc((a.rev-a.op)/a.rev*100))+"; всё, что выше 100%, и есть операционный убыток.",
        "В режиме «% от выручки» видно главное: доли продуктов держатся ровно, а скачут ФОТ и администрация — то есть проблема не в закупе сырья.",
        "Для сравнения: чтобы выйти в ноль, суммарная доля должна опуститься до 100% — это "+bb(mln(Math.abs(a.op)/n)+" в месяц")+" экономии или соответствующий рост выручки."
      ]);

      // 3 — факторы
      var p=PL()[st.month], fx=(st.cmp==="yoy"?p.yoy:p.fx);
      var base=fx?(st.cmp==="yoy"?fx.prev:MONTHS()[MONTHS().indexOf(st.month)-1]):null;
      var items3;
      if(fx && base){
        var big=[["Объём",fx.vol],["Маржинальность",fx.mar],["Постоянные",fx.fix]].sort(function(x,y){ return Math.abs(y[1])-Math.abs(x[1]); });
        items3=[
          "Результат "+bb(lbl(st.month))+" изменился на "+bb(sg(fx.d))+" против "+lbl(base)+".",
          "Главный фактор — "+bb(big[0][0]+" "+sg(big[0][1]))+": он объясняет "+pc(Math.abs(big[0][1])/(Math.abs(fx.vol)+Math.abs(fx.mar)+Math.abs(fx.fix))*100)+" всего движения.",
          "Второй по силе — "+big[1][0]+" "+sg(big[1][1])+", третий — "+big[2][0]+" "+sg(big[2][1])+".",
          "Эффект объёма "+sg(fx.vol)+" — это изменение выручки "+sg(PL()[st.month].rev-PL()[base].rev)+", умноженное на прежнюю маржинальность "+pc(PL()[base].cmr)+".",
          "Эффект маржинальности "+sg(fx.mar)+" — это сдвиг маржи на "+pp(PL()[st.month].cmr-PL()[base].cmr)+" на нынешнем объёме.",
          "Эффект постоянных "+sg(fx.fix)+" — они "+(fx.fix<0?"выросли":"снизились")+" на "+mln(Math.abs(PL()[st.month].fix-PL()[base].fix))+".",
          "Проверка: сумма трёх факторов равна фактическому изменению прибыли до тенге — модель замкнута.",
          "Объём и маржинальность работают вместе: падение выручки при одновременном снижении маржи бьёт дважды, и именно так выглядят слабые месяцы.",
          "Постоянные затраты — единственный фактор, который не зависит от рынка: его можно менять решением, а не переговорами.",
          "Переключите сравнение на «к прошлому году» — станет видно, что за 12 месяцев изменилось структурно, а не сезонно."
        ];
      } else {
        items3=["Для выбранного месяца нет базы сравнения — выберите другой месяц или переключите режим сравнения."];
      }
      obsBlock("fc-obs3","За счёт чего изменился результат",items3);

      // 4 — каналы / покупатель
      if(BO()){
        var bm4=months(), bp4=PL();
        var facR=0, facO=0; bm4.forEach(function(m){ facR+=D.pl[m].rev; facO+=D.pl[m].op; });
        var rank=(D.border||[]).indexOf(BUYER)+1;
        var byOp4=bm4.slice().sort(function(x,y){ return bp4[y].op-bp4[x].op; });
        var pos=bm4.filter(function(m){ return bp4[m].op>0; }).length;
        var cmA=bm4.reduce(function(s2,m){ return s2+bp4[m].cm; },0);
        var fixA=bm4.reduce(function(s2,m){ return s2+bp4[m].fix; },0);
        obsBlock("fc-obs4","Покупатель: "+BUYER,[
          "Доля в выручке завода за период — "+bb(pc(a.rev/facR*100))+", по обороту это "+bb(rank+"-е место")+" из "+(D.border||[]).length+".",
          "Маржинальная прибыль "+bb(mln(cmA))+" — столько он приносит на покрытие общезаводских постоянных затрат.",
          "Его доля постоянных затрат — "+bb(mln(fixA))+"; разница и есть результат "+bb(mln(a.op))+".",
          (cmA>fixA?"Он окупает свою долю постоянных затрат — каждый следующий тенге его выручки работает на прибыль завода."
                   :"Он не окупает свою долю постоянных затрат: не хватает "+bb(mln(fixA-cmA))+" за период."),
          "Прибыльных месяцев "+bb(pos+" из "+bm4.length)+"; лучший — "+lbl(byOp4[0])+" ("+mln(bp4[byOp4[0]].op)+"), худший — "+lbl(byOp4[byOp4.length-1])+" ("+mln(bp4[byOp4[byOp4.length-1]].op)+").",
          "Точка безубыточности лично для него — "+bb(mln(a.bep/bm4.length)+" выручки в месяц")+" при фактических "+mln(a.rev/bm4.length)+".",
          "Если бы завод состоял только из таких покупателей, результат был бы "+bb(pc(a.op/a.rev*100)+" к выручке")+" против "+pc(facO/facR*100)+" фактических.",
          "Чтобы он вышел в ноль без роста объёма, его маржинальность должна быть "+bb(pc(a.fix/a.rev*100))+" — сейчас "+pc(a.cmr)+".",
          "Рост его выручки на 10% при нынешней марже дал бы "+bb(sg(a.rev*0.1*a.cmr/100))+" результата за период (постоянные затраты при этом не растут).",
          "Постоянные затраты разнесены по объёму производства — это всё равно условная база: при уходе покупателя они остаются на заводе и лягут на остальных."
        ]);
      } else {
      var cms=D.cmonths.filter(function(m){ return st.period==="all"||m.indexOf(st.period)===0; });
      var rows=Object.keys(D.chan).map(function(nm){
        var t=0; cms.forEach(function(m){ t+=D.chan[nm][m]||0; });
        var f3=cms.slice(0,3).reduce(function(s2,m){ return s2+(D.chan[nm][m]||0); },0)/Math.min(3,cms.length);
        var l3=cms.slice(-3).reduce(function(s2,m){ return s2+(D.chan[nm][m]||0); },0)/Math.min(3,cms.length);
        return {n:nm,t:t,f:f3,l:l3,d:l3-f3};
      }).sort(function(x,y){ return y.t-x.t; });
      var tot=rows.reduce(function(s2,r){ return s2+r.t; },0);
      var gone=rows.filter(function(r){ return r.f>5e6 && r.l<r.f*0.2; }).sort(function(x,y){ return x.d-y.d; });
      var grown=rows.slice().sort(function(x,y){ return y.d-x.d; });
      obsBlock("fc-obs4","Каналы продаж",[
        "Всего каналов в выборке "+bb(rows.length)+", выручка за период "+bb(mln(tot))+".",
        "Крупнейший — "+bb(rows[0].n+": "+mln(rows[0].t))+", это "+bb(pc(rows[0].t/tot*100))+" всей выручки.",
        "Три крупнейших дают "+bb(pc((rows[0].t+rows[1].t+rows[2].t)/tot*100))+" — концентрация высокая, потеря любого из них критична.",
        (gone.length?("Полностью ушли: "+bb(gone.map(function(r){return r.n;}).join(", "))+" — минус "+bb(mln(Math.abs(gone.reduce(function(s2,r){return s2+r.d;},0)))+" выручки в месяц")+"."):"Полностью выпавших каналов в этом периоде нет."),
        "Больше всех вырос "+bb(grown[0].n)+": с "+mln(grown[0].f)+" до "+mln(grown[0].l)+" в месяц, "+bb(sg(grown[0].d))+".",
        "Второй по приросту — "+grown[1].n+" "+sg(grown[1].d)+", третий — "+grown[2].n+" "+sg(grown[2].d)+".",
        "Нетто по всем каналам: "+bb(sg(rows.reduce(function(s2,r){ return s2+r.d; },0))+" выручки в месяц")+" между началом и концом периода.",
        "При марже "+pc(a.cmr)+" это "+bb(sg(rows.reduce(function(s2,r){ return s2+r.d; },0)*a.cmr/100))+" результата ежемесячно.",
        "Мелкие каналы (за пределами топ-5) дают "+pc(rows.slice(5).reduce(function(s2,r){ return s2+r.t; },0)/tot*100)+" выручки — их много, но по деньгам это хвост.",
        "Диверсификация — не абстракция: чтобы заменить крупнейший канал, нужно "+bb(Math.ceil(rows[0].t/Math.max(1,rows[5]?rows[5].t:1))+" каналов")+" размера шестого по величине."
      ]);

      }

      // 5 — категории
      if(CATS().length>=3){
        var cs=CATS().slice(0,12);
        var byFc=cs.slice().sort(function(x,y){ return x.fc-y.fc; });
        var byGp=cs.slice().sort(function(x,y){ return y.gp-x.gp; });
        var trev=cs.reduce(function(s2,c){ return s2+c.rev; },0);
        var tgp=cs.reduce(function(s2,c){ return s2+c.gp; },0);
        obsBlock("fc-obs5","Фудкост по категориям",[
          "Средний фудкост по категориям — "+bb(pc((trev-tgp)/trev*100))+"; всё, что выше, съедает маржу быстрее среднего.",
          "Самая выгодная категория — "+bb(byFc[0].n+" ("+pc(byFc[0].fc)+")")+", самая тяжёлая — "+bb(byFc[byFc.length-1].n+" ("+pc(byFc[byFc.length-1].fc)+")")+".",
          "Больше всех валовой прибыли приносит "+bb(byGp[0].n+": "+mln(byGp[0].gp))+" — это "+pc(byGp[0].gp/tgp*100)+" всей валовой прибыли выборки.",
          "Топ-3 категории дают "+bb(pc((byGp[0].gp+byGp[1].gp+byGp[2].gp)/tgp*100))+" валовой прибыли: ассортимент держится на них.",
          "Категорий с фудкостом выше 55% — "+bb(cs.filter(function(c){ return c.fc>55; }).length)+"; при полных затратах "+pc((a.rev-a.op)/a.rev*100)+" они не окупаются даже по производству.",
          "Категорий с фудкостом ниже 45% — "+bb(cs.filter(function(c){ return c.fc<45; }).length)+": именно их стоит продвигать и ставить в приоритет по мощностям.",
          "Разброс фудкоста между лучшей и худшей категорией — "+bb(pp(byFc[byFc.length-1].fc-byFc[0].fc))+", то есть рецептура и цена решают больше, чем объём.",
          "Если тяжёлые категории подтянуть к среднему уровню, валовая прибыль выборки выросла бы примерно на "+bb(mln(cs.filter(function(c){ return c.fc>(trev-tgp)/trev*100; }).reduce(function(s2,c){ return s2+c.rev*(c.fc-(trev-tgp)/trev*100)/100; },0)))+".",
          "Выручка сильно концентрирована: первая категория — "+pc(cs[0].rev/trev*100)+" оборота выборки.",
          "Данные считаются только по позициям с заполненной себестоимостью в iiko — если у SKU не заведена калькуляция, он в этот график не попадает."
        ]);
      }
    }


    /* ── Прибыльность контрагентов ─────────────────────────────────────────
       Считаем по тем же правилам, что и вкладка целиком: выручка и продуктовая
       себестоимость — фактические из iiko, прочие переменные и постоянные затраты
       разнесены: переменные по выручке, постоянные по производству. Возвраты — из расходных накладных iiko,
       сгруппированы по номеру контрагента. */
    var PSORT="op", PBASE="prod", PGRP="grp";
    function prodRatio(ms){
      var num=0,den=0;
      ms.forEach(function(m){ var p=D.pl[m]; if(!p)return; den+=p.rev;
        ["food","povh","fot","rent","comm"].forEach(function(k){ num+=(p.layers||{})[k]||0; }); });
      return den?num/den:0;
    }
    function buyerRows(){
      var B=D.buyers||{}, ms=months(), out=[];
      (D.border||[]).forEach(function(n){
        var b=B[n]; if(!b) return;
        var mm=b.months.filter(function(m){ return ms.indexOf(m)>=0; });
        if(!mm.length) return;
        var rev=0,varc=0,fix=0,cm=0,qty=0,cost=0;
        mm=mm.filter(function(m){ return perMonths().indexOf(m)>=0; });
        if(!mm.length) return;
        mm.forEach(function(m){ var p=b.pl[m]; rev+=p.rev; varc+=p.var; cm+=p.cm;
          fix+=(PBASE==="rev"?(p.fixr||0):(p.fix||0)); qty+=(p.qty||0); cost+=(p.cost||0); });
        if(rev<=0) return;
        var op=cm-fix;
        var cmr=cm/rev*100, bep=cm>0?fix/(cm/rev):0;
        var r=b.ret||null, rr=0, rs=0;
        if(r){ if(ms.length===b.months.length){ rr=r.r; rs=r.s; }
          else { var g=0; mm.forEach(function(m){ var v=(r.m||{})[m]; if(v){ rr+=v[0]; g+=v[1]; } });
                 rs=g?rr/g*100:0; } }
        out.push({n:n,rev:rev,varc:varc,fix:fix,cm:cm,cmr:cmr,op:op,bep:bep,
                  ret:rr,rets:rs,mm:mm,qty:qty,cost:cost,items:b.items||[]});
      });
      return out;
    }

    /* Группировка по номерам: в iiko номер контрагента — это, по сути, точка сети.
       Одна сеть может идти под несколькими номерами (Базилик 1, 2, 5), и по отдельности
       каждая точка выглядит мелочью. Собираем их в группу по названию, номера показываем. */
    function brandOf(n){
      var s=String(n).replace(/^\s*\d+\s*[-–—]?\s*/,'').split('(')[0];
      s=s.replace(/\s*№?\s*\d+\s*$/,'').replace(/^(ТОО|ИП|АО|ЧЛ)\s+/i,'').trim();
      return s||String(n);
    }
    function numOf(n){ var m=String(n).match(/^\s*(\d+)/); return m?m[1]:null; }
    function groupRows(rows){
      if(PGRP!=="grp") return rows;
      var by={}, order=[];
      rows.forEach(function(r){
        var b=brandOf(r.n), k=b.toLowerCase().replace(/[^0-9a-zа-яё]+/gi,'');
        if(!by[k]){ by[k]={n:b,rev:0,varc:0,fix:0,cm:0,op:0,ret:0,qty:0,cost:0,mm:[],items:[],nums:[],members:[]}; order.push(k); }
        var g=by[k];
        g.rev+=r.rev; g.varc+=r.varc; g.fix+=r.fix; g.cm+=r.cm; g.op+=r.op; g.ret+=r.ret;
        g.qty+=r.qty; g.cost+=r.cost;
        r.mm.forEach(function(m){ if(g.mm.indexOf(m)<0) g.mm.push(m); });
        (r.items||[]).forEach(function(it){ g.items.push(it); });
        var nu=numOf(r.n); if(nu&&g.nums.indexOf(nu)<0) g.nums.push(nu);
        g.members.push(r);
      });
      return order.map(function(k){
        var g=by[k];
        g.cmr=g.rev?g.cm/g.rev*100:0;
        g.bep=(g.cm>0&&g.rev)?g.fix/(g.cm/g.rev):0;
        var gg=g.members.reduce(function(s2,m){ return s2+(m.ret?m.rev+m.ret:0); },0);
        g.rets=gg?g.ret/gg*100:0;
        g.mm.sort();
        if(g.nums.length) g.n=g.n+" ["+g.nums.sort(function(a,b){return +a-+b;}).join("+")+"]";
        // товарный разрез склеиваем по названию позиции
        if(g.members.length>1){
          var agg={};
          g.items.forEach(function(it){
            if(it.rest) return;
            var e=agg[it.n]||(agg[it.n]={n:it.n,cat:it.cat,q:0,r:0,c:0,cok:true});
            e.q+=it.q; e.r+=it.r; if(it.c==null) e.cok=false; else e.c+=it.c;
          });
          g.items=Object.keys(agg).map(function(nm){ var e=agg[nm];
            return {n:e.n,cat:e.cat,q:e.q,r:e.r,c:(e.cok?e.c:null),
                    p:(e.q?Math.round(e.r/e.q):null),
                    u:(e.cok&&e.q?Math.round(e.c/e.q):null)}; })
            .sort(function(a,b){ return b.r-a.r; }).slice(0,30);
        }
        return g;
      });
    }

    function psort(rows){
      var a=rows.slice();
      if(PSORT==="rev") a.sort(function(x,y){ return y.rev-x.rev; });
      else if(PSORT==="cmr") a.sort(function(x,y){ return y.cmr-x.cmr; });
      else if(PSORT==="ret") a.sort(function(x,y){ return y.ret-x.ret; });
      else a.sort(function(x,y){ return y.op-x.op; });
      return a;
    }
    function shortN(n){ var s=String(n).replace(/\s*\(дистрибьютор\)/,"").replace(/\s*\(все точки\)/," ▸все");
      return s.length>28?s.slice(0,27)+"…":s; }

    function profKpi(rows){
      var pos=rows.filter(function(r){ return r.op>0; });
      var neg=rows.filter(function(r){ return r.op<=0; });
      var lost=neg.reduce(function(s,r){ return s+r.op; },0);
      var earn=pos.reduce(function(s,r){ return s+r.op; },0);
      var trev=rows.reduce(function(s,r){ return s+r.rev; },0);
      var tret=rows.reduce(function(s,r){ return s+r.ret; },0);
      var c=[["Окупают себя",pos.length+" из "+rows.length,pc(pos.reduce(function(s,r){return s+r.rev;},0)/trev*100,0)+" выручки","#22c55e"],
             ["Приносят сверху",mln(earn),"после своей доли постоянных","#22c55e"],
             ["Не окупают",mln(lost),neg.length+" контрагентов","#ef4444"],
             ["Возвраты",mln(tret),tret?pc(tret/(trev+tret)*100)+" от отгрузки":"нет","#fb923c"]];
      document.getElementById("fc-prof-kpi").innerHTML=c.map(function(x){
        return '<div style="background:#0f172a;border:1px solid #1f2937;border-radius:11px;padding:10px 12px">'
          +'<div style="font-size:10px;color:#94a3b8;font-weight:700;text-transform:uppercase;letter-spacing:.04em">'+x[0]+'</div>'
          +'<div style="font-size:18px;font-weight:800;color:'+x[3]+';margin:4px 0 2px">'+x[1]+'</div>'
          +'<div style="font-size:10.5px;color:#64748b">'+x[2]+'</div></div>';}).join("");
    }

    function ch6(rows){
      var cv=destroy("fc-ch6"); if(!cv) return;
      var a=psort(rows);
      document.getElementById("fc-prof-wrap").style.height=Math.max(280,a.length*21+70)+"px";
      var labels=a.map(function(r){ return shortN(r.n); });
      new Chart(cv.getContext("2d"),{data:{labels:labels,datasets:[
        {type:"bar",label:"Выручка",data:a.map(function(r){ return +(r.rev/1e6).toFixed(2); }),
         backgroundColor:a.map(function(r){ return r.op>0?"rgba(34,197,94,.62)":"rgba(239,68,68,.55)"; }),
         borderRadius:4,order:3},
        {type:"scatter",label:"Точка безубыточности",showLine:false,
         data:a.map(function(r,i){ return {x:+(r.bep/1e6).toFixed(2), y:labels[i]}; }),
         backgroundColor:"#c9a94e",borderColor:"#c9a94e",pointStyle:"rectRot",radius:6,order:1},
        {type:"scatter",label:"Возвраты",showLine:false,
         data:a.map(function(r,i){ return r.ret>50000?{x:+(r.ret/1e6).toFixed(2), y:labels[i]}:null; }).filter(Boolean),
         backgroundColor:"#fb923c",borderColor:"#fb923c",pointStyle:"triangle",radius:5,order:2}
      ]},options:{indexAxis:"y",responsive:true,maintainAspectRatio:false,
        plugins:{legend:{labels:{color:"#cbd5e1",font:{size:11},boxWidth:11,usePointStyle:true}},datalabels:{display:false},
          tooltip:{callbacks:{label:function(c){
            var r=a[c.dataIndex]||a[0];
            if(c.dataset.label==="Выручка") return "Выручка "+mln(r.rev)+" · результат "+mln(r.op);
            if(c.dataset.label==="Возвраты") return "Возвраты "+mln(r.ret)+" · "+pc(r.rets)+" отгрузки";
            return "Порог безубыточности "+mln(r.bep)+" · маржинальность "+pc(r.cmr);
          }}}},
        scales:{x:{ticks:{color:"#64748b",font:{size:10},callback:function(v){return v+" млн";}},grid:{color:"rgba(51,65,85,.35)"}},
          y:{ticks:{color:"#cbd5e1",font:{size:10.5}},grid:{display:false}}}}});
    }


    function ch7(rows){
      var cv=destroy("fc-ch7"); if(!cv) return;
      var a=rows.slice().sort(function(x,y){ return y.op-x.op; });
      document.getElementById("fc-prof-wrap2").style.height=Math.max(280,a.length*21+70)+"px";
      var mx=Math.max.apply(null,a.map(function(r){ return Math.abs(r.op); }))/1e6||1;
      mx=Math.ceil(mx*1.35/10)*10 || 10;   // круглые границы + место под подпись суммы
      // Подписи сумм рисуем сами: плагин datalabels на этой странице может быть не подключён.
      var VAL={id:"vlab"+a.length,afterDatasetsDraw:function(ch){
        var m=ch.getDatasetMeta(0), cx=ch.ctx; if(!m||!m.data) return;
        cx.save(); cx.font="700 10px system-ui,-apple-system,sans-serif"; cx.textBaseline="middle";
        m.data.forEach(function(el,i){
          var r=a[i]; if(!r) return;
          var pos=r.op>0;
          cx.fillStyle=pos?"#7ff0c0":"#fda4b4";
          cx.textAlign=pos?"left":"right";
          cx.fillText(mlnS(r.op), el.x+(pos?7:-7), el.y);
        });
        cx.restore();}};
      var pls=[VAL];
      new Chart(cv.getContext("2d"),{type:"bar",
        data:{labels:a.map(function(r){ return shortN(r.n); }),datasets:[
          {label:"Результат за период",data:a.map(function(r){ return +(r.op/1e6).toFixed(2); }),
           backgroundColor:a.map(function(r){ return r.op>0?"#22c55e":"#ef4444"; }),borderRadius:4}
        ]},
        options:{indexAxis:"y",responsive:true,maintainAspectRatio:false,
          layout:{padding:{left:8,right:8}},
          plugins:{legend:{display:false},
            tooltip:{callbacks:{label:function(c){ var r=a[c.dataIndex];
              return "Результат "+mln(r.op)+" · выручка "+mln(r.rev)+" · маржинальная прибыль "+mln(r.cm)+" · доля постоянных "+mln(r.fix); }}},
            datalabels:{display:false}},
          scales:{x:{min:-mx,max:mx,ticks:{color:"#64748b",font:{size:10},callback:function(v){return v+" млн";}},
              grid:{color:"rgba(51,65,85,.35)"}},
            y:{ticks:{color:"#cbd5e1",font:{size:10.5}},grid:{display:false}}}},
        plugins:pls});
    }

    function profSum(rows){
      var el=document.getElementById("fc-prof-sum"); if(!el) return;
      var pos=rows.filter(function(r){ return r.op>0; }).sort(function(x,y){ return y.op-x.op; });
      var neg=rows.filter(function(r){ return r.op<=0; }).sort(function(x,y){ return x.op-y.op; });
      var sp=pos.reduce(function(s,r){ return s+r.op; },0), sn=neg.reduce(function(s,r){ return s+r.op; },0);
      function lst(a){ return a.slice(0,5).map(function(r){ return '<span style="white-space:nowrap">'+esc(shortN(r.n))+' <b style="color:'+(r.op>0?"#22c55e":"#ef4444")+'">'+mln(r.op)+'</b></span>'; }).join(" · "); }
      el.innerHTML='<div style="background:#0f172a;border:1px solid #1f2937;border-radius:12px;padding:11px 14px;font-size:12.5px;color:#cbd5e1;line-height:1.7">'
        +'<div><b style="color:#22c55e">Приносят '+mln(sp)+'</b> — '+pos.length+' контрагентов: '+lst(pos)+(pos.length>5?' и ещё '+(pos.length-5):'')+'</div>'
        +'<div style="margin-top:4px"><b style="color:#ef4444">Забирают '+mln(sn)+'</b> — '+neg.length+' контрагентов: '+lst(neg)+(neg.length>5?' и ещё '+(neg.length-5):'')+'</div>'
        +'<div style="margin-top:5px;color:#94a3b8">Нетто по всем контрагентам <b style="color:'+(sp+sn<0?"#ef4444":"#22c55e")+'">'+mln(sp+sn)+'</b> — это и есть операционный результат завода за период.</div></div>';
    }


    function baseNote(rows){
      var el=document.getElementById("fc-prof-base-note"); if(!el) return;
      var t = (PBASE==="prod")
        ? "<b>Постоянные затраты разнесены по производству.</b> База — продуктовая себестоимость покупателя: "
          + "сколько цех для него реально сделал. Так и должно быть на заводе: ФОТ производства, аренда и амортизация "
          + "тянутся объёмом выпуска, а не тем, по какой цене товар продан. При этой базе покупатель с низкой ценой "
          + "не получает скидку на накладные — и сразу видно, кто не отбивает завод."
        : "<b>Постоянные затраты разнесены по выручке.</b> Классическая база, но она даёт скидку на накладные тому, "
          + "кто покупает дёшево: чем ниже цена, тем меньше «его» доля постоянных. Для сравнения покупателей с разными "
          + "ценами это искажает картину — переключите на «по производству».";
      el.innerHTML='<div style="background:rgba(56,189,248,.08);border:1px solid rgba(56,189,248,.3);border-radius:11px;'
        +'padding:9px 13px;margin:6px 0 10px;font-size:12px;color:#cbd5e1;line-height:1.6">'+t+'</div>';
    }

    function unitTable(rows){
      var el=document.getElementById("fc-prof-unit"); if(!el) return;
      var a=psort(rows).filter(function(r){ return r.qty>0; });
      if(!a.length){ el.innerHTML='<div style="color:#64748b;font-size:12px;padding:8px 2px">Нет данных по количеству.</div>'; return; }
      var h='<table style="width:100%;border-collapse:collapse;font-size:12px;min-width:840px">'
        +'<tr style="color:#64748b;font-size:10.5px;font-weight:800;text-align:right;text-transform:uppercase;letter-spacing:.04em">'
        +'<th style="text-align:left;padding:6px 4px">Контрагент</th><th style="padding:6px 4px">Единиц</th>'
        +'<th style="padding:6px 4px">Цена за ед.</th><th style="padding:6px 4px">Переменные на ед.</th>'
        +'<th style="padding:6px 4px">Вклад с ед.</th><th style="padding:6px 4px">Постоянные на ед.</th>'
        +'<th style="padding:6px 4px">Результат с ед.</th></tr>';
      a.forEach(function(r){
        var pr=r.rev/r.qty, vu=r.varc/r.qty, cu=r.cm/r.qty, fu=r.fix/r.qty, ou=r.op/r.qty;
        h+='<tr style="border-top:1px solid #1b2636">'
          +'<td style="padding:6px 4px;color:#e2e8f0;max-width:230px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">'+esc(r.n)+'</td>'
          +'<td style="padding:6px 4px;text-align:right;color:#64748b">'+num(r.qty)+'</td>'
          +'<td style="padding:6px 4px;text-align:right;color:#cbd5e1;font-weight:700">'+num(pr)+' ₸</td>'
          +'<td style="padding:6px 4px;text-align:right;color:#fb923c">'+num(vu)+' ₸</td>'
          +'<td style="padding:6px 4px;text-align:right;color:#22d3ee">'+num(cu)+' ₸</td>'
          +'<td style="padding:6px 4px;text-align:right;color:#f59e0b">'+num(fu)+' ₸</td>'
          +'<td style="padding:6px 4px;text-align:right;font-weight:800;color:'+(ou<0?"#ef4444":"#22c55e")+'">'+(ou>0?"+":"")+num(ou)+' ₸</td></tr>';
      });
      el.innerHTML=h+'</table>';
    }


    /* Когда контрагент выходит в плюс.
       Считаем по предельной логике: у завода есть свободные мощности (выручка ниже
       порога безубыточности), поэтому при росте объёма конкретного покупателя
       общезаводские постоянные затраты НЕ растут — растут только переменные.
       op(Δq,Δp) = (1+Δq)·[выручка·(1+Δp) − переменные] − его доля постоянных. */
    function scenario(r){
      function op(dq,dp){ return (1+dq)*(r.rev*(1+dp)-r.varc)-r.fix; }
      var needP = r.rev? (r.fix + r.varc)/r.rev - 1 : 0;          // только цена
      var needQ = (r.cm>0)? r.fix/r.cm - 1 : null;                // только объём
      var needM = r.rev? r.fix/r.rev*100 : 0;                     // только маржинальность
      return {op:op, needP:needP, needQ:needQ, needM:needM};
    }
    function scenarioBlock(r){
      var S=scenario(r);
      var ok = r.op>0;
      var h='<div style="font-size:11.5px;font-weight:800;color:#c9a94e;letter-spacing:.08em;text-transform:uppercase;margin:12px 0 5px">'
        +(ok?'Что держит его в плюсе':'Когда он выйдет в плюс')+'</div>';
      h+='<div style="font-size:11.5px;color:#64748b;margin-bottom:7px">Считаем по предельной логике: завод недозагружен, поэтому '
        +'дополнительный объём этого покупателя не увеличивает общезаводские постоянные затраты — растут только переменные. '
        +'Его доля постоянных зафиксирована на уровне '+mln(r.fix)+'.</div>';

      // три пути к нулю
      var pr=r.qty?r.rev/r.qty:0, prNeed=pr*(1+S.needP);
      var L = ok ? ["Запас по цене","Запас по объёму","Запас по маржинальности"]
                 : ["Только цена","Только объём","Только маржинальность"];
      var rows=[
        [L[0], (S.needP>0?"+":"")+pc(S.needP*100),
          (ok?"цена может опуститься до ":"средняя цена единицы ")+(r.qty?(num(prNeed)+" ₸ вместо "+num(pr)+" ₸"):(mln(r.rev*(1+S.needP))+" вместо "+mln(r.rev))),
          S.needP<=0],
        [L[1], (S.needQ==null?"недостижимо":((S.needQ>0?"+":"")+pc(S.needQ*100))),
          (S.needQ==null?"маржинальная прибыль отрицательная — рост объёма только увеличит убыток"
                        :((ok?"объём может упасть до ":"")+(r.qty?(num(r.qty*(1+S.needQ))+" единиц вместо "+num(r.qty)):("выручки "+mln(r.rev*(1+S.needQ)))))),
          S.needQ!=null&&S.needQ<=0],
        [L[2], pc(S.needM),
          (ok?"порог: ниже этой маржинальности он уходит в минус, сейчас "+pc(r.cmr)+" ("+pp(r.cmr-S.needM)+" запаса)"
             :"вместо нынешних "+pc(r.cmr)+" — это "+pp(S.needM-r.cmr)+" за счёт набора товара, закупа или цены"),
          S.needM<=r.cmr]
      ];
      h+='<table style="width:100%;border-collapse:collapse;font-size:12px;margin-bottom:8px">';
      rows.forEach(function(x){
        h+='<tr style="border-top:1px solid #1b2636">'
          +'<td style="padding:5px 4px;color:#cbd5e1;white-space:nowrap">'+x[0]+'</td>'
          +'<td style="padding:5px 4px;text-align:right;font-weight:800;white-space:nowrap;color:'+(x[3]?"#22c55e":"#f59e0b")+'">'+x[1]+'</td>'
          +'<td style="padding:5px 4px;color:#94a3b8">'+x[2]+'</td></tr>';
      });
      h+='</table>';

      // матрица сценариев цена × объём
      var DP=[0,.03,.05,.10], DQ=[0,.05,.10,.20];
      h+='<div style="font-size:11.5px;color:#94a3b8;margin:8px 0 4px">Результат за период при сочетании цены и объёма, млн ₸ '
        +'<span style="color:#64748b">— зелёные ячейки это плюс</span></div><div style="overflow-x:auto">'
        +'<table style="width:100%;border-collapse:collapse;font-size:11.5px;min-width:420px">'
        +'<tr><th style="text-align:left;padding:5px 4px;color:#64748b;font-size:10px;text-transform:uppercase">Цена \\ Объём</th>'
        + DQ.map(function(q){ return '<th style="padding:5px 4px;text-align:right;color:#64748b;font-size:10px">'+(q?"+"+Math.round(q*100)+"%":"как сейчас")+'</th>'; }).join("")
        +'</tr>';
      DP.forEach(function(dp){
        h+='<tr style="border-top:1px solid #1b2636"><td style="padding:5px 4px;color:#cbd5e1;white-space:nowrap">'+(dp?"+"+Math.round(dp*100)+"%":"как сейчас")+'</td>';
        DQ.forEach(function(dq){
          var v=S.op(dq,dp), pos=v>0;
          h+='<td style="padding:5px 4px;text-align:right;font-weight:700;white-space:nowrap;'
            +'background:'+(pos?"rgba(34,197,94,.13)":"rgba(239,68,68,.10)")+';color:'+(pos?"#7ff0c0":"#fda4b4")+'">'
            +mlnS(v)+'</td>';
        });
        h+='</tr>';
      });
      h+='</table></div>';

      // возвраты как отдельный рычаг
      if(r.ret>50000&&r.rets>1){
        var half=r.ret*0.5;
        h+='<div style="font-size:12px;color:#cbd5e1;line-height:1.6;margin-top:8px">Отдельный рычаг — возвраты: '
          +'сейчас '+mln(r.ret)+' ('+pc(r.rets)+' отгрузки). Сокращение вдвое вернуло бы в выручку около '+mln(half)
          +', и результат стал бы '+mlnS(r.op+half)+' — '+(r.op+half>0?'этого уже достаточно для плюса.':'этого мало, нужен ещё один рычаг.')+'</div>';
      }

      // вывод одной строкой
      var best;
      if(ok) best='Он уже в плюсе. Запас: выручка может упасть на '+pc((r.rev-r.bep)/r.rev*100)+' до нуля.';
      else if(S.needQ!=null&&S.needQ>0&&S.needQ<0.25) best='Самый дешёвый путь — объём: ему не хватает '+pc(S.needQ*100)+' загрузки, цену трогать не обязательно.';
      else if(S.needP>0&&S.needP<0.08) best='Самый дешёвый путь — цена: +'+pc(S.needP*100)+' к прайсу закрывает разрыв целиком.';
      else if(S.needQ==null) best='Ни объём, ни скидка не помогут: у него отрицательная маржинальная прибыль, каждая новая отгрузка увеличивает убыток. Сначала цена или набор товара.';
      else best='Одним рычагом не закрыть: нужно сочетание — например '+pc(0.05*100)+' к цене и '+pc(20)+' к объёму даёт '+mlnS(S.op(0.20,0.05))+'.';
      h+='<div style="background:rgba(232,199,102,.10);border:1px solid rgba(232,199,102,.32);border-radius:10px;'
        +'padding:9px 12px;margin-top:9px;font-size:12.5px;color:#f2dd9e;line-height:1.6"><b>Вывод.</b> '+best+'</div>';
      return h;
    }

    function itemsTable(r){
      var it=r.items||[]; if(!it.length) return '';
      var tot=it.reduce(function(s,x){ return s+x.r; },0);
      var h='<div style="font-size:11.5px;font-weight:800;color:#c9a94e;letter-spacing:.08em;text-transform:uppercase;margin:12px 0 5px">Товарный анализ</div>'
        +'<div style="font-size:11.5px;color:#64748b;margin-bottom:6px">Что именно он берёт и по какой цене. «Наценка» — цена минус себестоимость единицы из iiko; это валовая наценка до постоянных затрат.</div>'
        +'<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:11.5px;min-width:700px">'
        +'<tr style="color:#64748b;font-size:10px;font-weight:800;text-align:right;text-transform:uppercase;letter-spacing:.04em">'
        +'<th style="text-align:left;padding:5px 4px">Позиция</th><th style="padding:5px 4px">Кол-во</th><th style="padding:5px 4px">Выручка</th>'
        +'<th style="padding:5px 4px">Доля</th><th style="padding:5px 4px">Цена ед.</th><th style="padding:5px 4px">С/с ед.</th>'
        +'<th style="padding:5px 4px">Наценка</th><th style="padding:5px 4px">Вал. прибыль</th><th style="padding:5px 4px">Маржа</th></tr>';
      it.forEach(function(x){
        var gp=(x.c!=null)?(x.r-x.c):null, mg=(gp!=null&&x.r)?gp/x.r*100:null;
        var mk=(x.p!=null&&x.u!=null)?(x.p-x.u):null;
        h+='<tr style="border-top:1px solid #1b2636'+(x.rest?';color:#64748b':'')+'">'
          +'<td style="padding:5px 4px;color:'+(x.rest?"#64748b":"#cbd5e1")+';max-width:250px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">'+esc(x.n)+'</td>'
          +'<td style="padding:5px 4px;text-align:right;color:#94a3b8">'+num(x.q)+'</td>'
          +'<td style="padding:5px 4px;text-align:right;color:#cbd5e1">'+mln(x.r)+'</td>'
          +'<td style="padding:5px 4px;text-align:right;color:#64748b">'+pc(x.r/tot*100,1)+'</td>'
          +'<td style="padding:5px 4px;text-align:right;color:#e2e8f0">'+(x.p!=null?num(x.p)+" ₸":"—")+'</td>'
          +'<td style="padding:5px 4px;text-align:right;color:#94a3b8">'+(x.u!=null?num(x.u)+" ₸":"—")+'</td>'
          +'<td style="padding:5px 4px;text-align:right;color:'+(mk!=null&&mk<0?"#ef4444":"#a78bfa")+'">'+(mk!=null?num(mk)+" ₸":"—")+'</td>'
          +'<td style="padding:5px 4px;text-align:right;color:'+(gp!=null&&gp<0?"#ef4444":"#22c55e")+'">'+(gp!=null?mln(gp):"—")+'</td>'
          +'<td style="padding:5px 4px;text-align:right;font-weight:700;color:'+(mg!=null&&mg<40?"#fb923c":"#22c55e")+'">'+(mg!=null?pc(mg):"—")+'</td></tr>';
      });
      // самые слабые позиции
      var weak=it.filter(function(x){ return !x.rest && x.c!=null && x.r>200000; })
                 .map(function(x){ return {n:x.n, m:(x.r-x.c)/x.r*100, r:x.r}; })
                 .sort(function(x,y){ return x.m-y.m; }).slice(0,3);
      var tail='';
      if(weak.length){
        tail='<div style="font-size:11.5px;color:#94a3b8;line-height:1.6;margin-top:7px">Самые слабые позиции у него: '
          + weak.map(function(w){ return '<b style="color:#fda4b4">'+esc(w.n)+'</b> ('+pc(w.m)+', '+mln(w.r)+')'; }).join(", ")
          + '. Именно по ним пересмотр цены или замена в матрице даёт больше всего.</div>';
      }
      return h+'</table></div>'+tail;
    }

    function profTable(rows){
      var a=psort(rows), trev=rows.reduce(function(s,r){ return s+r.rev; },0);
      var h='<table style="width:100%;border-collapse:collapse;font-size:12px;min-width:900px">'
        +'<tr style="color:#64748b;font-size:10.5px;font-weight:800;text-align:right;text-transform:uppercase;letter-spacing:.04em">'
        +'<th style="text-align:left;padding:6px 4px">Контрагент</th><th style="padding:6px 4px">Выручка</th><th style="padding:6px 4px">Доля</th>'
        +'<th style="padding:6px 4px">Маржин.</th><th style="padding:6px 4px">Маржин. прибыль</th><th style="padding:6px 4px">Доля постоянных</th>'
        +'<th style="padding:6px 4px">Результат</th><th style="padding:6px 4px">Порог</th><th style="padding:6px 4px">Запас</th>'
        +'<th style="padding:6px 4px">Возвраты</th></tr>';
      a.forEach(function(r){
        var zap=r.rev&&r.bep?(r.rev-r.bep)/r.rev*100:0;
        h+='<tr style="border-top:1px solid #1b2636">'
          +'<td style="padding:6px 4px;color:#e2e8f0;max-width:230px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">'+esc(r.n)+'</td>'
          +'<td style="padding:6px 4px;text-align:right;color:#cbd5e1">'+mln(r.rev)+'</td>'
          +'<td style="padding:6px 4px;text-align:right;color:#64748b">'+pc(r.rev/trev*100,1)+'</td>'
          +'<td style="padding:6px 4px;text-align:right;color:#22d3ee;font-weight:700">'+pc(r.cmr)+'</td>'
          +'<td style="padding:6px 4px;text-align:right;color:#94a3b8">'+mln(r.cm)+'</td>'
          +'<td style="padding:6px 4px;text-align:right;color:#f59e0b">'+mln(r.fix)+'</td>'
          +'<td style="padding:6px 4px;text-align:right;font-weight:800;color:'+(r.op<0?"#ef4444":"#22c55e")+'">'+mln(r.op)+'</td>'
          +'<td style="padding:6px 4px;text-align:right;color:#a78bfa">'+mln(r.bep)+'</td>'
          +'<td style="padding:6px 4px;text-align:right;color:'+(zap<0?"#ef4444":"#22c55e")+'">'+pc(zap,0)+'</td>'
          +'<td style="padding:6px 4px;text-align:right;color:'+(r.ret>50000?"#fb923c":"#475569")+'">'+(r.ret>50000?(mln(r.ret)+" · "+pc(r.rets,1)):"—")+'</td></tr>';
      });
      document.getElementById("fc-prof-tbl").innerHTML=h+'</table>';
    }

    function verdict(r, ctx){
      var b=function(t){ return '<b style="color:#f1f5f9">'+t+'</b>'; };
      var out=[];
      var facCmr=ctx.facCmr, need=r.rev?r.fix/r.rev*100:0;
      if(r.op>0){
        out.push("Окупает свою долю завода. Маржинальная прибыль "+b(mln(r.cm))+" против доли постоянных затрат "
          +b(mln(r.fix))+" — сверху остаётся "+b(mln(r.op))+". Выручка может упасть на "+b(mln(r.rev-r.bep))
          +" ("+pc((r.rev-r.bep)/r.rev*100)+"), прежде чем он уйдёт в ноль.");
      } else {
        out.push("Не окупает свою долю завода: разрыв "+b(mln(-r.op))+". Маржинальная прибыль "+mln(r.cm)
          +" не покрывает долю постоянных затрат "+mln(r.fix)+". Порог безубыточности для него — "
          +b(mln(r.bep))+" выручки, фактически "+mln(r.rev)+".");
        var dPrice=-r.op/r.rev*100, dVol=r.bep/r.rev-1;
        out.push("Три способа закрыть разрыв: поднять цену на "+b(pc(dPrice))+" при том же объёме; вырастить объём на "
          +b(pc(dVol*100))+" при той же марже; или поднять маржинальность с "+pc(r.cmr)+" до "+b(pc(need))
          +" — это "+pp(need-r.cmr)+" за счёт набора товара или закупа.");
      }
      var dm=r.cmr-facCmr;
      if(Math.abs(dm)>=1){
        out.push("Маржинальность "+b(pc(r.cmr))+" против "+pc(facCmr)+" по заводу — "+pp(dm)
          +(dm<0?". Он берёт товар с более дешёвой наценкой или по более низкой цене, поэтому каждый тенге его выручки доходит до покрытия постоянных затрат хуже среднего."
                :". Его набор товара выгоднее среднего по заводу — каждый тенге выручки работает лучше."));
      }
      if(r.ret>50000){
        var rc=r.ret*ctx.prodR;
        out.push("Возвраты "+b(mln(r.ret))+" — "+b(pc(r.rets))+" от отгрузки. В выручке они уже вычтены, но произвести и "
          +"привезти этот объём завод оплатил: около "+b(mln(rc))+" осело в затратах"
          +(r.op<0&&rc>=-r.op*0.5?" — это больше половины его разрыва, то есть возвраты и есть главная причина убытка.":"."));
      }
      var shr=r.rev/ctx.trev*100;
      if(shr>=20) out.push("Доля в выручке "+b(pc(shr))+" — на нём держится значимая часть загрузки. Решения по цене здесь двигают результат завода сильнее всего, но и риск потери объёма самый дорогой.");
      else if(shr<2) out.push("Доля в выручке "+pc(shr)+" — мелкий канал. Даже полное исправление его экономики меняет результат завода не более чем на "+b(mln(Math.abs(r.op)))+", поэтому решать его стоит пакетом с такими же мелкими, а не по отдельности.");
      return out;
    }

    function profCards(rows){
      var trev=rows.reduce(function(s,r){ return s+r.rev; },0);
      // Завод сравниваем по ТЕМ ЖЕ месяцам, что есть у покупателей, иначе
      // средняя маржинальность берётся за другой период и сравнение врёт.
      var set={}; rows.forEach(function(r){ r.mm.forEach(function(m){ set[m]=1; }); });
      var ms=Object.keys(set).sort();
      var tcm=rows.reduce(function(s,r){ return s+r.cm; },0);
      var ctx={trev:trev, facCmr:trev?tcm/trev*100:0, prodR:prodRatio(ms)};
      var a=psort(rows);
      var h="";
      a.forEach(function(r,i){
        var ok=r.op>0;
        h+='<details style="background:#0f172a;border:1px solid '+(ok?"rgba(34,197,94,.28)":"rgba(239,68,68,.28)")
          +';border-radius:12px;padding:0;margin-bottom:8px">'
          +'<summary style="cursor:pointer;list-style:none;padding:11px 14px;display:flex;flex-wrap:wrap;gap:10px;align-items:center">'
          +'<span style="font-size:13px;font-weight:800;color:#f1f5f9">'+esc(r.n)+'</span>'
          +'<span style="font-size:11px;font-weight:700;padding:2px 9px;border-radius:999px;background:'
          +(ok?"rgba(34,197,94,.14);color:#7ff0c0":"rgba(239,68,68,.14);color:#fda4b4")+'">'
          +(ok?"окупает себя":"не окупает")+'</span>'
          +(r.ret>50000?'<span style="font-size:11px;font-weight:700;padding:2px 9px;border-radius:999px;background:rgba(251,146,60,.14);color:#fdba74">возвраты '+pc(r.rets)+'</span>':'')
          +'<span style="margin-left:auto;font-size:12px;color:#94a3b8">выручка '+mln(r.rev)+' · маржа '+pc(r.cmr)
          +' · результат <b style="color:'+(ok?"#22c55e":"#ef4444")+'">'+mln(r.op)+'</b></span></summary>'
          +'<div style="padding:2px 14px 13px">'
          + ((r.members&&r.members.length>1)
              ? '<div style="font-size:11.5px;color:#94a3b8;line-height:1.6;margin:0 0 8px">В группе '+r.members.length+' контрагента: '
                + r.members.slice().sort(function(x,y){return y.rev-x.rev;}).map(function(m){
                    return esc(m.n)+' — '+mln(m.rev)+' выручки, результат <b style="color:'+(m.op<0?"#ef4444":"#22c55e")+'">'+mlnS(m.op)+'</b>'; }).join('; ')
                + '.</div>' : '')
          + verdict(r,ctx).map(function(t){ return '<p style="margin:0 0 7px;font-size:12.5px;line-height:1.65;color:#cbd5e1">'+t+'</p>'; }).join("")
          + scenarioBlock(r)
          + itemsTable(r)
          +'</div></details>';
      });
      document.getElementById("fc-prof-cards").innerHTML=h;
    }

    function profNote(rows){
      var neg=rows.filter(function(r){ return r.op<=0; }).sort(function(x,y){ return x.op-y.op; });
      var el=document.getElementById("fc-prof-note"); if(!el) return;
      if(!neg.length){ el.innerHTML=""; return; }
      var lost=neg.reduce(function(s,r){ return s+r.op; },0);
      var top3=neg.slice(0,3);
      el.innerHTML='<div style="background:rgba(239,68,68,.09);border:1px solid rgba(239,68,68,.3);border-radius:12px;'
        +'padding:11px 14px;font-size:12.5px;color:#cbd5e1;line-height:1.65">'
        +'<b style="color:#f1f5f9">Где сидит убыток.</b> '+neg.length+' контрагентов не окупают свою долю постоянных затрат, вместе это '
        +'<b style="color:#f1f5f9">'+mln(lost)+'</b>. Три главных: '
        + top3.map(function(r){ return esc(r.n)+" ("+mln(r.op)+")"; }).join(", ")+'.'
        +'<div style="color:#94a3b8;font-size:11.5px;margin-top:5px">Важно: доля постоянных затрат — не «его» расходы. '
        +'Если контрагент уйдёт, эти затраты останутся на заводе и лягут на остальных. Поэтому «не окупает» читается как '
        +'«его цена и объём не выдерживают своей доли завода», а не как «от него надо отказаться».</div></div>';
    }

    function profitability(){
      if(!D.buyers||!(D.border||[]).length){ var c=document.getElementById("fc-prof-card"); if(c) c.style.display="none"; return; }
      var rows=groupRows(buyerRows());
      if(!rows.length){ var c2=document.getElementById("fc-prof-card"); if(c2) c2.style.display="none"; return; }
      var cc=document.getElementById("fc-prof-card"); if(cc) cc.style.display="";
      var mlbl=st.month?(MN[+st.month.slice(5)]+" "+st.month.slice(0,4)):"месяц";
      seg("fc-prof-per",[["all","весь период"],["l3","последние 3 мес"],["m",mlbl]],PPER,
          function(v){ PPER=v; profitability(); });
      seg("fc-prof-grp",[["grp","группы по номерам"],["all","каждый контрагент"]],PGRP,
          function(v){ PGRP=v; profitability(); });
      seg("fc-prof-base",[["prod","постоянные: по производству"],["rev","по выручке"]],PBASE,
          function(v){ PBASE=v; profitability(); });
      seg("fc-prof-sort",[["op","по результату"],["rev","по выручке"],["cmr","по маржинальности"],["ret","по возвратам"]],
          PSORT,function(v){ PSORT=v; profitability(); });
      baseNote(rows); profKpi(rows); profSum(rows); profTable(rows); unitTable(rows); profNote(rows); profCards(rows);
      if(window.Chart){ ch7(rows); ch6(rows); }
      cutBlock();
    }


    /* ── Кого можно отключить ───────────────────────────────────────────────
       Ключевая развилка всей вкладки. Строка «результат» считается ПОСЛЕ
       разнесения постоянных затрат — но при отключении покупателя эти затраты
       никуда не денутся, они просто лягут на остальных. Поэтому решение
       принимается по маржинальной прибыли: пока она положительная, покупатель
       оплачивает часть общезаводских затрат, и отключать его — значит терять
       ровно этот вклад. */
    var CUTSEL={}, CUTFIX=0, CUTMOVE=0;
    function cutRows(){ return psort(groupRows(buyerRows())); }
    function cutVerdict(r){
      if(r.cm<=0) return ["cut","отключать или срочно поднимать цену"];
      if(r.op<0)  return ["fix","держать, но пересматривать цену"];
      return ["keep","держать"];
    }
    function cutRule(rows){
      var el=document.getElementById("fc-cut-rule"); if(!el) return;
      var neg=rows.filter(function(r){ return r.cm<=0; });
      var tot=rows.reduce(function(s2,r){ return s2+r.cm; },0);
      var loss=rows.filter(function(r){ return r.op<0; });
      var lossCm=loss.reduce(function(s2,r){ return s2+r.cm; },0);
      el.innerHTML='<div style="background:rgba(239,68,68,.09);border:1px solid rgba(239,68,68,.32);border-radius:12px;'
        +'padding:11px 14px;font-size:12.5px;color:#cbd5e1;line-height:1.7">'
        +'<b style="color:#f1f5f9">Главное, прежде чем кого-то отключать.</b> В таблице выше '+loss.length+' контрагентов с минусом — '
        +'но этот минус получился ПОСЛЕ того, как на них разнесли общезаводские постоянные затраты. '
        +'Если их отключить, затраты останутся на заводе, а вместе с покупателями уйдёт их маржинальная прибыль '
        +'<b style="color:#f1f5f9">'+mln(lossCm)+'</b> — и результат завода станет хуже, а не лучше.'
        +'<div style="margin-top:6px">Правило простое: <b style="color:#7ff0c0">пока маржинальная прибыль положительная</b> — покупатель '
        +'оплачивает часть постоянных затрат, и держать его выгоднее, чем не иметь вовсе. '
        +'<b style="color:#fda4b4">Отключать имеет смысл только тех, у кого она отрицательная</b> — там каждая отгрузка увеличивает убыток.</div>'
        +'<div style="margin-top:6px;color:#94a3b8;font-size:11.5px">Сейчас с отрицательной маржинальной прибылью: '
        + (neg.length? '<b style="color:#fda4b4">'+neg.map(function(r){ return esc(r.n)+" ("+mlnS(r.cm)+")"; }).join(", ")+'</b>. Это весь список — '
             +'вместе они дают '+mlnS(neg.reduce(function(s2,r){return s2+r.cm;},0))+', то есть отключение вообще никого не спасёт.'
          : 'таких нет. Отключение любого покупателя ухудшит результат завода.')
        +' Общий вклад всех контрагентов в покрытие постоянных — <b style="color:#f1f5f9">'+mln(tot)+'</b>.</div>'
        +'<div style="margin-top:6px;color:#94a3b8;font-size:11.5px">Отключение оправдано в трёх случаях: (1) отрицательный вклад; '
        +'(2) вместе с покупателем реально уходит затрата — маршрут доставки, смена, склад; '
        +'(3) цех загружен под завязку и его объём можно заменить более выгодным. Для третьего случая смотрите колонку '
        +'«вклад на 1 ₸ загрузки цеха» — она показывает, кто занимает мощность зря.</div></div>';
    }
    function cutTable(rows){
      var el=document.getElementById("fc-cut-tbl"); if(!el) return;
      var a=rows.slice().sort(function(x,y){ return (x.cost?x.cm/x.cost:9)-(y.cost?y.cm/y.cost:9); });
      var h='<table style="width:100%;border-collapse:collapse;font-size:12px;min-width:880px">'
        +'<tr style="color:#64748b;font-size:10.5px;font-weight:800;text-align:right;text-transform:uppercase;letter-spacing:.04em">'
        +'<th style="width:26px"></th><th style="text-align:left;padding:6px 4px">Контрагент</th><th style="padding:6px 4px">Выручка</th>'
        +'<th style="padding:6px 4px">Маржин. прибыль</th><th style="padding:6px 4px">Вклад на 1 ₸ загрузки цеха</th>'
        +'<th style="padding:6px 4px">Если отключить</th><th style="text-align:left;padding:6px 4px">Что делать</th></tr>';
      a.forEach(function(r){
        var v=cutVerdict(r), dens=r.cost?r.cm/r.cost:0;
        var col=v[0]==="cut"?"#fda4b4":(v[0]==="fix"?"#fdba74":"#7ff0c0");
        var bg=v[0]==="cut"?"rgba(239,68,68,.14)":(v[0]==="fix"?"rgba(251,146,60,.13)":"rgba(34,197,94,.13)");
        h+='<tr style="border-top:1px solid #1b2636">'
          +'<td style="padding:6px 2px;text-align:center"><input type="checkbox" data-cut="'+esc(r.n)+'"'+(CUTSEL[r.n]?" checked":"")+'></td>'
          +'<td style="padding:6px 4px;color:#e2e8f0;max-width:220px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">'+esc(r.n)+'</td>'
          +'<td style="padding:6px 4px;text-align:right;color:#cbd5e1">'+mln(r.rev)+'</td>'
          +'<td style="padding:6px 4px;text-align:right;font-weight:700;color:'+(r.cm>0?"#22d3ee":"#ef4444")+'">'+mlnS(r.cm)+'</td>'
          +'<td style="padding:6px 4px;text-align:right;color:'+(dens<0.5?"#fb923c":"#94a3b8")+'">'+dens.toFixed(2).replace(".",",")+'</td>'
          +'<td style="padding:6px 4px;text-align:right;font-weight:700;color:'+(r.cm>0?"#ef4444":"#22c55e")+'">'+mlnS(-r.cm)+'</td>'
          +'<td style="padding:6px 4px"><span style="font-size:11px;font-weight:700;padding:2px 9px;border-radius:999px;background:'+bg+';color:'+col+'">'+v[1]+'</span></td></tr>';
      });
      el.innerHTML=h+'</table>';
      [].forEach.call(el.querySelectorAll("input[data-cut]"),function(inp){
        inp.onchange=function(){ var k=inp.getAttribute("data-cut");
          if(inp.checked) CUTSEL[k]=1; else delete CUTSEL[k];
          cutSim(cutRows()); };
      });
    }
    function cutSim(rows){
      var el=document.getElementById("fc-cut-sim"); if(!el) return;
      var sel=rows.filter(function(r){ return CUTSEL[r.n]; });
      var rest=rows.filter(function(r){ return !CUTSEL[r.n]; });
      var totCm=rows.reduce(function(s2,r){ return s2+r.cm; },0);
      var totFix=rows.reduce(function(s2,r){ return s2+r.fix; },0);
      var totRev=rows.reduce(function(s2,r){ return s2+r.rev; },0);
      var base=totCm-totFix;
      var loseCm=sel.reduce(function(s2,r){ return s2+r.cm; },0);
      var freeFix=sel.reduce(function(s2,r){ return s2+r.fix; },0)*CUTFIX;
      var moveRev=sel.reduce(function(s2,r){ return s2+r.rev; },0)*CUTMOVE;
      var restRev=rest.reduce(function(s2,r){ return s2+r.rev; },0);
      var restCm=rest.reduce(function(s2,r){ return s2+r.cm; },0);
      var restCmr=restRev?restCm/restRev:0;
      var gainMove=moveRev*restCmr;
      var after=base-loseCm+freeFix+gainMove;
      var d=after-base;
      var beFix=freeFix+gainMove>0?0:0;
      var needFix=sel.length?(loseCm-gainMove)/Math.max(1,sel.reduce(function(s2,r){ return s2+r.fix; },0)):0;
      var slider=function(id,val,label,hint){
        return '<div style="flex:1;min-width:230px"><div style="font-size:11.5px;color:#94a3b8;margin-bottom:3px">'+label
          +' <b style="color:#f1f5f9">'+Math.round(val*100)+'%</b></div>'
          +'<input type="range" id="'+id+'" min="0" max="100" step="5" value="'+Math.round(val*100)+'" style="width:100%">'
          +'<div style="font-size:10.5px;color:#64748b;margin-top:2px">'+hint+'</div></div>';
      };
      var h='<div style="background:#0f172a;border:1px solid #1f2937;border-radius:12px;padding:12px 14px">'
        +'<div style="display:flex;gap:18px;flex-wrap:wrap;margin-bottom:12px">'
        + slider("fc-cut-f",CUTFIX,"Какую долю его постоянных реально удастся срезать",
                 "по умолчанию 0: аренда, ФОТ и амортизация от ухода покупателя не исчезают")
        + slider("fc-cut-m",CUTMOVE,"Какая доля его объёма перейдёт к оставшимся",
                 "по умолчанию 0: свободные мощности сами по себе выручку не приносят")
        +'</div>';
      if(!sel.length){
        h+='<div style="font-size:12.5px;color:#94a3b8">Отметьте галочками тех, кого рассматриваете к отключению — покажу, что станет с результатом завода.</div>';
      } else {
        var rowsHtml=[
          ["Результат завода сейчас",mln(base),"#e2e8f0"],
          ["Уходит маржинальная прибыль","−"+mln(loseCm).replace("−",""),"#ef4444"],
          ["Экономия постоянных затрат",(freeFix?("+"+mln(freeFix)):"0"),freeFix?"#22c55e":"#64748b"],
          ["Объём перешёл к оставшимся",(gainMove?("+"+mln(gainMove)):"0"),gainMove?"#22c55e":"#64748b"],
          ["Результат завода после",mln(after),after>base?"#22c55e":"#ef4444"]
        ];
        h+='<table style="width:100%;border-collapse:collapse;font-size:12.5px">';
        rowsHtml.forEach(function(x,i){
          h+='<tr style="'+(i===4?"border-top:1px solid #334155":"border-top:1px solid #1b2636")+'">'
            +'<td style="padding:6px 4px;color:#94a3b8">'+x[0]+'</td>'
            +'<td style="padding:6px 4px;text-align:right;font-weight:'+(i===0||i===4?"800":"700")+';color:'+x[2]+'">'+x[1]+'</td></tr>';
        });
        h+='</table>';
        h+='<div style="margin-top:10px;background:'+(d>=0?"rgba(34,197,94,.10)":"rgba(239,68,68,.10)")
          +';border:1px solid '+(d>=0?"rgba(34,197,94,.32)":"rgba(239,68,68,.32)")+';border-radius:10px;padding:10px 13px;'
          +'font-size:12.5px;color:#cbd5e1;line-height:1.65">'
          +'<b style="color:#f1f5f9">Итог: результат завода '+(d>=0?"улучшится":"ухудшится")+' на '+mln(Math.abs(d))+'.</b> '
          + (d>=0
              ? "При заданных условиях отключение оправдано. Проверьте только, что срезать постоянные вы действительно сможете — это решение, а не следствие."
              : ("Отключение при этих условиях невыгодно. Чтобы выйти хотя бы в ноль, нужно либо срезать "
                 + pc(Math.max(0,Math.min(1,needFix))*100,0) + " их доли постоянных затрат, либо передать оставшимся "
                 + pc(Math.max(0,Math.min(1,(loseCm-freeFix)/Math.max(1,sel.reduce(function(s2,r){return s2+r.rev;},0)*restCmr)))*100,0)
                 + " их объёма."))
          +'</div>';
      }
      h+='</div>';
      el.innerHTML=h;
      var f=document.getElementById("fc-cut-f"), m=document.getElementById("fc-cut-m");
      if(f) f.oninput=function(){ CUTFIX=(+f.value)/100; cutSim(rows); };
      if(m) m.oninput=function(){ CUTMOVE=(+m.value)/100; cutSim(rows); };
    }

    var PPER="all";
    function perMonths(){
      var ms=months();
      if(PPER==="l3") return ms.slice(-3);
      if(PPER==="m") return ms.filter(function(m){ return m===st.month; });
      return ms;
    }

    /* Позиции, которые не окупают даже переменные затраты.
       Переменные на единицу = продуктовая себестоимость × k, где
       k = все переменные затраты ОПиУ / продуктовая себестоимость. Так учитываются
       логистика, электроэнергия, расходники и потери, а не только сырьё. */
    function killRows(){
      var ms=months(), rev=0,varc=0,food=0,fix=0;
      ms.forEach(function(m){ var p=D.pl[m]; if(!p) return;
        rev+=p.rev; varc+=p.var; fix+=p.fix; food+=(p.layers||{}).food||0; });
      if(!food||!rev) return null;
      var k=varc/food, need=fix/rev*100;
      var rows=groupRows(buyerRows()), neg=[], low=0, lowRev=0, lowGap=0;
      rows.forEach(function(r){
        (r.items||[]).forEach(function(it){
          if(it.rest||!it.q||!it.r||it.c==null) return;
          var p=it.r/it.q, vu=(it.c/it.q)*k, cm=(p-vu)*it.q, mr=p?(p-vu)/p*100:0;
          if(cm<0) neg.push({b:r.n,n:it.n,q:it.q,r:it.r,p:p,vu:vu,cm:cm,mr:mr});
          else if(mr<need){ low++; lowRev+=it.r; lowGap+=(need/100*it.r)-cm; }
        });
      });
      neg.sort(function(a,b2){ return a.cm-b2.cm; });
      return {k:k,need:need,neg:neg,low:low,lowRev:lowRev,lowGap:lowGap};
    }
    function killBlock(){
      var el=document.getElementById("fc-kill"); if(!el) return;
      var K=killRows();
      if(!K){ el.innerHTML=""; return; }
      var totCm=K.neg.reduce(function(s2,r){ return s2+r.cm; },0);
      var totRev=K.neg.reduce(function(s2,r){ return s2+r.r; },0);
      var h='<div style="background:rgba(239,68,68,.09);border:1px solid rgba(239,68,68,.32);border-radius:12px;padding:11px 14px;'
        +'font-size:12.5px;color:#cbd5e1;line-height:1.7;margin-bottom:10px">'
        +'<b style="color:#f1f5f9">Вот это и есть «убрать невыгодные продажи».</b> Нашлось <b style="color:#f1f5f9">'+K.neg.length+'</b> связок '
        +'покупатель × товар, где цена ниже переменных затрат: выручка '+mln(totRev)+', вклад <b style="color:#fda4b4">'+mlnS(totCm)+'</b>. '
        +'Убрать их из матрицы или поднять по ним цену — единственное сокращение продаж, которое улучшает результат, а не ухудшает.'
        +'<div style="margin-top:5px;color:#94a3b8;font-size:11.5px">Переменные на единицу = продуктовая себестоимость × '
        +K.k.toFixed(2).replace(".",",")+' — коэффициент переводит сырьё в полные переменные затраты (логистика, электроэнергия, расходники, потери). '
        +'Ещё <b>'+K.low+'</b> связок на '+mln(K.lowRev)+' выручки дают положительный вклад, но ниже порога '+pc(K.need)+', '
        +'который нужен, чтобы нести свою долю постоянных: недобор '+mln(K.lowGap)+'. Их убирать нельзя — только поднимать цену.</div></div>';
      if(!K.neg.length){ el.innerHTML=h; return; }
      h+='<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:12px;min-width:840px">'
        +'<tr style="color:#64748b;font-size:10.5px;font-weight:800;text-align:right;text-transform:uppercase;letter-spacing:.04em">'
        +'<th style="text-align:left;padding:6px 4px">Покупатель</th><th style="text-align:left;padding:6px 4px">Позиция</th>'
        +'<th style="padding:6px 4px">Штук</th><th style="padding:6px 4px">Цена</th><th style="padding:6px 4px">Переменные на ед.</th>'
        +'<th style="padding:6px 4px">Теряем на штуке</th><th style="padding:6px 4px">Всего</th>'
        +'<th style="padding:6px 4px">Минимальная цена</th></tr>';
      K.neg.slice(0,25).forEach(function(r){
        var minp=r.vu/(1-K.need/100);
        h+='<tr style="border-top:1px solid #1b2636">'
          +'<td style="padding:6px 4px;color:#cbd5e1;max-width:170px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">'+esc(r.b)+'</td>'
          +'<td style="padding:6px 4px;color:#e2e8f0;max-width:250px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">'+esc(r.n)+'</td>'
          +'<td style="padding:6px 4px;text-align:right;color:#94a3b8">'+num(r.q)+'</td>'
          +'<td style="padding:6px 4px;text-align:right;color:#e2e8f0">'+num(r.p)+' ₸</td>'
          +'<td style="padding:6px 4px;text-align:right;color:#fb923c">'+num(r.vu)+' ₸</td>'
          +'<td style="padding:6px 4px;text-align:right;color:#ef4444;font-weight:700">−'+num(r.vu-r.p)+' ₸</td>'
          +'<td style="padding:6px 4px;text-align:right;color:#ef4444;font-weight:800">'+mlnS(r.cm)+'</td>'
          +'<td style="padding:6px 4px;text-align:right;color:#7ff0c0">'+num(minp)+' ₸</td></tr>';
      });
      h+='</table></div>';
      if(K.neg.length>25) h+='<div style="font-size:11.5px;color:#64748b;margin-top:6px">Показаны 25 самых дорогих из '+K.neg.length+'.</div>';
      h+='<div style="font-size:11.5px;color:#94a3b8;margin-top:8px;line-height:1.6">«Минимальная цена» — при которой позиция покрывает '
        +'переменные затраты и свою долю постоянных ('+pc(K.need)+' маржинальности). Ниже неё отгружать нет смысла ни при каком объёме.</div>';
      el.innerHTML=h;
    }

    function cutBlock(){
      var card=document.getElementById("fc-cut-card"); if(!card) return;
      var rows=cutRows();
      if(!rows.length){ card.style.display="none"; return; }
      card.style.display="";
      cutRule(rows); cutTable(rows); cutSim(rows); killBlock();
    }

    function momTable(){
      var ms=months();
      var h='<table style="width:100%;border-collapse:collapse;font-size:12.5px;min-width:820px">'
        +'<tr style="color:#64748b;font-size:10.5px;font-weight:800;text-align:right;text-transform:uppercase;letter-spacing:.04em">'
        +'<th style="text-align:left;padding:6px 4px">Месяц</th><th style="padding:6px 4px">Выручка</th><th style="padding:6px 4px">Δ выручки</th>'
        +'<th style="padding:6px 4px">Маржин.</th><th style="padding:6px 4px">Δ маржин.</th><th style="padding:6px 4px">Постоянные</th>'
        +'<th style="padding:6px 4px">Δ постоян.</th><th style="padding:6px 4px">Результат</th><th style="padding:6px 4px">Δ результата</th><th style="padding:6px 4px">Запас</th></tr>';
      ms.forEach(function(m,i){
        var p=PL()[m], base=(st.cmp==="yoy"?(p.yoy?p.yoy.prev:null):(MONTHS()[MONTHS().indexOf(m)-1]||null));
        var b=base?PL()[base]:null;
        var dr=b?(p.rev-b.rev)/b.rev*100:null, dm=b?(p.cmr-b.cmr):null, df=b?(p.fix-b.fix)/b.fix*100:null, dop=b?(p.op-b.op):null;
        function cell(v,txt,inv){ var c=v==null?"#64748b":((inv?-v:v)>0?"#22c55e":((inv?-v:v)<0?"#ef4444":"#94a3b8"));
          return '<td style="padding:6px 4px;text-align:right;white-space:nowrap;color:'+c+';font-weight:700">'+txt+'</td>'; }
        h+='<tr data-m="'+m+'" style="border-top:1px solid #1b2636;cursor:pointer;background:'+(m===st.month?"rgba(201,169,78,.09)":"transparent")+'">'
          +'<td style="padding:6px 4px;color:#e2e8f0;font-weight:700">'+MS[+m.slice(5)]+" "+m.slice(0,4)+'</td>'
          +'<td style="padding:6px 4px;text-align:right;color:#cbd5e1">'+mln(p.rev)+'</td>'
          +cell(dr,dr==null?"—":pc(dr))
          +'<td style="padding:6px 4px;text-align:right;color:#22d3ee;font-weight:700">'+pc(p.cmr)+'</td>'
          +cell(dm,dm==null?"—":((dm>0?"+":"")+dm.toFixed(1).replace(".",",")+" пп"))
          +'<td style="padding:6px 4px;text-align:right;color:#f59e0b">'+mln(p.fix)+'</td>'
          +cell(df,df==null?"—":pc(df),true)
          +'<td style="padding:6px 4px;text-align:right;font-weight:800;color:'+(p.op<0?"#ef4444":"#22c55e")+'">'+mln(p.op)+'</td>'
          +cell(dop,dop==null?"—":sg(dop))
          +'<td style="padding:6px 4px;text-align:right;color:'+(p.safety<0?"#ef4444":"#22c55e")+'">'+pc(p.safety)+'</td></tr>';
      });
      var el=document.getElementById("fc-mom"); el.innerHTML=h+'</table>';
      el.onclick=function(e){ var tr=e.target.closest("tr[data-m]"); if(!tr) return; st.month=tr.getAttribute("data-m"); render(); };
    }

    function linesTable(){
      var m=st.month, base=(st.cmp==="yoy"?(PL()[m].yoy?PL()[m].yoy.prev:null):(MONTHS()[MONTHS().indexOf(m)-1]||null));
      var rows=LINES().map(function(l){ var cur=l.m[m]||0, prv=base?(l.m[base]||0):0;
        return {n:l.n,g:l.g,cur:cur,prv:prv,d:cur-prv,dp:prv?((cur-prv)/Math.abs(prv)*100):null}; })
        .filter(function(r){ return Math.abs(r.cur)>300000||Math.abs(r.d)>300000; });
      rows.sort(function(a,b){ return Math.abs(b.d)-Math.abs(a.d); });
      var note="";
      if(!rows.length){
        note='<div style="font-size:11px;color:#c4b5fd;background:rgba(167,139,250,.1);border:1px solid rgba(167,139,250,.35);'
           +'border-radius:8px;padding:6px 9px;margin-bottom:7px">&#9432; Постатейной расшифровки ОПиУ за этот месяц ещё нет &mdash; '
           +'показаны шесть групп затрат из iiko. Детализация появится после закрытия месяца.</div>';
        rows=D.layers.map(function(l){
          var cur=(PL()[m].layers||{})[l.k]||0, prv=base?((PL()[base].layers||{})[l.k]||0):0;
          return {n:l.t,g:"группа",cur:cur,prv:prv,d:cur-prv,dp:prv?((cur-prv)/Math.abs(prv)*100):null};
        }).filter(function(r){ return r.cur||r.d; });
        rows.sort(function(a,b){ return Math.abs(b.d)-Math.abs(a.d); });
      }
      var h=note+'<div style="font-size:11.5px;color:#64748b;margin-bottom:6px">'+MN[+m.slice(5)]+" "+m.slice(0,4)+(base?(" против "+MN[+base.slice(5)]+" "+base.slice(0,4)):"")+'</div>'
        +'<table style="width:100%;border-collapse:collapse;font-size:12px;min-width:380px">';
      rows.slice(0,14).forEach(function(r){
        h+='<tr style="border-top:1px solid #1b2636">'
          +'<td style="padding:5px 4px;color:#cbd5e1;max-width:210px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">'+esc(r.n.replace(/^\d+(\.\d+)*\./,""))+' <span style="color:#475569;font-size:10px">'+r.g+'</span></td>'
          +'<td style="padding:5px 4px;text-align:right;color:#94a3b8;white-space:nowrap">'+mln(r.cur)+'</td>'
          +'<td style="padding:5px 4px;text-align:right;font-weight:700;white-space:nowrap;color:'+(r.d>0?"#ef4444":"#22c55e")+'">'+sg(r.d)+'</td></tr>';
      });
      document.getElementById("fc-lines").innerHTML=h+'</table>';
    }

    function chanTable(){
      var tt=document.getElementById("fc-chan-title");
      if(BO()){
        if(tt) tt.innerHTML="&#128100; "+esc(BUYER)+" по месяцам";
        var bm=months(), bp=PL();
        var h2='<div style="font-size:11.5px;color:#64748b;margin:6px 0">выручка, доля завода и результат с учётом разнесённых затрат</div>'
          +'<table style="width:100%;border-collapse:collapse;font-size:12px;min-width:380px">'
          +'<tr style="color:#64748b;font-size:10.5px;font-weight:800;text-transform:uppercase;letter-spacing:.04em">'
          +'<th style="text-align:left;padding:5px 4px">Месяц</th><th style="text-align:right;padding:5px 4px">Выручка</th>'
          +'<th style="text-align:right;padding:5px 4px">Доля</th><th style="text-align:right;padding:5px 4px">Маржин.</th>'
          +'<th style="text-align:right;padding:5px 4px">Результат</th></tr>';
        bm.forEach(function(m){ var p=bp[m];
          h2+='<tr style="border-top:1px solid #1b2636">'
            +'<td style="padding:5px 4px;color:#cbd5e1">'+MS[+m.slice(5)]+" "+m.slice(0,4)+'</td>'
            +'<td style="padding:5px 4px;text-align:right;color:#94a3b8">'+mln(p.rev)+'</td>'
            +'<td style="padding:5px 4px;text-align:right;color:#64748b">'+pc(SHARE(m)*100,0)+'</td>'
            +'<td style="padding:5px 4px;text-align:right;color:#22d3ee;font-weight:700">'+pc(p.cmr)+'</td>'
            +'<td style="padding:5px 4px;text-align:right;font-weight:700;color:'+(p.op<0?"#ef4444":"#22c55e")+'">'+mln(p.op)+'</td></tr>';
        });
        document.getElementById("fc-chan").innerHTML=h2+"</table>";
        return;
      }
      if(tt) tt.innerHTML="&#127978; Каналы продаж";
      var ms=D.cmonths.filter(function(m){ return st.period==="all"||m.indexOf(st.period)===0; });
      var last=ms[ms.length-1], prev=ms[ms.length-2];
      var rows=Object.keys(D.chan).map(function(n){
        var tot=0; ms.forEach(function(m){ tot+=D.chan[n][m]||0; });
        var a=D.chan[n][last]||0, b=prev?(D.chan[n][prev]||0):0;
        return {n:n,t:tot,a:a,d:a-b,dp:b?((a-b)/b*100):null};
      }).filter(function(r){ return r.t>0; });
      rows.sort(function(x,y){ return y.t-x.t; });
      var tot=rows.reduce(function(s,r){return s+r.t;},0);
      var h='<div style="font-size:11.5px;color:#64748b;margin:6px 0">последний месяц в данных — '+MN[+last.slice(5)]+" "+last.slice(0,4)+'</div>'
        +'<table style="width:100%;border-collapse:collapse;font-size:12px;min-width:380px">';
      rows.slice(0,10).forEach(function(r){
        h+='<tr style="border-top:1px solid #1b2636">'
          +'<td style="padding:5px 4px;color:#cbd5e1;max-width:190px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">'+esc(r.n)+'</td>'
          +'<td style="padding:5px 4px;text-align:right;color:#94a3b8">'+mln(r.t)+'</td>'
          +'<td style="padding:5px 4px;text-align:right;color:#64748b">'+pc(r.t/tot*100,0)+'</td>'
          +'<td style="padding:5px 4px;text-align:right;font-weight:700;white-space:nowrap;color:'+(r.d>0?"#22c55e":(r.d<0?"#ef4444":"#64748b"))+'">'+sg(r.d)+'</td></tr>';
      });
      document.getElementById("fc-chan").innerHTML=h+'</table>';
    }
    __MODAL__

    function render(){
      var yrsA=[]; MONTHS().forEach(function(m){ var y=m.slice(0,4); if(yrsA.indexOf(y)<0) yrsA.push(y); });
      seg("fc-period",[["all","Весь период"]].concat(yrsA.map(function(y){ return [y,y]; })),st.period,function(v){ st.period=v;
        var ms=months(); if(ms.indexOf(st.month)<0) st.month=ms[ms.length-1]; render(); });
      seg("fc-mode",[["abs","₸"],["pct","% от выручки"]],st.mode,function(v){ st.mode=v; render(); });
      seg("fc-cmp",BO()?[["mom","к прошлому месяцу"]]:[["mom","к прошлому месяцу"],["yoy","к прошлому году"]],st.cmp,function(v){ st.cmp=v; render(); });
      var bs=document.getElementById("fc-buyer");
      if(bs){
        bs.innerHTML='<option value="">&#127981; Весь завод</option>'+((D.border||[]).map(function(n){
          return '<option value="'+esc(n)+'"'+(n===BUYER?" selected":"")+'>'+esc(n)+'</option>'; }).join(""));
        bs.onchange=function(){
          BUYER=this.value; var ms=MONTHS();
          if(st.period!=="all" && !ms.filter(function(m){return m.indexOf(st.period)===0;}).length) st.period="all";
          if(ms.indexOf(st.month)<0) st.month=ms[ms.length-1];
          if(BUYER && st.cmp==="yoy") st.cmp="mom";
          render();
        };
      }
      var sel=document.getElementById("fc-month");
      sel.innerHTML=months().map(function(m){ var p=PL()[m];
        return '<option value="'+m+'"'+(m===st.month?" selected":"")+'>'+MN[+m.slice(5)]+" "+m.slice(0,4)
          +(p.est?(" · iiko, "+gapNames(p)+" оценкой"):(p.src==="iiko"?" · iiko":""))+'</option>'; }).join("");
      sel.onchange=function(){ st.month=this.value; render(); };
      gapNote(); bnote(); kpi(); alertBox(); profitability(); momTable(); linesTable(); chanTable(); observations();
      if(window.Chart){ ch1(); ch2(); ch3(); ch4(); ch5(); }
    }

    function boot(){
      var a=agg(MONTHS());
      document.getElementById("fc-sum").textContent="выручка "+mln(a.rev)+" · результат "+mln(a.op)+" · "+MONTHS().length+" мес.";
      var det=document.getElementById("fc-details");
      det.addEventListener("toggle",function(){ var c=document.getElementById("fc-caret"); if(c) c.innerHTML=det.open?"&#9662;":"&#9656;"; if(det.open) render(); });
      document.getElementById("fc-open").onclick=function(){ openModal(); };
      document.getElementById("fc-close").onclick=function(){ document.getElementById("fc-modal").style.display="none"; document.body.style.overflow=""; };
      document.getElementById("fc-modal").onclick=function(e){ if(e.target===this){ this.style.display="none"; document.body.style.overflow=""; } };
      if(det.open) render();
    }
    if(document.readyState==="loading") document.addEventListener("DOMContentLoaded",boot); else boot();
  })();
  </script>
</div>
'''


MODAL_JS = r'''
    function openModal(){
      var ms=months(), a=agg(ms), n=ms.length;
      var y25=YRS()["2025"], y26=YRS()["2026"];
      var cmp=null;
      if(y26){
        var same=y26.months.map(function(m){ return "2025"+m.slice(4); }).filter(function(m){ return PL()[m]; });
        if(same.length===y26.months.length){ cmp={a:agg(same),b:agg(y26.months),ms:same,ms2:y26.months}; }
      }
      var sorted=ms.slice().sort(function(x,y){ return PL()[y].op-PL()[x].op; });
      var best=sorted[0], worst=sorted[sorted.length-1];
      var revS=ms.slice().sort(function(x,y){ return PL()[y].rev-PL()[x].rev; });
      var LOSS=["1.30.Возвраты от дистрибьютеров","1.28.Брак","1.7.Истек срок хранения (порча)","1.3.Недостача инвентаризации","1.24.Бракераж","1.27.Нарушение тех.процесса","1.11.Списание сломанных ТМЗ"];
      function lineSum(name,mm){ var l=null; LINES().forEach(function(x){ if(x.n===name) l=x; }); if(!l) return 0;
        var s=0; mm.forEach(function(m){ s+=l.m[m]||0; }); return s; }
      var lossTot=0; LOSS.forEach(function(k){ lossTot+=lineSum(k,ms); });

      var chanMs=D.cmonths.filter(function(m){ return st.period==="all"||m.indexOf(st.period)===0; });
      var chRows=Object.keys(D.chan).map(function(nm){
        var t=0; chanMs.forEach(function(m){ t+=D.chan[nm][m]||0; });
        var last3=chanMs.slice(-3).reduce(function(s,m){ return s+(D.chan[nm][m]||0); },0)/Math.max(1,Math.min(3,chanMs.length));
        var first3=chanMs.slice(0,3).reduce(function(s,m){ return s+(D.chan[nm][m]||0); },0)/Math.max(1,Math.min(3,chanMs.length));
        return {n:nm,t:t,last:last3,first:first3,d:last3-first3};
      }).sort(function(x,y){ return y.t-x.t; });
      var gone=chRows.filter(function(r){ return r.first>5e6 && r.last<r.first*0.2; }).sort(function(x,y){ return x.d-y.d; });
      var grown=chRows.filter(function(r){ return r.d>3e6; }).sort(function(x,y){ return y.d-x.d; });

      var deficit=-a.op, perMonth=deficit/n;
      var needRev=a.cmr>0?(deficit/(a.cmr/100)):0;

      function H(t){ return '<div style="font-size:11px;font-weight:800;color:#c9a94e;letter-spacing:.09em;text-transform:uppercase;margin:22px 0 8px;padding-bottom:6px;border-bottom:1px solid #1f2937">'+t+'</div>'; }
      function P(t){ return '<p style="margin:0 0 9px;font-size:13px;line-height:1.72;color:#cbd5e1">'+t+'</p>'; }
      function KV(rows){
        return '<table style="width:100%;border-collapse:collapse;font-size:12.5px;margin:4px 0 10px">'+rows.map(function(r){
          return '<tr style="border-bottom:1px solid #16202f"><td style="padding:6px 2px;color:#94a3b8">'+r[0]+'</td>'
            +'<td style="padding:6px 2px;text-align:right;font-weight:700;color:'+(r[2]||"#e2e8f0")+';white-space:nowrap;font-variant-numeric:tabular-nums">'+r[1]+'</td></tr>';
        }).join("")+'</table>';
      }
      function LI(items){ return '<ol style="margin:2px 0 10px;padding-left:20px;font-size:13px;line-height:1.75;color:#cbd5e1">'+items.map(function(t){ return '<li style="margin-bottom:6px">'+t+'</li>'; }).join("")+'</ol>'; }
      function b(t){ return '<b style="color:#f1f5f9">'+t+'</b>'; }

      var h="";
      h+=H("1 · Резюме");
      h+=P("За "+n+" мес. выручка "+b(mln(a.rev))+", полная себестоимость "+b(mln(a.full))+", результат "+b(mln(a.op))+" ("+pc(a.op/a.rev*100)+" к выручке). На каждые 100 ₸ выручки приходится "+b(Math.round(a.full/a.rev*100)+" ₸")+" затрат.");
      h+=KV([["Выручка",mln(a.rev)],["Переменные затраты",mln(a.var_),"#fb923c"],["Маржинальная прибыль",mln(a.cm)+" · "+pc(a.cmr),"#22d3ee"],
             ["Постоянные затраты",mln(a.fix),"#f59e0b"],["Операционный результат",mln(a.op),a.op<0?"#ef4444":"#22c55e"],
             ["Точка безубыточности, в месяц",mln(a.bep/n),"#a78bfa"],["Фактическая выручка, в месяц",mln(a.rev/n)],
             ["Запас прочности",pc(a.safety),a.safety<0?"#ef4444":"#22c55e"]]);

      h+=H("2 · Выручка");
      h+=P("Лучший месяц по выручке — "+b(MN[+revS[0].slice(5)]+" "+revS[0].slice(0,4))+" ("+mln(PL()[revS[0]].rev)+"), худший — "+b(MN[+revS[revS.length-1].slice(5)]+" "+revS[revS.length-1].slice(0,4))+" ("+mln(PL()[revS[revS.length-1]].rev)+"). Разброс "+mln(PL()[revS[0]].rev-PL()[revS[revS.length-1]].rev)+" — это "+pc((PL()[revS[0]].rev/PL()[revS[revS.length-1]].rev-1)*100)+" к минимуму.");
      if(cmp){
        var dr=(cmp.b.rev-cmp.a.rev)/cmp.a.rev*100;
        h+=P("Сопоставимый период "+b(MN[+cmp.ms2[0].slice(5)]+"–"+MN[+cmp.ms2[cmp.ms2.length-1].slice(5)])+": 2026 год дал "+b(mln(cmp.b.rev))+" против "+mln(cmp.a.rev)+" в 2025 — "+b(pc(dr))+" ("+sg(cmp.b.rev-cmp.a.rev)+")."
          +(dr<0?" Провал выручки и есть основная причина убытка."
                :" То есть выручка год к году не упала — значит результат ухудшили не продажи, а маржинальность и постоянные затраты."));
        var h2=MONTHS().filter(function(m){ return m.indexOf("2025")===0 && +m.slice(5)>=7; });
        if(h2.length){
          var a2=agg(h2), am2=a2.rev/h2.length, am26=cmp.b.rev/cmp.ms2.length;
          h+=P("Но если сравнивать со вторым полугодием 2025 ("+b(mln(am2)+" в месяц")+"), то текущий уровень "+b(mln(am26)+" в месяц")+" — это "+b(pc((am26/am2-1)*100))+", то есть "+b(mln(Math.abs(am2-am26))+" выручки в месяц")+", которых не хватает при неизменных постоянных затратах.");
        }
      }
      h+=P("Прибыльных месяцев "+b(ms.filter(function(m){return PL()[m].op>0;}).length+" из "+n)+". Лучший результат — "+b(MN[+best.slice(5)]+" "+best.slice(0,4)+": "+mln(PL()[best].op))+", худший — "+b(MN[+worst.slice(5)]+" "+worst.slice(0,4)+": "+mln(PL()[worst].op))+".");

      h+=H("3 · Маржинальность: сколько остаётся после переменных затрат");
      h+=P("Средняя маржинальность периода — "+b(pc(a.cmr))+". Это доля выручки, которая доходит до покрытия постоянных затрат. Каждый процентный пункт маржинальности стоит "+b(mln(a.rev/n/100))+" в месяц.");
      if(cmp){
        var dm=cmp.b.cmr-cmp.a.cmr;
        h+=P("В сопоставимом периоде маржинальность "+(dm<0?"упала":"выросла")+" с "+b(pc(cmp.a.cmr))+" до "+b(pc(cmp.b.cmr))+" — это "+b((dm>0?"+":"")+dm.toFixed(1).replace(".",",")+" пункта")+", или "+b(sg(cmp.b.rev*dm/100))+" результата на нынешних объёмах.");
      }
      var fl=[]; D.layers.forEach(function(l){ fl.push([l.t,mln(a.layers[l.k])+" · "+pc(a.layers[l.k]/a.rev*100),LC[l.k]]); });
      h+=P("Структура полной себестоимости за период:");
      h+=KV(fl);

      h+=H("4 · Постоянные затраты");
      h+=P("Постоянные затраты — "+b(mln(a.fix))+" за период, в среднем "+b(mln(a.fix/n))+" в месяц. Они не зависят от объёма продаж, поэтому при падении выручки бьют по результату напрямую.");
      if(cmp){
        var nb=cmp.ms2.length, na=cmp.ms.length;
        var df=cmp.b.fix/nb-cmp.a.fix/na;
        h+=P("В сопоставимом периоде постоянные "+(df>0?"выросли":"снизились")+" на "+b(mln(Math.abs(df))+" в месяц")+" ("+pc((cmp.b.fix/nb)/(cmp.a.fix/na)*100-100)+"). Это "+b(mln(df*nb))+" за период.");
      }

      h+=H("5 · Точка безубыточности");
      h+=P("При маржинальности "+b(pc(a.cmr))+" для покрытия постоянных затрат нужна выручка "+b(mln(a.bep/n)+" в месяц")+". Фактическая — "+b(mln(a.rev/n))+". Разрыв "+b(mln(Math.abs(a.bep/n-a.rev/n)))+" в месяц"+(a.op<0?" — это и есть месячный убыток в пересчёте на выручку.":"."));
      var bepRows=ms.slice(-6).map(function(m){ var p=PL()[m];
        return [MN[+m.slice(5)]+" "+m.slice(0,4), mln(p.rev)+" при пороге "+mln(p.bep)+" · "+pc(p.safety), p.safety<0?"#ef4444":"#22c55e"]; });
      h+=KV(bepRows);

      if(BO()){
        h+=H("6 · Покупатель в разрезе завода");
        var fR=0,fO=0; ms.forEach(function(m){ fR+=D.pl[m].rev; fO+=D.pl[m].op; });
        h+=P("Выбран "+b(esc(BUYER))+". Все цифры выше — его доля завода: выручка по расходным накладным iiko, продуктовая себестоимость по его фактическому товарному набору, прочие переменные — пропорционально выручке, постоянные — по доле в производстве.");
        h+=KV([["Доля в выручке завода",pc(a.rev/fR*100)],["Место по обороту",((D.border||[]).indexOf(BUYER)+1)+" из "+(D.border||[]).length],
               ["Маржинальная прибыль",mln(a.cm)+" · "+pc(a.cmr),"#22d3ee"],["Доля постоянных затрат",mln(a.fix),"#f59e0b"],
               ["Результат",mln(a.op),a.op<0?"#ef4444":"#22c55e"]]);
        h+=P(a.cm>a.fix?"Он окупает свою долю общезаводских постоянных затрат.":"Маржинальной прибыли не хватает на его долю постоянных затрат: разрыв "+b(mln(a.fix-a.cm))+". Это не значит, что от него надо отказаться — при его уходе постоянные затраты останутся и лягут на остальных. Значит, нужна либо цена, либо набор с более высокой маржой, либо объём.");
      } else {
      h+=H("6 · Каналы продаж: кто ушёл и кто пришёл");
      if(gone.length){
        h+=P(b("Потерянные каналы.")+" Сравнение первых и последних трёх месяцев данных:");
        h+=LI(gone.slice(0,4).map(function(r){ return b(r.n)+" — было "+mln(r.first)+" в месяц, стало "+mln(r.last)+". Потеря "+b(mln(r.d))+" выручки ежемесячно."; }));
      }
      if(grown.length){
        h+=P(b("Выросшие каналы:"));
        h+=LI(grown.slice(0,4).map(function(r){ return b(r.n)+" — с "+mln(r.first)+" до "+mln(r.last)+" в месяц, "+b(sg(r.d))+"."; }));
      }
      var netCh=(grown.reduce(function(s,r){return s+r.d;},0))+(gone.reduce(function(s,r){return s+r.d;},0));
      h+=P("Нетто-эффект по каналам: "+b(sg(netCh)+" выручки в месяц")+". При маржинальности "+pc(a.cmr)+" это "+b(sg(netCh*a.cmr/100))+" результата ежемесячно — "+(netCh<0?"ровно та дыра, которую нечем закрыть при неизменных постоянных затратах.":"вклад в прибыль."));
      var top=chRows.slice(0,5);
      h+=KV(top.map(function(r){ return [r.n, mln(r.t)+" · "+pc(r.t/chRows.reduce(function(s,x){return s+x.t;},0)*100,0)]; }));
      h+=P("Концентрация: три крупнейших канала дают "+b(pc(top.slice(0,3).reduce(function(s,r){return s+r.t;},0)/chRows.reduce(function(s,x){return s+x.t;},0)*100))+" всей выручки. Уход любого из них повторит сценарий текущего года.");
      }

      h+=H("7 · Продукт и фудкост");
      if(CATS().length){
        var cs=CATS().slice(0,10);
        var bestC=cs.slice().sort(function(x,y){ return x.fc-y.fc; })[0];
        var worstC=cs.slice().sort(function(x,y){ return y.fc-x.fc; })[0];
        h+=P("Средний фудкост по категориям с заполненной себестоимостью — "+b(pc(cs.reduce(function(s,c){return s+c.cost;},0)/cs.reduce(function(s,c){return s+c.rev;},0)*100))+". Самая выгодная категория — "+b(bestC.n+" ("+pc(bestC.fc)+")")+", самая тяжёлая — "+b(worstC.n+" ("+pc(worstC.fc)+")")+".");
        h+=KV(cs.map(function(c){ return [c.n, mln(c.rev)+" выручки · фудкост "+pc(c.fc)+" · вал. прибыль "+mln(c.gp), c.fc>55?"#ef4444":(c.fc<45?"#22c55e":"#e2e8f0")]; }));
        h+=P("Категории с фудкостом выше 55% при полной себестоимости завода около "+pc(a.full/a.rev*100)+" не окупают даже производство — по ним нужен либо пересмотр цены, либо вывод из матрицы.");
      } else { h+=P("Данные по категориям недоступны."); }

      h+=H("8 · Потери внутри себестоимости");
      h+=P("Сумма статей потерь за период — "+b(mln(lossTot))+", это "+b(pc(lossTot/a.rev*100))+" выручки и "+b(pc(lossTot/Math.max(1,Math.abs(a.op))*100,0))+" от абсолютной величины результата.");
      h+=KV(LOSS.map(function(k){ var s=lineSum(k,ms); return [k.replace(/^\d+(\.\d+)*\./,""), mln(s)+" · "+pc(s/a.rev*100,2), s>a.rev*0.005?"#ef4444":"#e2e8f0"]; }));

      h+=H("9 · За счёт чего прибыль и за счёт чего убыток");
      h+=LI([
        b("Прибыль дают:")+" маржинальность "+pc(a.cmr)+" — каждый тенге выручки приносит "+b(Math.round(a.cmr)+" тиын")+" на покрытие постоянных затрат; объём в сильные месяцы (в лучшем месяце результат "+mln(PL()[best].op)+"); каналы с растущим объёмом"+(grown.length?" — прежде всего "+grown[0].n:"")+".",
        b("Убыток создают:")+" падение выручки ниже порога "+mln(a.bep/n)+" в месяц; постоянные затраты "+mln(a.fix/n)+" в месяц, не сокращённые вслед за объёмом; потери в себестоимости "+mln(lossTot)+"; концентрация на нескольких крупных покупателях.",
        (function(){ var lv=a.rev/n*0.01*a.cmr/100, lf=a.fix/n*0.01, k=lv/lf;
          return b("Главный рычаг:")+" при нынешней структуре "+b("+1% выручки")+" даёт "+b(mln(lv))+" результата в месяц, а "+b("−1% постоянных затрат")+" — "+b(mln(lf))+". "
            +(k>1.15?("Наращивать объём примерно в "+b(k.toFixed(1).replace(".",",")+" раза")+" эффективнее, чем резать постоянные на тот же процент.")
              :(k<0.87?("Сокращать постоянные примерно в "+b((1/k).toFixed(1).replace(".",",")+" раза")+" эффективнее, чем наращивать объём на тот же процент.")
                :"Оба рычага дают почти одинаковый эффект, поэтому работать нужно с обоими сразу."));
        })()
      ]);

      h+=H("10 · Сценарии выхода в ноль");
      if(a.op<0){
        h+=LI([
          "Только за счёт объёма: нужно "+b("+"+mln(needRev/n)+" выручки в месяц")+" ("+b(pc(needRev/a.rev*100))+" к текущей) при сохранении маржинальности и постоянных.",
          "Только за счёт постоянных: сократить их на "+b(mln(perMonth)+" в месяц")+" ("+b(pc(perMonth/(a.fix/n)*100))+" от нынешних "+mln(a.fix/n)+").",
          "Только за счёт маржинальности: поднять её на "+b((deficit/a.rev*100).toFixed(1).replace(".",",")+" пункта")+" — с "+pc(a.cmr)+" до "+b(pc(a.cmr+deficit/a.rev*100))+". Это либо цена, либо фудкост, либо сокращение потерь.",
          "Комбинация, наиболее реалистичная: половина разрыва объёмом (+"+mln(needRev/n/2)+" выручки), четверть — маржинальностью (+"+(deficit/a.rev*100/4).toFixed(1).replace(".",",")+" пункта), четверть — постоянными (−"+mln(perMonth/4)+" в месяц)."
        ]);
      } else {
        h+=P("Период прибыльный. Запас прочности "+b(pc(a.safety))+": выручка может упасть на "+b(mln(a.rev-a.bep))+" до точки безубыточности.");
      }

      h+=H("11 · Методика и ограничения");
      h+=P("Данные — управленческий ОПиУ ("+MONTHS().length+" мес., "+MN[+MONTHS()[0].slice(5)]+" "+MONTHS()[0].slice(0,4)+" — "+MN[+MONTHS()[MONTHS().length-1].slice(5)]+" "+MONTHS()[MONTHS().length-1].slice(0,4)+") и отчёт о продажах по контрагентам из iiko. Переменными считаются статьи, зависящие от объёма: продуктовая себестоимость, расходные материалы, логистика, потери, электроэнергия. Постоянными — ФОТ производства и АУП, аренда, администрация, маркетинг, ремонты. Деление условное: часть статей полупеременные, поэтому точка безубыточности — ориентир, а не бухгалтерская величина. Выручка по каналам берётся по расходным накладным и может незначительно расходиться с ОПиУ из-за возвратов и внутренних перемещений. Прочие доходы и КПН в операционный результат не включены.");
      h+=P('<span style="color:#475569;font-size:11.5px">Собрано '+D.built+' · система «Пульс» · Ольга Герасименко</span>');

      document.getElementById("fc-modal-body").innerHTML=h;
      document.getElementById("fc-modal-sub").textContent=(st.period==="all"?"весь период":st.period)+" · "+n+" мес. · выручка "+mln(a.rev)+" · результат "+mln(a.op);
      document.getElementById("fc-modal").style.display="block";
      document.body.style.overflow="hidden";
    }
'''


def _match_div(s, start):
    """Индекс сразу за парным </div> для тега <div, начинающегося в start."""
    depth = 0
    for m in re.finditer(r"<(/?)div\b[^>]*?(/?)>", s[start:]):
        if m.group(2) == "/":
            continue
        depth += -1 if m.group(1) else 1
        if depth == 0:
            return start + m.end()
    return -1


def inject(html, data):
    """Кладём блок аналитики СРАЗУ ПОД ШАПКОЙ — он главный на странице, а не сноска внизу."""
    block = SECTION.replace("__MODAL__", MODAL_JS).replace(
        "__FCDATA__", json.dumps(data, ensure_ascii=False, separators=(",", ":")))
    # вырезаем прошлую версию блока (ровно её, а не всё до конца страницы)
    i = html.find('<div id="fullcost-analytics"')
    if i >= 0:
        j = _match_div(html, i)
        rest = html[j:] if j > 0 else ""
        # Пустоту по краям разреза раньше не подчищали: каждый прогон оставлял три пустые
        # строки, а сборка идёт 4 раза в сутки — их накопилось под сотню. Срезаем с обеих сторон.
        html = html[:i].rstrip(" \t\n") + "\n" + rest.lstrip(" \t\n")
    # вставляем сразу после шапки
    a = html.find('<div class="topbar"')
    if a >= 0:
        b = _match_div(html, a)
        if b > 0:
            return html[:b] + "\n" + block + html[b:]
    k = html.find("</body>")
    if k >= 0:
        return html[:k] + block + "\n" + html[k:]
    return html + block


def main():
    data = build()
    p = os.path.join(HERE, TARGET)
    html = open(p, encoding="utf-8").read()
    open(p, "w", encoding="utf-8").write(inject(html, data))
    a = data["years"]
    tot_rev = sum(data["pl"][m]["rev"] for m in data["months"])
    tot_op = sum(data["pl"][m]["op"] for m in data["months"])
    print("Полная себестоимость: %d мес., выручка %.1f млн, результат %.1f млн, каналов %d, категорий %d"
          % (len(data["months"]), tot_rev / 1e6, tot_op / 1e6, len(data["chan"]), len(data["cats"])))


if __name__ == "__main__":
    main()
