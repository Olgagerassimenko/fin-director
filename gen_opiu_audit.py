# -*- coding: utf-8 -*-
"""gen_opiu_audit.py — аудит ОПиУ 2025–2026 постатейно.

Читает «Отчет о прибылях и убытках 2025-2026.xlsx» (эталон бухгалтерии, помесячно)
и собирает opiu_audit.js для страницы «Аудит ОПиУ».

Что считается:
  months   — список месяцев по порядку;
  rows     — все строки ОПиУ с иерархией и значениями по месяцам;
  kpi      — выручка, валовая, операционная, чистая по месяцам и годам;
  findings — замечания аудитора по каждой статье:
             перенос между периодами, всплеск, знак, круглая сумма,
             статья прекратилась/появилась, зеркальная пара, дрейф доли,
             спорная классификация.

Только чтение xlsx. Запускается в GitHub Actions после остальных сборок.
"""
import datetime as _dt
import json, os, re, statistics
import os as _os, sys as _sys
_sys.path.insert(0, _os.path.dirname(_os.path.abspath(__file__)))
import almaty  # время завода — Алматы (UTC+5), не UTC раннера

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "Отчет о прибылях и убытках 2025-2026.xlsx")

MON = ["янв", "фев", "мар", "апр", "май", "июн", "июл", "авг", "сен", "окт", "ноя", "дек"]

# ── расчётные строки: это не статьи учёта, а результат вычитания
COMPUTED = {"Валовая прибыль"}

# ── итоги, которые считаются вычитанием, а не сложением строк
SUBTRACT = {"Итого Прибыль от основной деятельности", "ИТОГО ЧИСТАЯ ПРИБЫЛЬ"}

# ── строки выручки: к ним расходные проверки не применяем
REVENUE = {"Торговая выручка", "Выручка", "Итого Выручка"}

# ── строки, где минус — норма (контр-статьи), а не ошибка разноски
CONTRA = {
    "1.4.Излишки инвентаризации",
    "1.13.Коррекция отрицательных остатков на складе",
}

# ── экспертные пометки о месте статьи в отчёте
PLACEMENT = {
    "Логистика доставка": (
        "Сидит в «Расходы по реализации», но это доставка готовой продукции покупателю. "
        "По МСФО это расходы на сбыт — место верное, однако валовая прибыль завода "
        "их не видит: маржа по продукту выглядит лучше, чем есть. "
        "Для управленческих решений считайте её вместе с себестоимостью."),
    "1.18.Электроэнергия": (
        "Отнесена в «С/с реализованной продукции». Постоянная по природе: "
        "при падении объёма себестоимость единицы автоматически растёт. "
        "Проверьте, не сидит ли здесь энергия склада и офиса."),
    "1.16.Мусор": (
        "В себестоимости продукции. Это вывоз ТБО по договору — постоянный расход площадки, "
        "не переменный расход продукта."),
    "3.3.2.Налоги НДС": (
        "НДС в составе административных расходов. НДС — не расход периода, а расчёты с бюджетом; "
        "к расходу относится только невозмещаемая часть. Суммы ровные (5/10/15 млн) — "
        "это начисление «на глаз», а не факт декларации."),
    "1.14.Ремонт помещений": (
        "В себестоимости продукции. Крупные ремонты — капитализируемые затраты либо "
        "административные; в себестоимости они искажают маржу конкретного месяца."),
    "1.9.Ремонт/Обслуживание производ.оборудования": (
        "В себестоимости продукции. Плановое ТО — да, но капитальный ремонт "
        "должен идти через основные средства."),
    "1.30.Возвраты от дистрибьютеров": (
        "Возвраты показаны расходом в себестоимости, а не сторнированием выручки. "
        "Из-за этого выручка завышена, а себестоимость раздута на ту же сумму — "
        "валовая маржа считается неверно."),
    "Удержания из ЗП": (
        "Показаны в «Прочих доходах». Удержание из зарплаты — это не доход, "
        "а уменьшение расхода по ФОТ. ФОТ завышен примерно на эту сумму."),
    "3.3.1.Админ.расходы ПРОЧИЕ": (
        "Статья-«помойка»: суммы прыгают от 0,3 до 17,5 млн. "
        "Расшифровку по этой статье нужно требовать ежемесячно."),
}


def load():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.worksheets[0]

    # ── шапка: строка с датами. iiko отдаёт их то датой, то строкой «31.01.2025»
    def ym(v):
        if hasattr(v, "year"):
            return v.year, v.month
        m = re.match(r"^\s*(\d{2})\.(\d{2})\.(\d{4})\s*$", str(v or ""))
        return (int(m.group(3)), int(m.group(2))) if m else None

    hdr = None
    for r in range(1, 12):
        if sum(1 for c in range(2, ws.max_column + 1) if ym(ws.cell(r, c).value)) >= 6:
            hdr = r
            break
    if hdr is None:
        raise SystemExit("не нашёл строку с датами месяцев")

    cols, months, labels = [], [], []
    for c in range(2, ws.max_column + 1):
        p = ym(ws.cell(hdr, c).value)
        if p:
            cols.append(c)
            months.append("%04d-%02d" % p)
            labels.append("%s %d" % (MON[p[1] - 1], p[0]))

    rows = []
    for r in range(hdr + 1, ws.max_row + 1):
        raw = ws.cell(r, 1).value
        if raw is None:
            continue
        s = str(raw)
        name = s.strip()
        if not name:
            continue
        ind = len(s) - len(s.lstrip())
        vals = []
        for c in cols:
            v = ws.cell(r, c).value
            vals.append(float(v) if isinstance(v, (int, float)) else 0.0)
        low = name.lower()
        if low.startswith(("итого", "всего")) or name in COMPUTED:
            kind = "total"
        elif ind == 0:
            kind = "grand"
        elif not any(vals):
            kind = "sub"           # заголовок группы: значений нет
        else:
            kind = "line"
        rows.append({"n": name, "ind": ind, "kind": kind,
                     "v": [round(x) for x in vals]})

    # ── знак статьи: рост дохода — хорошо, рост расхода — плохо.
    #    Определяем по блоку отчёта, а не по названию.
    blk = "rev"
    for r in rows:
        n = r["n"]
        if n == "Себестоимость":
            blk = "cogs"
        elif n == "Расходы":
            blk = "opex"
        elif n == "Прочие доходы":
            blk = "oinc"
        elif n == "Прочие расходы":
            blk = "oexp"
        r["g"] = blk
        r["inc"] = 1 if blk in ("rev", "oinc") else 0

    return months, labels, rows


def checksums(rows):
    """Каждый «Итого» против суммы своих прямых строк.

    Отступ в выгрузке iiko — 3 пробела на уровень. Прямые дети группы имеют
    отступ ровно на 3 больше заголовка; вложенные группы участвуют своим «Итого».
    Три строки считаются вычитанием, их проверяем отдельно.
    """
    out = []
    idx = {r["n"]: r for r in rows}
    n = len(rows[0]["v"]) if rows else 0

    for j, t in enumerate(rows):
        if t["kind"] != "total" or t["n"] in COMPUTED or t["n"] in SUBTRACT:
            continue
        # ищем заголовок группы выше: тот же отступ, без значений
        h = None
        for i in range(j - 1, -1, -1):
            if rows[i]["ind"] < t["ind"]:
                break
            if rows[i]["ind"] == t["ind"] and rows[i]["kind"] == "sub":
                h = i
                break
        if h is None:
            continue
        kids = [rows[i] for i in range(h + 1, j)
                if rows[i]["ind"] == t["ind"] + 3 and rows[i]["kind"] in ("line", "total")]
        if len(kids) < 2:
            continue
        s = [sum(k["v"][m] for k in kids) for m in range(n)]
        out.append({"name": t["n"], "kids": [k["n"] for k in kids], "tot": t["v"],
                    "sum": s, "diff": [t["v"][m] - s[m] for m in range(n)]})

    # расчётные строки: разность, а не сумма
    for name, a, b in (("Валовая прибыль", "Итого Выручка", "Итого Себестоимость"),
                       ("Итого Прибыль от основной деятельности",
                        "Валовая прибыль", "Итого Расходы")):
        if name in idx and a in idx and b in idx:
            s = [idx[a]["v"][m] - idx[b]["v"][m] for m in range(n)]
            out.append({"name": name, "kids": [a, "минус " + b], "tot": idx[name]["v"],
                        "sum": s, "diff": [idx[name]["v"][m] - s[m] for m in range(n)]})
    if all(k in idx for k in ("ИТОГО ЧИСТАЯ ПРИБЫЛЬ", "Итого Прибыль от основной деятельности",
                              "Итого Прочие доходы", "Итого Прочие расходы")):
        s = [idx["Итого Прибыль от основной деятельности"]["v"][m]
             + idx["Итого Прочие доходы"]["v"][m]
             - idx["Итого Прочие расходы"]["v"][m] for m in range(n)]
        out.append({"name": "ИТОГО ЧИСТАЯ ПРИБЫЛЬ",
                    "kids": ["Итого Прибыль от основной деятельности",
                             "Итого Прочие доходы", "минус Итого Прочие расходы"],
                    "tot": idx["ИТОГО ЧИСТАЯ ПРИБЫЛЬ"]["v"], "sum": s,
                    "diff": [idx["ИТОГО ЧИСТАЯ ПРИБЫЛЬ"]["v"][m] - s[m] for m in range(n)]})
    return out


def pick(rows, name):
    for r in rows:
        if r["n"] == name:
            return r["v"]
    return None


def build():
    months, labels, rows = load()
    n = len(months)

    rev = pick(rows, "Итого Выручка") or [0] * n
    gross = pick(rows, "Валовая прибыль") or [0] * n
    cogs = pick(rows, "Итого Себестоимость") or [0] * n
    opex = pick(rows, "Итого Расходы") or [0] * n
    oper = pick(rows, "Итого Прибыль от основной деятельности") or [0] * n
    net = pick(rows, "ИТОГО ЧИСТАЯ ПРИБЫЛЬ") or [0] * n

    lines = [r for r in rows if r["kind"] == "line"]

    # ── проверки
    F = []

    def add(sev, typ, art, mon, title, txt, amt=0):
        F.append({"sev": sev, "t": typ, "a": art, "m": mon,
                  "h": title, "d": txt, "s": round(amt)})

    for r in lines:
        name, v = r["n"], r["v"]
        if name in ("Торговая выручка", "Выручка", "Итого Выручка"):
            continue
        nz = [x for x in v if x]
        if not nz:
            continue
        med = statistics.median([abs(x) for x in nz])
        fill = len(nz) / float(n)

        # 1. перенос между периодами: ноль в регулярной статье + удвоение рядом
        if fill >= 0.7 and med >= 200000:
            for i, x in enumerate(v):
                if x != 0:
                    continue
                nb = []
                if i > 0:
                    nb.append((i - 1, v[i - 1]))
                if i < n - 1:
                    nb.append((i + 1, v[i + 1]))
                big = [(j, y) for j, y in nb if abs(y) >= med * 1.6]
                if big:
                    j, y = max(big, key=lambda p: abs(p[1]))
                    add(2, "shift", name, months[i],
                        "Пропуск месяца и удвоение в соседнем",
                        "В %s по статье ноль, а в %s — %s при обычных %s. "
                        "Похоже, документ провели не тем месяцем: расход одного месяца "
                        "лёг на другой и исказил оба." % (
                            labels[i], labels[j], fmt(y), fmt(med)),
                        abs(y) - med)
                elif fill >= 0.85:
                    add(1, "zero", name, months[i],
                        "Регулярная статья обнулилась",
                        "Статья платится почти каждый месяц (обычно %s), "
                        "а в %s — ноль. Либо документ не провели, либо расход "
                        "ушёл на другую статью." % (fmt(med), labels[i]), med)

        # 2. всплеск — не больше двух самых крупных на статью, иначе список не читается
        if med >= 150000:
            sp = [(i, x) for i, x in enumerate(v)
                  if abs(x) > med * 2.5 and abs(x) - med > 1000000]
            sp.sort(key=lambda p: -(abs(p[1]) - med))
            for i, x in sp[:2]:
                add(2 if abs(x) - med > 5000000 else 1, "spike", name, months[i],
                    "Разовый всплеск",
                    "%s: %s против обычных %s — в %.1f раза больше нормы%s. "
                    "Проверьте первичку: разовый расход, начисление за несколько "
                    "месяцев сразу или ошибка счёта." % (
                        labels[i], fmt(x), fmt(med), abs(x) / med if med else 0,
                        "" if len(sp) <= 2 else " (всего таких месяцев %d)" % len(sp)),
                    abs(x) - med)

        # 3. знак (только для расходных статей: у доходных минус — это возврат, норма)
        if name not in CONTRA and not r["inc"]:
            neg = [(i, x) for i, x in enumerate(v) if x < -50000]
            pos = [x for x in v if x > 0]
            if neg and pos:
                i, x = min(neg, key=lambda p: p[1])
                add(2, "sign", name, months[i],
                    "Отрицательная сумма в расходной статье",
                    "В %s статья ушла в минус (%s), хотя в остальные месяцы это расход. "
                    "Обычно это сторно прошлого периода — прибыль этого месяца "
                    "приукрашена на эту сумму." % (labels[i], fmt(x)), abs(x))

        # 4. ровные суммы — ручное начисление
        rnd = [(i, x) for i, x in enumerate(v)
               if x >= 1000000 and x % 500000 == 0]
        if len(rnd) >= 3:
            add(1, "round", name, months[rnd[-1][0]],
                "Ровные суммы — начисление «на глаз»",
                "В %d месяцах сумма кратна 500 тыс. (%s). Так выглядит оценочное "
                "начисление, а не факт по документу. Расхождение с фактом накопится "
                "и вылезет при закрытии года." % (
                    len(rnd), ", ".join("%s %s" % (labels[i], fmt(x)) for i, x in rnd[-3:])),
                sum(x for _, x in rnd) / len(rnd))

        # 5. статья прекратилась
        tail = v[-3:]
        head = v[:-3]
        if not any(tail) and head and statistics.median([abs(x) for x in head if x] or [0]) >= 200000:
            hm = statistics.median([abs(x) for x in head if x])
            add(1, "stopped", name, months[-1],
                "Статья прекратилась",
                "Последние 3 месяца по статье ноль, до этого регулярно %s в месяц. "
                "Если расход никуда не делся — он теперь сидит в другой строке, "
                "и сравнение с прошлым годом врёт." % fmt(hm), hm * 3)

        # 6. статья появилась
        first_nz = next((i for i, x in enumerate(v) if x), None)
        if first_nz is not None and first_nz >= 5 and med >= 500000:
            add(1, "started", name, months[first_nz],
                "Статья появилась в середине периода",
                "До %s статьи не было, дальше — регулярно %s в месяц. "
                "Сравнение «год к году» по этой строке некорректно, "
                "и стоит проверить, из какой статьи она выделилась." % (
                    labels[first_nz], fmt(med)), med)

        # 7. дрейф доли в выручке
        sh = [x / rev[i] if rev[i] else 0 for i, x in enumerate(v)]
        if med >= 1000000:
            msh = statistics.median(sh)
            for i, s in enumerate(sh):
                if abs(s - msh) > 0.015 and abs(v[i]) > 3000000:
                    add(1, "drift", name, months[i],
                        "Доля в выручке скакнула",
                        "%s: %.1f%% выручки против обычных %.1f%% — "
                        "%+.1f п.п. Это %s в деньгах." % (
                            labels[i], s * 100, msh * 100, (s - msh) * 100,
                            fmt((s - msh) * rev[i])),
                        abs((s - msh) * rev[i]))
                    break

    # 8. зеркальные пары — перекинули между счетами.
    #    Проверка чувствительна к случайным совпадениям сумм, поэтому берём только
    #    расходные статьи, требуем совпадение лучше 1,5 % и не больше одной пары
    #    на статью за месяц. Помечаем как «гипотеза», а не как факт.
    seen, used = set(), set()
    for i in range(1, n):
        d = []
        for r in lines:
            if r["n"] in REVENUE:
                continue
            dd = r["v"][i] - r["v"][i - 1]
            if abs(dd) > 1500000:
                d.append((r["n"], dd))
        ups = sorted([x for x in d if x[1] > 0], key=lambda p: -p[1])
        dns = [x for x in d if x[1] < 0]
        for na, da in ups:
            if (i, na) in used:
                continue
            cand = [(nb, db) for nb, db in dns
                    if (i, nb) not in used and abs(da + db) <= abs(da) * 0.015]
            if not cand:
                continue
            nb, db = min(cand, key=lambda p: abs(da + p[1]))
            key = tuple(sorted((na, nb)))
            if key in seen:
                continue
            seen.add(key)
            used.add((i, na))
            used.add((i, nb))
            add(1, "mirror", na, months[i],
                "Зеркальное движение двух статей",
                "В %s «%s» выросла на %s, и почти ровно на столько же упала «%s» "
                "(совпадение %.1f%%). Это гипотеза, а не приговор: суммы могли совпасть "
                "случайно. Стоит открыть проводки обеих статей за месяц — если это "
                "перенос с одного счёта на другой, расход не изменился, "
                "изменилось только место." % (
                    labels[i], na, fmt(da), nb,
                    100 - abs(da + db) / abs(da) * 100),
                da)

    # 10. свод не сходится: «Итого» против суммы своих строк
    CHK = checksums(rows)
    foot_ok = sum(1 for t in CHK if not any(abs(d) >= 1 for d in t["diff"]))
    for t in CHK:
        bad = [(i, t["diff"][i]) for i in range(n) if abs(t["diff"][i]) >= 1]
        if not bad:
            continue
        i, dv = max(bad, key=lambda p: abs(p[1]))
        add(2, "foot", t["name"], months[i],
            "Итог не сходится со своими строками",
            "«%s» в %s = %s, а сумма входящих строк = %s. Разница %s%s. "
            "Либо в отчёт не попала строка, либо итог считается по другой формуле — "
            "любая аналитика по этой группе поедет." % (
                t["name"], labels[i], fmt(t["tot"][i]), fmt(t["sum"][i]), fmt(dv),
                "" if len(bad) == 1 else " (расходится в %d месяцах из %d)" % (len(bad), n)),
            abs(dv))

    # 9. экспертная классификация
    for r in lines:
        if r["n"] in PLACEMENT:
            tot = sum(abs(x) for x in r["v"])
            add(1, "class", r["n"], months[-1], "Спорное место статьи в отчёте",
                PLACEMENT[r["n"]], tot / n)

    F.sort(key=lambda f: (-f["sev"], -f["s"]))

    # ── годовые своды
    def yr(y):
        idx = [i for i, m in enumerate(months) if m[:4] == y]
        if not idx:
            return None
        return {"y": y, "n": len(idx), "from": months[idx[0]], "to": months[idx[-1]],
                "rev": sum(rev[i] for i in idx), "cogs": sum(cogs[i] for i in idx),
                "gross": sum(gross[i] for i in idx), "opex": sum(opex[i] for i in idx),
                "oper": sum(oper[i] for i in idx), "net": sum(net[i] for i in idx)}

    years = [y for y in (yr("2025"), yr("2026")) if y]

    # сопоставимый период: те же месяцы 2026 против 2025
    lp = None
    m26 = [m[5:] for m in months if m[:4] == "2026"]
    if m26:
        i25 = [i for i, m in enumerate(months) if m[:4] == "2025" and m[5:] in m26]
        i26 = [i for i, m in enumerate(months) if m[:4] == "2026"]
        if i25:
            lp = {"mm": len(m26),
                  "a": {k: sum(src[i] for i in i25) for k, src in
                        (("rev", rev), ("cogs", cogs), ("gross", gross),
                         ("opex", opex), ("oper", oper), ("net", net))},
                  "b": {k: sum(src[i] for i in i26) for k, src in
                        (("rev", rev), ("cogs", cogs), ("gross", gross),
                         ("opex", opex), ("oper", oper), ("net", net))}}

    # ── насколько бухгалтерский отчёт отстаёт от календаря
    t = almaty.today()
    ly, lm = int(months[-1][:4]), int(months[-1][5:])
    lag = (t.year - ly) * 12 + (t.month - lm) - 1   # сколько месяцев закрыто, но не в отчёте
    src = os.path.basename(XLSX)
    try:
        mt = _dt.datetime.fromtimestamp(os.path.getmtime(XLSX)).strftime("%d.%m.%Y")
    except Exception:
        mt = ""

    return {
        "updated": almaty.now().strftime("%d.%m.%Y %H:%M"),
        "src": src, "srcDate": mt, "lag": max(0, lag), "lastLabel": labels[-1],
        "months": months, "labels": labels,
        "rows": rows,
        "kpi": {"rev": rev, "cogs": cogs, "gross": gross,
                "opex": opex, "oper": oper, "net": net},
        "years": years, "lp": lp,
        "foot": {"ok": foot_ok, "all": len(CHK)},
        "findings": F[:220],
        "fstat": {k: sum(1 for f in F if f["t"] == k) for k in
                  {f["t"] for f in F}},
    }


def fmt(x):
    x = float(x)
    a = abs(x)
    if a >= 1e9:
        return "%.2f млрд ₸" % (x / 1e9)
    if a >= 1e6:
        return "%.1f млн ₸" % (x / 1e6)
    if a >= 1e3:
        return "%.0f тыс ₸" % (x / 1e3)
    return "%.0f ₸" % x


if __name__ == "__main__":
    d = build()
    open(os.path.join(HERE, "opiu_audit.js"), "w", encoding="utf-8").write(
        "window.OPIU_AUDIT=" + json.dumps(d, ensure_ascii=False, separators=(",", ":")) + ";")
    print("opiu_audit.js: месяцев %d, строк %d, замечаний %d"
          % (len(d["months"]), len(d["rows"]), len(d["findings"])))
    for k, c in sorted(d["fstat"].items(), key=lambda p: -p[1]):
        print("   %-9s %d" % (k, c))
