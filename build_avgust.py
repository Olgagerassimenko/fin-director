# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "ДДС МЕСЯЦ АВГУСТ"
FONT="Arial"
blue=Font(name=FONT,size=10,color="0000FF")      # план-вход (правьте)
black=Font(name=FONT,size=10,color="000000")     # формулы
boldb=Font(name=FONT,size=10,bold=True)
white=Font(name=FONT,size=10,bold=True,color="FFFFFF")
grayf=Font(name=FONT,size=9,italic=True,color="666666")
yfill=PatternFill("solid",fgColor="FFF9C4")       # факт — заполнять
hfill=PatternFill("solid",fgColor="1F4E79")
sfill=PatternFill("solid",fgColor="D9E1F2")       # итоги
ifill=PatternFill("solid",fgColor="E2EFDA")       # выручка/операц
thin=Side(style="thin",color="BFBFBF")
bd=Border(left=thin,right=thin,top=thin,bottom=thin)
money='#,##0;(#,##0);-'
center=Alignment(horizontal="center",vertical="center",wrap_text=True)
right=Alignment(horizontal="right"); left=Alignment(horizontal="left")

# недели августа: (label, датаднапазон)
weeks=[("Неделя 1","03.08–07.08"),("Неделя 2","10.08–14.08"),("Неделя 3","17.08–21.08"),
       ("Неделя 4","24.08–28.08"),("Неделя 5","31.08–04.09")]
# колонки: C..L  (План/Факт на каждую неделю)
wcols=[("C","D"),("E","F"),("G","H"),("I","J"),("K","L")]

# ---- ПЛАН (из «Неделя ОПЛАТ»), seed {(row_key, week_index): amount} ----
# доход по источникам (июльский план, как образец распределения на август)
inc = {
 "85-ДФЗ дистрибьютер ФЗ":[22250000,22250000,23000000,22000000,22000000],
 "Яндекс":[11000000,12000000,13000000,11500000,11500000],
 "7-Kaspi Банк":[5000000,5000000,6000000,7000000,7000000],
 "99-RP АЗС":[4000000,4000000,6000000,5000000,5000000],
 "96- DSF покупатель":[4000000,3000000,5000000,0,0],
 "Май Март":[1000000,1500000,9000000,0,3000000],
 "Б1, Б2, Б5":[2600000,2600000,2500000,1800000,1800000],
 "Глово, Вольт":[1300000,400000,1000000,500000,500000],
 "110 АЗС №67 \"Sinooil\"":[600000,0,1115000,0,0],
 "Прочие":[500000,4400000,3250000,1320000,1320000],
}
# доход по кассе
inc_k = {
 "Частное лицо (Покупатель)":[1000000,0,1300000,400000,400000],
 "Crave":[0,0,1000000,0,0],
 "97- Кейтеринг гость":[0,0,500000,0,0],
 "Торты":[0,500000,0,520000,520000],
}
# расход по банку (безнал) — отрицательные
eb = {
 "Сырье Платежи поставщикам":[-34000000,-26000000,0,-35000000,-35000000],
 "IT-обслуживание,телефония,интернет":[-500000,-32800,0,0,0],
 "Административные расходы":[-500000,0,0,0,0],
 "Аренда помещений":[0,-6655000,0,0,0],
 "Ремонт оборудования":[-500000,-600000,0,-1000000,0],
 "Оплата труда":[0,-30000000,0,0,0],
 "Налоги и другие платежи в бюджет":[0,0,-15000000,0,-8625948],
 "Погашение кредита и %%":[-1749942,0,-4013564,0,0],
 "Инвентарь":[0,0,0,-6301360,0],
 "Мусор":[0,0,0,0,-100000],
 "Электроэнергия":[0,0,0,0,-9432232],
}
# расход по кассе (нал) — отрицательные
ek = {
 "Подотчет (Рынок сырье, Рынок ТМЦ)":[-1100000,-2000000,0,-3000000,0],
 "Оплата труда":[-3100000,-2650000,-40404000,0,-2200000],
 "Инвентарь":[-200000,0,0,0,0],
 "Расходы по реализации Разные":[-2500000,-2500000,0,-3500000,-1000000],
 "Аренда квартир":[0,0,-390000,0,-350149],
 "Производственные расходы":[0,0,0,0,-95600],
}

def colnum(letter): return openpyxl.utils.column_index_from_string(letter)

r=1
# Заголовок
ws.cell(r,1,"ДДС · Недельное планирование · Август 2026").font=Font(name=FONT,size=13,bold=True)
r+=1
ws.cell(r,1,"Формат «по аналогии с июнем»: недельные колонки План/Факт, статьи, остатки по банку и кассе. План — на основе недельного плана оплат (лист «Неделя ОПЛАТ»).").font=grayf
r+=2
top=r
# строка недель
ws.cell(r,1,"Статья").font=boldb; ws.cell(r,2,"").font=boldb
for i,(wl,wd) in enumerate(weeks):
    pc,fc=wcols[i]
    c1=ws.cell(r,colnum(pc),f"{wl}  ({wd})"); c1.font=white; c1.fill=hfill; c1.alignment=center
    ws.merge_cells(f"{pc}{r}:{fc}{r}")
    ws.cell(r,colnum(fc)).fill=hfill
r+=1
ws.cell(r,1,"").font=boldb
for i in range(5):
    pc,fc=wcols[i]
    a=ws.cell(r,colnum(pc),"План"); a.font=white; a.fill=hfill; a.alignment=center
    b=ws.cell(r,colnum(fc),"Факт"); b.font=white; b.fill=hfill; b.alignment=center
r+=1
hdr_rows=(top,r-1)

rowidx={}   # (label,section)->row
def addrow(label, section=None, marker="", kind="input", seed=None, key=None, fill_fact=True):
    """kind: input(blue plan), formula, subtotal, calc"""
    global r
    ws.cell(r,1,label).font=black; ws.cell(r,1).alignment=left
    if marker: ws.cell(r,2,marker).font=grayf
    for i in range(5):
        pc,fc=wcols[i]
        pcell=ws.cell(r,colnum(pc)); fcell=ws.cell(r,colnum(fc))
        pcell.number_format=money; fcell.number_format=money
        pcell.border=bd; fcell.border=bd
        if kind=="input":
            v=None
            if seed is not None: v=seed[i]
            if v: pcell.value=v; pcell.font=blue
            else: pcell.font=blue
            if fill_fact: fcell.fill=yfill
            fcell.font=black
    rowidx[(label,section)]=r
    rr=r; r+=1; return rr

def style_total(row, fill=sfill, bold=True):
    for c in range(1, colnum("L")+1):
        cell=ws.cell(row,c)
        cell.fill=fill
        if bold and c==1: cell.font=boldb
        elif bold: 
            f=cell.font; cell.font=Font(name=FONT,size=10,bold=True,color=f.color)

# --- Остатки на начало ---
sb=addrow("Остаток на начало по Банку","start",kind="input",fill_fact=True)
sk=addrow("Остаток на начало по Кассе","start",kind="input",fill_fact=True)
# начальные остатки W1 (вход) — из факта июля (касса iiko на 01.08)
ws.cell(sb,colnum("C"),3377730).font=blue
ws.cell(sk,colnum("C"),0).font=blue
ws.cell(sb,1).comment=Comment("Остаток на 01.08 — впишите фактические остатки по банку и кассе на начало месяца. Сейчас стоит остаток кассы iiko на конец июля.","Пульс")

# --- Приход по банку ---
inc_first=r
inc_bank_rows=[]
for lbl,seed in inc.items():
    inc_bank_rows.append(addrow(lbl,"incb","Приход",kind="input",seed=seed))
itg_ib=addrow("Итого по банку","incb_tot","Приход",kind="subtotal")
style_total(itg_ib)
# --- Приход по кассе ---
inc_kassa_rows=[]
for lbl,seed in inc_k.items():
    inc_kassa_rows.append(addrow(lbl,"inck","Приход",kind="input",seed=seed))
itg_ik=addrow("Итого по кассе","inck_tot","Приход",kind="subtotal")
style_total(itg_ik)
rev=addrow("1.Выручка","rev","Приход",kind="formula")
style_total(rev, fill=ifill)

# --- Расход по банку ---
exp_bank_rows=[]
for lbl,seed in eb.items():
    exp_bank_rows.append(addrow(lbl,"expb","Расход",kind="input",seed=seed))
itg_eb=addrow("Итого по банку","expb_tot","Расход",kind="subtotal")
style_total(itg_eb)
# --- Расход по кассе ---
exp_kassa_rows=[]
for lbl,seed in ek.items():
    exp_kassa_rows.append(addrow(lbl,"expk","Расход",kind="input",seed=seed))
itg_ek=addrow("Итого по кассе","expk_tot","Расход",kind="subtotal")
style_total(itg_ek)
tot_exp=addrow("ИТОГО РАСХОДЫ","totexp",kind="formula"); style_total(tot_exp)
oper=addrow("Операционная деятельность","oper",kind="formula"); style_total(oper, fill=ifill)
# --- Взаиморасчёты банк/касса ---
vfp_b=addrow("Касса взаиморасчёта банк","vfpb","Расход",kind="input")
vfp_k=addrow("Касса взаиморасчёта касса","vfpk","Приход",kind="input")
# --- Остатки на конец ---
eb_row=addrow("Остаток на конец ПО БАНКУ","endb",kind="formula"); style_total(eb_row)
ek_row=addrow("Остаток на конец ПО КАССЕ","endk",kind="formula"); style_total(ek_row)
tot_row=addrow("Итого","tot",kind="formula"); style_total(tot_row, fill=ifill)

# ---------- ФОРМУЛЫ ----------
def R(name): return rowidx[(name,None)] if (name,None) in rowidx else None
for i in range(5):
    pc,fc=wcols[i]
    for col in (pc,fc):
        # Итого по банку приход
        ws.cell(itg_ib,colnum(col)).value=f"=SUM({col}{inc_bank_rows[0]}:{col}{inc_bank_rows[-1]})"
        ws.cell(itg_ik,colnum(col)).value=f"=SUM({col}{inc_kassa_rows[0]}:{col}{inc_kassa_rows[-1]})"
        ws.cell(rev,colnum(col)).value=f"={col}{itg_ib}+{col}{itg_ik}"
        ws.cell(itg_eb,colnum(col)).value=f"=SUM({col}{exp_bank_rows[0]}:{col}{exp_bank_rows[-1]})"
        ws.cell(itg_ek,colnum(col)).value=f"=SUM({col}{exp_kassa_rows[0]}:{col}{exp_kassa_rows[-1]})"
        ws.cell(tot_exp,colnum(col)).value=f"={col}{itg_eb}+{col}{itg_ek}"
        ws.cell(oper,colnum(col)).value=f"={col}{rev}+{col}{tot_exp}"
        # остаток начало для недель >1 = конец пред недели (той же категории План/Факт)
        if i>0:
            ppc,pfc=wcols[i-1]
            prev = ppc if col==pc else pfc
            ws.cell(sb,colnum(col)).value=f"={prev}{eb_row}"; ws.cell(sb,colnum(col)).font=black
            ws.cell(sk,colnum(col)).value=f"={prev}{ek_row}"; ws.cell(sk,colnum(col)).font=black
        # конец банк = начало + приход банк + расход банк - взаиморасч банк
        ws.cell(eb_row,colnum(col)).value=f"={col}{sb}+{col}{itg_ib}+{col}{itg_eb}-{col}{vfp_b}"
        ws.cell(ek_row,colnum(col)).value=f"={col}{sk}+{col}{itg_ik}+{col}{itg_ek}+{col}{vfp_k}"
        ws.cell(tot_row,colnum(col)).value=f"={col}{eb_row}+{col}{ek_row}"

# ширины
ws.column_dimensions["A"].width=40; ws.column_dimensions["B"].width=9
for pc,fc in wcols:
    ws.column_dimensions[pc].width=13; ws.column_dimensions[fc].width=13
ws.freeze_panes="C6"

# Легенда снизу
r+=1
ws.cell(r,1,"Как пользоваться:").font=boldb; r+=1
for t in ["• Синие ячейки — план, правьте под август.",
          "• Жёлтые ячейки (Факт) — вписывайте фактические суммы по мере месяца.",
          "• Итоги, выручка, остатки — формулы, считаются сами; остаток на начало недели = остаток на конец предыдущей.",
          "• Расходы вносятся со знаком «минус».",
          "• План на август взят по образцу недельного плана оплат (июль) — скорректируйте суммы и даты под факт."]:
    ws.cell(r,1,t).font=grayf; r+=1

wb.save("/tmp/ДДС_недельное_планирование_август_2026.xlsx")
print("saved xlsx; rows used:", r)
