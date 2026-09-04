# -*- coding: utf-8 -*-
"""
rebuild_sales.py — пересобирает данные раздела «Продажи 2026» из выгрузок
«I Отчет  ПРОДАЖИ MM.2026.xlsx» (это проверенный источник: суммы сходятся
с отчётом iiko до тенге).

Для каждого месяца считает: выручку, категории, топ-20 SKU, Magnum Chef,
кол-во SKU. Валовую прибыль (total_gp) выгрузка не содержит — переносим
из прежних данных. Затем перезаписывает объект `const DS = {...}`
в продажи_2026.html.

Запуск: python rebuild_sales.py
"""
import sys, os, re, json, glob, warnings
from collections import defaultdict

warnings.filterwarnings("ignore")
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

import openpyxl

HERE      = os.path.dirname(os.path.abspath(__file__))
HTML_FILE = os.path.join(HERE, "продажи_2026.html")

RU_MONTHS = ["", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
             "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]

REVENUE_TYPE = "Выручка расходной накладной"


def log(*a):
    print(*a, flush=True)


def get_cat(name):
    """Классификатор категорий. Должен совпадать с skuCat() в sku_analytics.js,
    иначе диаграмма категорий и блок аналитики покажут разные цифры."""
    n = str(name).upper()
    def has(*w): return any(x in n for x in w)
    if has('КОМПОТ','МОРС','ЛИМОНАД','СОК ','ЧАЙ ','КОФЕ','ВОДА','НАПИТ','СМУЗИ','АЙРАН','КВАС'):
        return 'Напитки'
    if has('ТОРТ','БЕНТО'): return 'Торты'
    if has('КИМПАБ','ОНИГИР','УДОН','РАМЕН','СУШИ','ГИОЗА','ПОКЕ','ЯПОН','ВОК ','ЯННЕМ','ТОКПОК'):
        return 'Япония'
    if has('БЛИН','СЫРНИК','КАША','ОМЛЕТ','ЗАВТРАК','ГРАНОЛА','ХЛОПЬ'): return 'Завтраки'
    if has('САЛАТ','ШУБА','ВИНЕГРЕТ'): return 'Салаты'
    if has('СЭНДВИЧ','СЕНДВИЧ','БУРГЕР','ХОТ-ДОГ','ХОТДОГ','ДОГ (','ШАУРМА','ЧИАБАТТА','БАГЕТ С','ТОСТ'):
        return 'Сэндвичи'
    if has('ЧИЗКЕЙК','ТИРАМИСУ','БРАУНИ','МЕДОВИК','НАПОЛЕОН','ПИРОЖН','ЭКЛЕР','ДЕСЕРТ','ЧИА ',
           'ПУДДИНГ','ПУДИНГ','МАФФИН','КУКИС','ОРЕШКИ','СИННАБОН','МОРОЖ','ШАРЛОТКА','ПАХЛАВА',
           'ЗЕФИР','МАКАРУН'):
        return 'Десерты'
    if has('СОСИСКА В ТЕСТЕ','ПИРОЖОК','ПИРОГ','САМСА','СЛОЙК','КРУАССАН','БУЛОЧК','ХЛЕБ','ЛАВАШ',
           'БРЕЦЕЛЬ','СОЧНИК','ЛЕПЁШ','ЛЕПЕШ','ХАЧАПУРИ','ВЫПЕЧК','БАГЕТ','ШТРУДЕЛЬ','РУЛЕТ'):
        return 'Выпечка'
    if has('П/Ф','ПП*','КУРИЦ','ГОВЯД','СВИН','ИНДЕЙК','КОТЛЕТ','ШНИЦЕЛЬ','МАНТЫ','ПЛОВ','ПАСТА',
           'ПЕННЕ','ФУЗИЛЛИ','ЛАГМАН','ГУЙРУ','ЦОМЯН','ПЕЛЬМЕН','ВАРЕНИК','ТЕФТЕЛ','БРИЗОЛЬ','ЛЮЛЯ',
           'БЕФСТРОГ','ЗРАЗ','СУП ','БОРЩ','ТОМ ЯМ','КРЫЛЬЯ','ЗАПЕКАНК','ПЮРЕ','ГРЕЧК','РИС ','РАГУ',
           'ЖАРЕН','БИФШТЕКС','ГУЛЯШ','ТУЧИКЕН','ГАРНИР','ГОЛУБЦ','ФАРШ','ШАШЛЫК','СТЕЙК','НАГГЕТС',
           'КАРТОФ'):
        return 'Горячее'
    return 'Прочее'


def is_magnum(name):
    u = str(name).upper()
    return 'MAGNUM' in u or 'МАГНУМ' in u


def parse_report(path):
    """Возвращает (период_текст, dict[item] = {'rev':..,'qty':..})."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    period = ''
    items = defaultdict(lambda: {'rev': 0.0, 'qty': 0.0})
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 2 and row and row[0]:
            period = str(row[0])
        if i < 5 or not row or len(row) < 5:
            continue
        if str(row[2]).strip() != REVENUE_TYPE:
            continue
        name = str(row[1] or '').strip()
        if not name:
            continue
        try:
            rev = float(row[4])
        except (TypeError, ValueError):
            continue
        try:
            qty = abs(float(row[3]))
        except (TypeError, ValueError):
            qty = 0.0
        items[name]['rev'] += rev
        items[name]['qty'] += qty
    wb.close()
    return period, items


def build_month(items):
    cat_agg = defaultdict(lambda: {'rev': 0.0, 'qty': 0.0, 'count': 0})
    total_rev = 0.0
    for name, v in items.items():
        cat = get_cat(name)
        total_rev += v['rev']
        cat_agg[cat]['rev'] += v['rev']
        cat_agg[cat]['qty'] += v['qty']
        cat_agg[cat]['count'] += 1

    cats = [{'cat': c, 'rev': round(x['rev']), 'qty': round(x['qty']), 'count': x['count'],
             'pct': round(x['rev'] / total_rev * 100, 1) if total_rev else 0}
            for c, x in sorted(cat_agg.items(), key=lambda kv: -kv[1]['rev'])]

    ranked = sorted(items.items(), key=lambda kv: -kv[1]['rev'])
    top20 = [{'name': n, 'cat': get_cat(n), 'rev': round(v['rev']), 'qty': round(v['qty']),
              'magnum': is_magnum(n)} for n, v in ranked[:20]]

    mag = [(n, v) for n, v in items.items() if is_magnum(n)]
    mag_items = [{'name': n, 'rev': round(v['rev']), 'qty': round(v['qty'])}
                 for n, v in sorted(mag, key=lambda kv: -kv[1]['rev'])[:30]]
    mag_rev = sum(v['rev'] for _, v in mag)

    return {
        'total_rev': round(total_rev),
        'mag_rev': round(mag_rev),
        'mag_pct': round(mag_rev / total_rev * 100, 1) if total_rev else 0,
        'sku_count': len(items),
        'categories': cats,
        'top20': top20,
        'magnum_items': mag_items,
    }


def read_period(path):
    """Быстро читает строку «Период: с ДД.ММ.ГГГГ по ДД.ММ.ГГГГ»."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    txt = ''
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 2:
            txt = str(row[0]) if row and row[0] else ''
            break
    wb.close()
    m = re.search(r'с\s*(\d{2})\.(\d{2})\.(\d{4})\s*по\s*(\d{2})\.(\d{2})\.(\d{4})', txt)
    if not m:
        return None
    d1, m1, y1 = int(m.group(1)), int(m.group(2)), int(m.group(3))
    d2, m2, y2 = int(m.group(4)), int(m.group(5)), int(m.group(6))
    return {'txt': txt, 'd1': d1, 'm1': m1, 'y1': y1, 'd2': d2, 'm2': m2, 'y2': y2}


def scan_reports(folder, year):
    """По каждому месяцу выбирает выгрузку с 1-го числа и самой поздней датой конца."""
    best = {}
    for path in glob.glob(os.path.join(folder, "I Отчет*ПРОДАЖИ*.xlsx")):
        if os.path.basename(path).startswith('~$'):
            continue
        try:
            p = read_period(path)
        except Exception:
            continue
        # берём только выгрузки «с 1-го числа» внутри одного месяца нужного года
        if not p or p['y1'] != year or p['m1'] != p['m2'] or p['d1'] != 1:
            continue
        m = p['m1']
        if m not in best or p['d2'] > best[m][2]['d2']:
            best[m] = (path, p['txt'], p)
    return best


def period_label(month_num, period_text):
    """«Период: с 01.07.2026 по 13.07.2026» -> ('Июль (1–13)', True)."""
    base = RU_MONTHS[month_num]
    m = re.search(r'с\s*(\d{2})\.(\d{2})\.(\d{4})\s*по\s*(\d{2})\.(\d{2})\.(\d{4})', period_text or '')
    if not m:
        return base, False
    d1, m1, d2, m2 = int(m.group(1)), int(m.group(2)), int(m.group(4)), int(m.group(5))
    import calendar
    last = calendar.monthrange(int(m.group(3)), m1)[1]
    if d1 == 1 and d2 >= last:
        return base, False
    return f"{base} ({d1}–{d2})", True


def load_html_and_old_ds():
    h = open(HTML_FILE, encoding='utf-8').read()
    i = h.find('const DS =')
    seg = h[i + len('const DS ='):]
    depth = 0; end = -1; st = False
    for j, ch in enumerate(seg):
        if ch == '{':
            depth += 1; st = True
        elif ch == '}':
            depth -= 1
            if st and depth == 0:
                end = j + 1; break
    obj = re.sub(r"'([A-Za-z_][A-Za-z0-9_]*)'\s*:", r'"\1":', seg[:end])
    return h, json.loads(obj)


def write_sales_sum(ds):
    """Короткая сводка продаж для плашки «Пульс» на главной.

    Главная не может грузить продажи_2026.html целиком (2 МБ), поэтому
    рядом кладём маленький файл с готовыми числами: последний месяц,
    год с начала, топ-категория и топ-товар. Считается из тех же DS,
    что и сам дашборд, поэтому расхождений между ними быть не может.
    """
    import re as _re
    mk = sorted(k for k in ds if _re.match(r"^\d{4}-\d{2}$", k))
    if not mk:
        return
    cur, prev = ds[mk[-1]], (ds[mk[-2]] if len(mk) > 1 else None)
    cats = cur.get("categories") or []
    top = (cur.get("top20") or [])
    year = ds.get("year") or {}
    out = {
        "month": mk[-1],
        "label": cur.get("label") or mk[-1],
        "rev": cur.get("total_rev") or 0,
        "gp": cur.get("total_gp") or 0,
        "gpEst": bool(cur.get("gp_est")),
        "sku": cur.get("sku_count") or 0,
        "magRev": cur.get("mag_rev") or 0,
        "magPct": cur.get("mag_pct") or 0,
        "prevRev": (prev or {}).get("total_rev") or 0,
        "prevLabel": (prev or {}).get("label") or "",
        "yearRev": year.get("total_rev") or sum((ds[k].get("total_rev") or 0) for k in mk),
        "yearGp": year.get("total_gp") or sum((ds[k].get("total_gp") or 0) for k in mk),
        "months": [{"k": k, "rev": ds[k].get("total_rev") or 0} for k in mk],
        "topCat": ({"n": cats[0].get("cat"), "rev": cats[0].get("rev"), "pct": cats[0].get("pct")} if cats else None),
        "topSku": ({"n": top[0].get("name"), "rev": top[0].get("rev"), "qty": top[0].get("qty")} if top else None),
    }
    path = os.path.join(HERE, "sales_sum.js")
    open(path, "w", encoding="utf-8").write(
        "window.SALES_SUM=" + json.dumps(out, ensure_ascii=False) + ";")
    print("sales_sum.js: %s, выручка %.1f млн" % (out["label"], out["rev"] / 1e6))


def inject_ds(html, ds):
    i = html.find('const DS =')
    seg = html[i + len('const DS ='):]
    depth = 0; end = -1; st = False
    for j, ch in enumerate(seg):
        if ch == '{':
            depth += 1; st = True
        elif ch == '}':
            depth -= 1
            if st and depth == 0:
                end = j + 1; break
    return html[:i] + 'const DS = ' + json.dumps(ds, ensure_ascii=False) + seg[end:]



RETURNS_ANCHOR = '<div class="card"><canvas id="ch-year" height="340"></canvas></div>'
RETURNS_BLOCK = '\n  <div class="card" id="returns-card" style="margin-top:12px;display:none">\n    <div style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px;margin-bottom:10px">\n      <div style="font-size:14px;font-weight:700;color:#f1f5f9">&#8617;&#65038; Возвраты покупателей <span style="font-weight:500;font-size:12px;color:#94a3b8">— уже вычтены из выручки на графике выше</span></div>\n      <div id="returns-total" style="font-size:13px;color:#c9a94e;font-weight:700"></div>\n    </div>\n    <div id="returns-chips" style="display:flex;flex-wrap:wrap;gap:8px"></div>\n  </div>\n  <script>\n  (function(){\n    function fmt(v){ v=Math.round(Math.abs(v));\n      if(v>=1e6) return (v/1e6).toFixed(1).replace(".",",")+" млн";\n      if(v>=1e3) return Math.round(v/1e3)+" тыс"; return String(v); }\n    function render(){\n      var card=document.getElementById("returns-card"); if(!card) return;\n      var R=(window.SALES_META&&window.SALES_META.returns)||null;\n      var keys=R?Object.keys(R).sort():[];\n      if(!keys.length){ card.style.display="none"; return; }\n      var MN=["","Янв","Фев","Мар","Апр","Май","Июн","Июл","Авг","Сен","Окт","Ноя","Дек"];\n      var tot=0, gtot=0, chips="";\n      keys.forEach(function(k){\n        var val=R[k]||0, mo=parseInt(k.split("-")[1],10);\n        var net=(window.DS&&window.DS[k]&&window.DS[k].total_rev)||0, gross=net+val;\n        tot+=val; gtot+=gross;\n        var pctTxt=gross>0?" · "+(val/gross*100).toFixed(1).replace(".",",")+"%":"";\n        chips+=\'<span style="background:#0f172a;border:1px solid #334155;border-radius:9px;padding:6px 10px;font-size:12px;color:#cbd5e1">\'\n             +\'<b style="color:#e2e8f0">\'+(MN[mo]||k)+\'</b> &minus;\'+fmt(val)\n             +\'<span style="color:#94a3b8">\'+pctTxt+\'</span></span>\';\n      });\n      document.getElementById("returns-chips").innerHTML=chips;\n      var gpct=gtot>0?" ("+(tot/gtot*100).toFixed(1).replace(".",",")+"% от выручки)":"";\n      document.getElementById("returns-total").textContent="Итого возвраты: −"+fmt(tot)+gpct;\n      card.style.display="block";\n    }\n    if(document.readyState==="loading"){ document.addEventListener("DOMContentLoaded",render); } else { render(); }\n  })();\n  </script>'

def inject_returns_block(html):
    """Вставляет блок «Возвраты покупателей» под графиком выручки (идемпотентно)."""
    if 'id="returns-card"' in html:
        return html
    i = html.find(RETURNS_ANCHOR)
    if i < 0:
        return html
    j = i + len(RETURNS_ANCHOR)
    return html[:j] + RETURNS_BLOCK + html[j:]



ANALYTICS_ANCHOR = '<div class="section" style="padding-top:6px">\n  <details class="opiu-check"'
RETURNS_ANALYTICS_SECTION = '<div class="section" id="returns-analytics" style="padding-top:6px" data-rv="9">\n  <details id="ra-details" class="opiu-check" style="background:#1e293b;border:1px solid #334155;border-radius:14px;padding:0;overflow:hidden">\n    <summary style="cursor:pointer;list-style:none;padding:14px 18px;font-size:14px;font-weight:700;color:#f1f5f9;display:flex;align-items:center;gap:8px;user-select:none;flex-wrap:wrap">\n      <span style="color:#c9a94e"><span id="ra-caret">▸</span> 📉 Аналитика возвратов</span>\n      <span id="ra-sum" style="font-weight:600;font-size:12px;color:#e2896b"></span>\n      <span style="font-weight:500;font-size:12px;color:#94a3b8;margin-left:auto">нажмите, чтобы раскрыть</span>\n    </summary>\n    <div style="padding:4px 14px 16px">\n      <div id="ra-warn" style="display:none;background:rgba(251,191,36,.10);border:1px solid rgba(251,191,36,.4);border-radius:10px;padding:10px 13px;font-size:12.5px;line-height:1.55;color:#fcd34d;margin:2px 0 12px"></div>\n      <div id="ra-note2" style="display:none;background:rgba(96,165,250,.10);border:1px solid rgba(96,165,250,.35);border-radius:10px;padding:10px 13px;font-size:12.5px;line-height:1.55;color:#bfdbfe;margin:2px 0 12px"></div>\n      <div class="card" style="height:320px"><canvas id="ch-returns"></canvas></div>\n      <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:12px;margin-top:12px">\n        <div class="card"><div style="font-size:13px;font-weight:700;color:#f1f5f9;margin-bottom:6px">🏢 По контрагентам <span style="font-weight:500;font-size:11px;color:#64748b">— клик по сектору фильтрует список</span></div><div style="height:250px;position:relative"><canvas id="ch-ret-ctr"></canvas></div></div>\n        <div class="card"><div style="font-size:13px;font-weight:700;color:#f1f5f9;margin-bottom:6px">📦 По товарам <span style="font-weight:500;font-size:11px;color:#64748b">— клик по сектору фильтрует список</span></div><div style="height:250px;position:relative"><canvas id="ch-ret-prod"></canvas></div></div>\n      </div>\n      <div class="card" style="margin-top:12px">\n        <div style="display:flex;flex-wrap:wrap;gap:8px;align-items:center;margin-bottom:12px">\n          <div id="rf-tabs" style="display:inline-flex;background:#0f172a;border:1px solid #334155;border-radius:9px;padding:2px">\n            <button type="button" data-tab="contractors" style="border:0;background:transparent;color:#cbd5e1;font-size:12px;font-weight:600;padding:6px 12px;border-radius:7px;cursor:pointer">🏢 Контрагенты</button>\n            <button type="button" data-tab="products" style="border:0;background:transparent;color:#cbd5e1;font-size:12px;font-weight:600;padding:6px 12px;border-radius:7px;cursor:pointer">📦 Товары</button>\n          </div>\n          <select id="rf-month" style="background:#0f172a;color:#e2e8f0;border:1px solid #334155;border-radius:8px;padding:6px 10px;font-size:12px;cursor:pointer"></select>\n          <select id="rf-sort" style="background:#0f172a;color:#e2e8f0;border:1px solid #334155;border-radius:8px;padding:6px 10px;font-size:12px;cursor:pointer">\n            <option value="ret">сортировка: по сумме</option>\n            <option value="share">сортировка: по доле %</option>\n            <option value="name">сортировка: по названию</option>\n          </select>\n          <input id="rf-search" type="text" placeholder="Поиск по названию…" style="flex:1;min-width:150px;background:#0f172a;color:#e2e8f0;border:1px solid #334155;border-radius:8px;padding:6px 10px;font-size:12px">\n          <span id="rf-count" style="color:#94a3b8;font-size:12px;white-space:nowrap"></span>\n        </div>\n        <div id="rf-head" style="display:flex;justify-content:space-between;align-items:center;font-size:11px;color:#64748b;font-weight:600;letter-spacing:.03em;padding:2px 2px 6px;border-bottom:1px solid #334155;margin-bottom:2px"><span id="rf-head-l">Контрагент</span><span>Возвраты, ₸ · доля</span></div>\n        <div id="rf-list"></div>\n      </div>\n    </div>\n  </details>\n  <script>\n  (function(){\n    function fmt(v){ v=Math.round(Math.abs(v)); if(v>=1e6) return (v/1e6).toFixed(1).replace(".",",")+" млн"; if(v>=1e3) return Math.round(v/1e3)+" тыс"; return String(v); }\n    function esc(s){ return String(s).replace(/[&<>]/g,function(c){return {"&":"&amp;","<":"&lt;",">":"&gt;"}[c];}); }\n    var MN=["","Январь","Февраль","Март","Апрель","Май","Июнь","Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"];\n    var MS=["","Янв","Фев","Мар","Апр","Май","Июн","Июл","Авг","Сен","Окт","Ноя","Дек"];\n    var PAL=["#60a5fa","#f59e0b","#34d399","#a78bfa","#f472b6","#22d3ee","#fb923c","#4ade80","#e879f9","#94a3b8"];\n    var KNOWN={"90":"Маймарт","110":"Sinooil","99":"КЗ","105":"Алга экспресс"};\n    var GCOLOR={"90":"#ef4444","110":"#38bdf8","99":"#f59e0b","105":"#a78bfa"};\n    function numKey(n){ var m=String(n).match(/^(\\d+)/); return m?m[1]:null; }\n    function brandOf(name){ var s=String(name).replace(/^\\d+\\s*[-–]?\\s*/,"").split(",")[0].trim(); if(s.length>26) s=s.slice(0,26)+"…"; return s; }\n    function colorFor(k,i){ return GCOLOR[k]||PAL[(i||0)%PAL.length]; }\n    var LIM=12;\n    var st={tab:"contractors",month:"",sort:"ret",q:"",showAll:false,expand:{}};\n    function rowVal(e){\n      if(!st.month){ return {ret:e.r, gross:e.g, share:e.s, qty:e.q}; }\n      var mm=e.m&&e.m[st.month];\n      if(!mm){ return {ret:0, gross:0, share:null, qty:null}; }\n      return {ret:mm[0], gross:mm[1], share:mm[1]>0?+(mm[0]/mm[1]*100).toFixed(1):null, qty:null};\n    }\n    function sortRows(a){\n      if(st.sort==="ret") a.sort(function(x,y){return y.ret-x.ret;});\n      else if(st.sort==="share") a.sort(function(x,y){return (y.share||0)-(x.share||0);});\n      else a.sort(function(x,y){return String(x.n||x.gname).localeCompare(String(y.n||y.gname),"ru");});\n    }\n    // авто-группировка контрагентов по ведущему номеру (одинаковый номер => одна сеть)\n    function groupContractors(arr){\n      var byNum={}, singles=[];\n      arr.forEach(function(r){ var k=numKey(r.n); if(k){ (byNum[k]=byNum[k]||[]).push(r); } else singles.push(r); });\n      var groups=[], ci=0;\n      Object.keys(byNum).forEach(function(k){\n        var mem=byNum[k];\n        if(mem.length<2){ singles.push(mem[0]); return; }\n        var gr=mem.reduce(function(s,r){return s+r.ret;},0), gg=mem.reduce(function(s,r){return s+(r.gross||0);},0);\n        sortRows(mem);\n        groups.push({group:true,gid:"g"+k,key:k,gname:k+" "+(KNOWN[k]||brandOf(mem[0].n)),color:colorFor(k,ci++),members:mem,ret:gr,gross:gg,share:gg>0?+(gr/gg*100).toFixed(1):null,count:mem.length});\n      });\n      return {groups:groups,singles:singles};\n    }\n    function bar(w,col){ return \'<div style="height:6px;background:#0f172a;border-radius:4px;overflow:hidden"><div style="height:100%;width:\'+w.toFixed(1)+\'%;background:\'+col+\';border-radius:4px"></div></div>\'; }\n    function rowHTML(r,i,max,indent){\n      var w=max>0?(r.ret/max*100):0;\n      var sh=(r.share!=null)?(\'<span style="color:#94a3b8;font-weight:500"> · \'+String(r.share).replace(".",",")+"% возвр.</span>"):\'\';\n      var qt=(r.qty!=null&&r.qty>0)?(\'<span style="color:#64748b;font-weight:500"> · \'+r.qty+" шт</span>"):\'\';\n      var pad=indent?\'padding-left:16px;border-left:2px solid #334155;margin-left:6px;\':\'\';\n      var nm=indent?esc(String(r.n).replace(/^\\d+\\s*[-–]?\\s*/,"")):esc(r.n);\n      var num=indent?"":(\'<span style="color:#475569;font-variant-numeric:tabular-nums">\'+i+\'.</span> \');\n      return \'<div style="padding:6px 2px;\'+pad+\'border-bottom:1px solid #1b2636">\'\n        +\'<div style="display:flex;justify-content:space-between;gap:10px;font-size:12.5px;color:#cbd5e1;margin-bottom:4px">\'\n        +\'<span style="max-width:56%;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">\'+num+nm+\'</span>\'\n        +\'<span style="white-space:nowrap;color:#e2e8f0;font-weight:700">−\'+fmt(r.ret)+sh+qt+\'</span></div>\'\n        +bar(w,"#e2896b")+\'</div>\';\n    }\n    function groupHTML(g,max,expanded){\n      var w=max>0?(g.ret/max*100):0;\n      var sh=(g.share!=null)?(\'<span style="color:#94a3b8;font-weight:500"> · \'+String(g.share).replace(".",",")+"% возвр.</span>"):\'\';\n      return \'<div data-grp="\'+g.gid+\'" style="padding:8px 2px;cursor:pointer;border-bottom:1px solid #1b2636">\'\n        +\'<div style="display:flex;justify-content:space-between;gap:10px;font-size:13px;color:#f1f5f9;font-weight:700;margin-bottom:4px">\'\n        +\'<span style="max-width:56%;overflow:hidden;text-overflow:ellipsis;white-space:nowrap"><span style="color:#c9a94e">\'+(expanded?"▾":"▸")+\'</span> 🏪 \'+esc(g.gname)+\' <span style="color:#64748b;font-weight:500">\'+g.count+" точек</span></span>"\n        +\'<span style="white-space:nowrap;color:#fff;font-weight:800">−\'+fmt(g.ret)+sh+\'</span></div>\'\n        +bar(w,g.color)+\'</div>\';\n    }\n    function renderList(){\n      var D=window.RETURNS_DETAIL; if(!D) return;\n      var raw=(D[st.tab]||[]).map(function(e){ var v=rowVal(e); return {n:e.n,ret:v.ret,gross:v.gross,share:v.share,qty:v.qty}; }).filter(function(r){ return r.ret>0; });\n      if(st.q){ var qq=st.q.toLowerCase(); raw=raw.filter(function(r){ return r.n.toLowerCase().indexOf(qq)>=0; }); }\n      var grouping=(st.tab==="contractors" && !st.q);\n      var rows;\n      if(grouping){ var gc=groupContractors(raw); rows=gc.singles.concat(gc.groups); }\n      else { rows=raw; }\n      sortRows(rows);\n      var max=rows.length?Math.max.apply(null,rows.map(function(r){return r.ret;})):0;\n      var tot=raw.reduce(function(s,r){return s+r.ret;},0);\n      var units=[], i=0;\n      rows.forEach(function(r){\n        if(r.group){ units.push(groupHTML(r,max,st.expand[r.gid]));\n          if(st.expand[r.gid]){ var mmax=r.members.length?Math.max.apply(null,r.members.map(function(x){return x.ret;})):0; r.members.forEach(function(mm){ units.push(rowHTML(mm,0,mmax,true)); }); } }\n        else { i++; units.push(rowHTML(r,i,max,false)); }\n      });\n      var el=document.getElementById("rf-list");\n      if(!units.length){ el.innerHTML=\'<div style="color:#94a3b8;font-size:12px;padding:10px">Ничего не найдено</div>\'; }\n      else{\n        var shown=st.showAll?units:units.slice(0,LIM), html=shown.join("");\n        if(units.length>LIM){ html+=\'<div data-showall="1" style="text-align:center;padding:10px 0 2px;cursor:pointer;color:#c9a94e;font-size:12.5px;font-weight:600">\'+(st.showAll?"▴ Свернуть список":("▾ Показать все ("+units.length+")"))+\'</div>\'; }\n        el.innerHTML=html;\n      }\n      var lbl=st.month?(MN[parseInt(st.month.split("-")[1],10)]||st.month):"с начала года";\n      document.getElementById("rf-count").textContent=(st.tab==="contractors"?"Контрагентов ":"Товаров ")+raw.length+" · "+lbl+" · итого −"+fmt(tot);\n      var hl=document.getElementById("rf-head-l"); if(hl) hl.textContent=(st.tab==="contractors"?"Контрагент / сеть":"Товар");\n    }\n    function styleTabs(){\n      var t=document.getElementById("rf-tabs"); if(!t) return;\n      [].forEach.call(t.querySelectorAll("button"),function(b){ var on=b.getAttribute("data-tab")===st.tab; b.style.background=on?"#334155":"transparent"; b.style.color=on?"#fff":"#cbd5e1"; });\n    }\n    function fillMonths(){\n      var D=window.RETURNS_DETAIL, sel=document.getElementById("rf-month"); if(!sel||!D) return;\n      var opts=\'<option value="">Все месяцы</option>\';\n      (D.months||[]).forEach(function(k){ opts+=\'<option value="\'+k+\'">\'+(MN[parseInt(k.split("-")[1],10)]||k)+\'</option>\'; });\n      sel.innerHTML=opts;\n    }\n    /* месяц, в котором возвраты явно ещё не проведены: меньше 30% от медианы остальных */\n    function unpostedMonth(){\n      var D=window.RETURNS_DETAIL; if(!D) return null;\n      var BM=D.all_by_month||D.by_month; if(!BM) return null;\n      var ks=Object.keys(BM).sort(); if(ks.length<3) return null;\n      var last=ks[ks.length-1], prev=ks.slice(0,-1).map(function(k){return BM[k]||0;}).sort(function(a,b){return a-b;});\n      var med=prev[Math.floor(prev.length/2)];\n      if(med>0 && (BM[last]||0) < med*0.3) return {k:last,v:BM[last]||0,med:med};\n      return null;\n    }\n    function setSummary(){\n      var D=window.RETURNS_DETAIL, e=document.getElementById("ra-sum"); if(!e||!D) return;\n      var t=D.total||Object.keys(D.by_month||{}).reduce(function(s,k){return s+(D.by_month[k]||0);},0);\n      var sv=D.svc_total||0, all=(D.all_total||t+sv);\n      e.textContent="возвраты −"+fmt(all)+" с начала года"+(sv?(" (из них −"+fmt(sv)+" актами услуг)"):"");\n      var u=unpostedMonth(), w=document.getElementById("ra-warn");\n      if(w&&u){\n        var nm=MN[parseInt(u.k.split("-")[1],10)]||u.k;\n        w.innerHTML="⚠️ "+nm+": возвраты ещё не проведены — "+fmt(u.v)+" ₸ против медианы "+fmt(u.med)\n          +". Выручка "+nm.toLowerCase()+"а завышена примерно на "+fmt(u.med)+" и уменьшится, когда возвраты закроют.";\n        w.style.display="";\n      } else if(w){ w.style.display="none"; }\n      var w2=document.getElementById("ra-note2");\n      if(w2){\n        if(D.svc_total){\n          var mk=Object.keys(D.svc_by_month||{}).sort();\n          var MG=["","января","февраля","марта","апреля","мая","июня","июля","августа","сентября","октября","ноября","декабря"];\n          w2.innerHTML="С "+(mk.length?(MG[parseInt(mk[0].split(\'-\')[1],10)]||mk[0]):"недавнего времени")\n            +" часть возвратов проводится не обратной реализацией, а <b>актом приёма услуг</b> (счёт «"\n            +(D.svc_account||"Торговая выручка")+"»): "+fmt(D.svc_total)+" ₸. В таком документе нет номенклатуры, "\n            +"поэтому по товарам эти возвраты не расписаны и из выручки по SKU <b>не вычтены</b> — "\n            +"в разрезе товаров ниже их нет, а в помесячном графике они показаны отдельным цветом.";\n          w2.style.display="";\n        } else { w2.style.display="none"; }\n      }\n    }\n    function popList(tab){ return (window.RETURNS_DETAIL[tab]||[]).map(function(e){ var v=rowVal(e); return {n:e.n,ret:v.ret,gross:v.gross}; }).filter(function(r){ return r.ret>0; }); }\n    function slices(arr,contractors){\n      var items;\n      if(contractors){ var gc=groupContractors(arr); items=gc.singles.map(function(r){return {n:r.n,ret:r.ret};}); gc.groups.forEach(function(g){ items.push({n:g.gname,ret:g.ret,color:g.color,grp:true,key:g.key}); }); }\n      else items=arr.map(function(r){return {n:r.n,ret:r.ret};});\n      items.sort(function(a,b){return b.ret-a.ret;});\n      var TOP=8, top=items.slice(0,TOP), rest=items.slice(TOP).reduce(function(s,r){return s+r.ret;},0);\n      if(rest>0) top.push({n:"Прочие",ret:rest,rest:true});\n      return top;\n    }\n    function centerPlugin(tref){ return {id:"ctr"+Math.round(tref.v%9973),afterDraw:function(ch){var a=ch.chartArea;if(!a)return;var x=(a.left+a.right)/2,y=(a.top+a.bottom)/2,c=ch.ctx;c.save();c.textAlign="center";c.textBaseline="middle";c.fillStyle="#f1f5f9";c.font="800 18px system-ui,-apple-system,sans-serif";c.fillText("−"+fmt(tref.v),x,y-8);c.fillStyle="#64748b";c.font="600 10px system-ui,-apple-system,sans-serif";c.fillText("возвраты",x,y+12);c.restore();}}; }\n    function buildDonut(cid,sl,onClick){\n      var cv=document.getElementById(cid); if(!cv||!window.Chart) return;\n      try{var ex=Chart.getChart?Chart.getChart(cv):null; if(ex) ex.destroy();}catch(e){}\n      if(!sl.length){ return; }\n      var total=sl.reduce(function(s,r){return s+r.ret;},0), tref={v:total};\n      var cols=sl.map(function(r,i){ return r.rest?"#475569":(r.color||PAL[i%PAL.length]); });\n      /* подписи рисуем сами: плагин datalabels падает на этих бубликах\n         (они строятся внутри закрытого <details>, координаты секторов ещё null) */\n      var faces={id:"rd"+cid,afterDraw:function(ch){\n        var a=ch.chartArea; if(!a) return;\n        var meta=ch.getDatasetMeta(0); if(!meta||!meta.data) return;\n        var dd=ch.data.datasets[0].data, c=ch.ctx;\n        c.save(); c.textAlign="center"; c.textBaseline="middle";\n        meta.data.forEach(function(el,i){\n          if(!el||typeof el.x!=="number"||typeof el.startAngle!=="number") return;\n          var v=dd[i]||0, p=total?v/total*100:0; if(!(p>=7)) return;\n          var ang=(el.startAngle+el.endAngle)/2, rr=(el.innerRadius+el.outerRadius)/2;\n          var lx=el.x+Math.cos(ang)*rr, ly=el.y+Math.sin(ang)*rr;\n          c.font="800 10px system-ui,-apple-system,sans-serif";\n          c.shadowColor="rgba(0,0,0,.9)"; c.shadowBlur=6;\n          c.fillStyle="#fff"; c.fillText(Math.round(p)+"%",lx,ly);\n          c.shadowBlur=0;\n        });\n        var x=(a.left+a.right)/2, y=(a.top+a.bottom)/2;\n        c.fillStyle="#f1f5f9"; c.font="800 16px system-ui,-apple-system,sans-serif";\n        c.fillText("−"+fmt(tref.v),x,y-7);\n        c.fillStyle="#64748b"; c.font="600 10px system-ui,-apple-system,sans-serif";\n        c.fillText("возвраты",x,y+11);\n        c.restore();\n      }};\n      new Chart(cv.getContext("2d"),{type:"doughnut",\n        data:{labels:sl.map(function(r){return r.n;}),datasets:[{data:sl.map(function(r){return r.ret;}),backgroundColor:cols,borderColor:"#0f172a",borderWidth:2,hoverOffset:9}]},\n        options:{responsive:true,maintainAspectRatio:false,cutout:"62%",layout:{padding:6},onClick:onClick,\n          plugins:{legend:{display:false},\n            tooltip:{callbacks:{label:function(c){var v=c.parsed;return " "+c.label+": −"+fmt(v)+" ("+(v/total*100).toFixed(1).replace(".",",")+"%)";}}},\n            datalabels:{display:false}}},\n        plugins:[faces]});\n    }\n    function renderDonuts(){\n      if(!window.RETURNS_DETAIL||!window.Chart) return;\n      var cs=slices(popList("contractors"),true), ps=slices(popList("products"),false);\n      buildDonut("ch-ret-ctr",cs,function(e,els){ if(!els.length)return; var sl=cs[els[0].index]; st.tab="contractors"; st.showAll=false; styleTabs();\n        if(sl.grp){ st.expand["g"+sl.key]=true; st.q=""; } else if(sl.rest){ st.q=""; } else { st.q=sl.n; }\n        var se=document.getElementById("rf-search"); if(se) se.value=(sl.grp||sl.rest)?"":sl.n; renderList(); });\n      buildDonut("ch-ret-prod",ps,function(e,els){ if(!els.length)return; var sl=ps[els[0].index]; st.tab="products"; st.showAll=false; styleTabs();\n        st.q=sl.rest?"":sl.n; var se=document.getElementById("rf-search"); if(se) se.value=st.q; renderList(); });\n    }\n    function renderChart(){\n      var D=window.RETURNS_DETAIL; if(!D) return;\n      var BM=D.all_by_month||D.by_month||{};\n      var keys=Object.keys(BM).sort();\n      var labels=keys.map(function(k){return MS[parseInt(k.split("-")[1],10)]||k;});\n      var vals=keys.map(function(k){return +(((D.by_month[k]||0))/1e6).toFixed(2);});\n      var svals=keys.map(function(k){return +((((D.svc_by_month||{})[k]||0))/1e6).toFixed(2);});\n      var hasSvc=svals.some(function(v){return v>0;});\n      var pcts=keys.map(function(k){var net=(window.DS&&window.DS[k]&&window.DS[k].total_rev)||0;var g=net+(BM[k]||0);return g>0?+(((BM[k]||0)/g*100)).toFixed(2):0;});\n      var cv=document.getElementById("ch-returns");\n      if(cv && window.Chart){\n        try{var ex=Chart.getChart?Chart.getChart(cv):null; if(ex) ex.destroy();}catch(e){}\n        new Chart(cv.getContext("2d"),{type:"bar",\n          data:{labels:labels,datasets:[\n            {label:hasSvc?"Обратной реализацией":"Возвраты, ₸",data:vals,backgroundColor:"#e2896b",borderRadius:6,yAxisID:"y",order:3,stack:"r"},\n            {label:"Актом приёма услуг",data:svals,backgroundColor:"#60a5fa",borderRadius:6,yAxisID:"y",order:2,stack:"r",hidden:!hasSvc},\n            {label:"% от выручки",data:pcts,type:"line",borderColor:"#c9a94e",backgroundColor:"#c9a94e",borderWidth:2,tension:.3,pointRadius:3,yAxisID:"y1",order:1}\n          ]},\n          options:{responsive:true,maintainAspectRatio:false,\n            plugins:{legend:{labels:{color:"#cbd5e1",font:{size:12}}},\n              tooltip:{callbacks:{label:function(c){return c.dataset.label+": "+(c.datasetIndex===0?(c.parsed.y+" М"):(String(c.parsed.y).replace(".",",")+"%"));}}},\n              datalabels:{display:false}},\n            scales:{\n              x:{stacked:true,ticks:{color:"#94a3b8",font:{size:12,weight:"600"}},grid:{display:false}},\n              y:{stacked:true,position:"left",beginAtZero:true,ticks:{color:"#64748b",font:{size:10},callback:function(v){return v+" М";}},grid:{color:"rgba(51,65,85,.4)"}},\n              y1:{position:"right",beginAtZero:true,ticks:{color:"#c9a94e",font:{size:10},callback:function(v){return v+"%";}},grid:{display:false}}\n            }}});\n      }\n    }\n    function wire(){\n      var t=document.getElementById("rf-tabs");\n      if(t) t.addEventListener("click",function(e){ var b=e.target.closest?e.target.closest("button"):null; if(!b)return; st.tab=b.getAttribute("data-tab"); st.showAll=false; styleTabs(); renderList(); });\n      var mo=document.getElementById("rf-month"); if(mo) mo.addEventListener("change",function(){ st.month=this.value; renderList(); renderDonuts(); });\n      var so=document.getElementById("rf-sort"); if(so) so.addEventListener("change",function(){ st.sort=this.value; renderList(); });\n      var se=document.getElementById("rf-search"); if(se) se.addEventListener("input",function(){ st.q=this.value; st.showAll=false; renderList(); });\n      var lst=document.getElementById("rf-list");\n      if(lst) lst.addEventListener("click",function(e){\n        var g=e.target.closest?e.target.closest("[data-grp]"):null;\n        if(g){ var id=g.getAttribute("data-grp"); st.expand[id]=!st.expand[id]; renderList(); return; }\n        var sa=e.target.closest?e.target.closest("[data-showall]"):null;\n        if(sa){ st.showAll=!st.showAll; renderList(); }\n      });\n      var det=document.getElementById("ra-details");\n      if(det) det.addEventListener("toggle",function(){ var c=document.getElementById("ra-caret"); if(c) c.textContent=det.open?"▾":"▸"; if(det.open){ renderChart(); renderDonuts(); renderList(); } });\n    }\n    function render(){ if(!window.RETURNS_DETAIL) return; fillMonths(); styleTabs(); setSummary(); renderList(); var det=document.getElementById("ra-details"); if(det&&det.open){ renderChart(); renderDonuts(); } }\n    function boot(){ wire(); if(window.Chart){ render(); } else { var n=0,t=setInterval(function(){ n++; if(window.Chart||n>50){ clearInterval(t); render(); } },100); } }\n    var s=document.createElement("script");\n    s.src="/returns_meta.js?v="+(window.SALES_META&&window.SALES_META.pulled?encodeURIComponent(window.SALES_META.pulled):"1");\n    s.onload=function(){ if(document.readyState==="loading"){ document.addEventListener("DOMContentLoaded", boot); } else { boot(); } };\n    s.onerror=function(){};\n    (document.head||document.documentElement).appendChild(s);\n  })();\n  </script>\n</div>\n\n'

def inject_returns_analytics(html):
    """Заменяет/вставляет раздел «Аналитика возвратов» (обновляется каждый прогон)."""
    anchor_i = html.find(ANALYTICS_ANCHOR)
    if anchor_i < 0:
        return html
    start = html.find('<div class="section" id="returns-analytics"')
    if 0 <= start < anchor_i:
        html = html[:start] + html[anchor_i:]
        anchor_i = html.find(ANALYTICS_ANCHOR)
    return html[:anchor_i] + RETURNS_ANALYTICS_SECTION + html[anchor_i:]


def inject_opiu_columns(html):
    """Столбцы «I продажи» и «Возвраты» в таблице сверки ОПиУ (идемпотентно)."""
    if 'I продажи, млн' in html:
        return html
    reps = [
        ('<th>Продажи, млн ₸</th>', '<th>I продажи, млн</th><th>Возвраты, млн</th><th>Продажи, млн ₸</th>'),
        ('var O=window.OPIU_REV||{}, DS=window.DS||{};', 'var O=window.OPIU_REV||{}, DS=window.DS||{}; var R=(window.SALES_META&&window.SALES_META.returns)||{};'),
        ('var sP=0,sT=0,sI=0,html=', 'var sP=0,sT=0,sI=0,sG=0,sR=0,html='),
        ('var p=(DS[k]&&typeof DS[k].total_rev==="number")?DS[k].total_rev:null;', 'var p=(DS[k]&&typeof DS[k].total_rev==="number")?DS[k].total_rev:null; var ret=R[k]||0; var gross=(p!==null)?p+ret:null;'),
        ('\'</td><td>\'+(p!==null?mln(p):"—")+\'</td>', '\'</td><td>\'+(gross!==null?mln(gross):"—")+\'</td><td class="mut">\'+(ret?("−"+mln(ret)):"—")+\'</td><td>\'+(p!==null?mln(p):"—")+\'</td>'),
        ('sP+=p; sT+=t; sI+=it;', 'sP+=p; sT+=t; sI+=it; sG+=gross; sR+=ret;'),
        ("'+nm+'</td><td>'+mln(p)+'</td>", '\'+nm+\'</td><td>\'+mln(gross)+\'</td><td class="mut">−\'+mln(ret)+\'</td><td>\'+mln(p)+\'</td>'),
        ("Итого (сошедшиеся месяцы)</td><td>'+mln(sP)+'</td>", 'Итого (сошедшиеся месяцы)</td><td>\'+mln(sG)+\'</td><td class="mut">−\'+mln(sR)+\'</td><td>\'+mln(sP)+\'</td>'),
    ]
    for _o, _n in reps:
        if _o in html:
            html = html.replace(_o, _n, 1)
    return html


def inject_footer(html):
    """Футер с автором внизу страницы (идемпотентно)."""
    if 'psig-sales' in html:
        return html
    foot = '<footer id="psig-sales" style="text-align:center;color:#475569;font-size:12px;line-height:1.7;padding:24px 16px 36px;border-top:1px solid #1e293b;margin-top:26px;font-family:Inter,-apple-system,sans-serif">Система «Пульс» · Фуд завод (Мастерская Сегодня)<br><b style="color:#94a3b8">Ольга Герасименко</b> · финансовый директор · данные из iiko, обновляются ежедневно<br><a href="/metrics.html" style="color:#a78bfa;text-decoration:none;font-weight:600">🔒 Аналитика посещений</a></footer>'
    if '</body>' in html:
        return html.replace('</body>', foot + '\n</body>', 1)
    if '</html>' in html:
        return html.replace('</html>', foot + '\n</html>', 1)
    return html + foot


def inject_bar_style(html):
    """Единый стиль двух сворачивающихся блоков (идемпотентно)."""
    if 'dc-caret' in html:
        return html
    html = html.replace('details.opiu-check summary::-webkit-details-marker{display:none}', 'details.opiu-check summary::-webkit-details-marker{display:none}\n.dc-caret{display:inline-block;transition:transform .2s;color:#c9a94e}\ndetails[open] .dc-caret{transform:rotate(90deg)}', 1)
    html = html.replace('    <summary style="cursor:pointer;list-style:none;padding:14px 18px;font-size:14px;font-weight:700;color:#f1f5f9;display:flex;align-items:center;gap:8px;user-select:none">\n      <span style="color:#c9a94e">&#128269; Сверка выручки с ОПиУ</span>\n      <span style="font-weight:500;font-size:12px;color:#94a3b8">— нажмите, чтобы раскрыть</span>\n      <span style="margin-left:auto;color:#64748b;font-size:12px" class="opiu-hint">&#9660;</span>\n    </summary>', '    <summary style="cursor:pointer;list-style:none;padding:14px 18px;font-size:14px;font-weight:700;color:#f1f5f9;display:flex;align-items:center;gap:8px;user-select:none">\n      <span style="color:#c9a94e"><span class="dc-caret">▸</span> &#128269; Сверка выручки с ОПиУ</span>\n      <span style="font-weight:500;font-size:12px;color:#94a3b8;margin-left:auto">нажмите, чтобы раскрыть</span>\n    </summary>', 1)
    html = html.replace('id="ra-caret"', 'class="dc-caret"', 1)
    return html


def inject_beacon(html):
    """Счётчик просмотров для вкладки Метрики (идемпотентно)."""
    if '/track?p=' in html:
        return html
    tag = '<script>try{fetch("/track?p="+encodeURIComponent(location.pathname),{method:"GET",keepalive:true})}catch(e){}</script>'
    if '</head>' in html:
        return html.replace('</head>', tag + '</head>', 1)
    if '</html>' in html:
        return html.replace('</html>', tag + '\n</html>', 1)
    return html + tag


def main():
    log("=" * 60)
    log("  Пересборка продаж из выгрузок «I Отчет ПРОДАЖИ»")
    log("=" * 60)

    html, old = load_html_and_old_ds()
    year = 2026
    months = {}
    year_months = []

    reports = scan_reports(HERE, year)
    log(f"  Найдены выгрузки по месяцам: {sorted(reports.keys())}\n")

    all_items = defaultdict(lambda: {'rev': 0.0, 'qty': 0.0})   # накопление за год

    for m in range(1, 13):
        if m not in reports:
            continue
        path = reports[m][0]
        log(f"  · {RU_MONTHS[m]}: {os.path.basename(path)}")
        period, items = parse_report(path)
        if not items:
            log(f"  {RU_MONTHS[m]:8} — пусто, пропуск ({os.path.basename(path)})")
            continue
        for _n, _v in items.items():
            all_items[_n]['rev'] += _v['rev']
            all_items[_n]['qty'] += _v['qty']

        md = build_month(items)
        label, partial = period_label(m, period)
        md['label'] = label

        mk = f"{year}-{m:02d}"
        gp_old = (old.get(mk) or {}).get('total_gp')
        md['total_gp'] = gp_old if gp_old else 0
        if mk in old:
            for k, v in old[mk].items():
                if k not in md:
                    md[k] = v
        months[mk] = md
        year_months.append({'mk': mk, 'label': label, 'rev': md['total_rev'],
                            'gp': gp_old if gp_old else None,
                            **({'partial': True} if partial else {})})
        old_rev = (old.get(mk) or {}).get('total_rev', 0)
        diff = md['total_rev'] - old_rev
        log(f"  {label:16} {md['total_rev']:>14,}   было {old_rev:>14,}   Δ {diff:>+13,}   SKU={md['sku_count']}")

    # ── валовая прибыль неполного месяца ───────────────────────────────
    # Выгрузка продаж не содержит себестоимости, а перенесённая ВП относится
    # к прежнему (более короткому) периоду — из-за этого маржа занижалась.
    # Для неполного месяца оцениваем ВП по средней марже последних полных
    # месяцев, где ВП известна, и помечаем как оценку.
    full_m = [x for x in year_months if not x.get('partial') and x['gp'] and x['rev']]
    if full_m:
        base = full_m[-2:]
        margin = sum(x['gp'] for x in base) / sum(x['rev'] for x in base)
        for x in year_months:
            if x.get('partial'):
                est = round(x['rev'] * margin)
                x['gp'] = est
                x['gp_est'] = True
                months[x['mk']]['total_gp'] = est
                months[x['mk']]['gp_est'] = True
                log(f"  ~ {x['label']}: ВП оценена по марже {margin*100:.1f}% → {est:,} "
                    f"(была перенесена за другой период)")

    total_rev = sum(x['rev'] for x in year_months)
    total_gp = sum(x['gp'] for x in year_months if x['gp'])
    last_m = int(year_months[-1]['mk'][5:7]) if year_months else 1
    ds = dict(months)
    # Год оформляем так же, как месяц: категории, топ-SKU, Magnum — но итогово
    ya = build_month(all_items)
    ds['year'] = {'label': f"{year} — Янв–{RU_MONTHS[last_m][:3]}",
                  'total_rev': total_rev, 'total_gp': total_gp,
                  'is_year': True, 'months': year_months,
                  'categories': ya['categories'], 'top20': ya['top20'],
                  'magnum_items': ya['magnum_items'], 'mag_rev': ya['mag_rev'],
                  'mag_pct': ya['mag_pct'], 'sku_count': ya['sku_count']}
    log(f"  {'ГОД (итого)':16} категорий={len(ya['categories'])}  SKU={ya['sku_count']}")

    old_year = (old.get('year') or {}).get('total_rev', 0)
    log("-" * 60)
    log(f"  {'ИТОГО ГОД':16} {total_rev:>14,}   было {old_year:>14,}   Δ {total_rev-old_year:>+13,}")

    # ── Защита от пустой пересборки ────────────────────────────────────
    # 02.09.2026 iiko лежал (ServerState=WAITING_LICENSE), выгрузок продаж не
    # появилось, scan_reports вернул пусто — и страница была перезаписана
    # набором данных без единого месяца. График выручки исчез, период
    # схлопнулся в «Янв–Янв», а возвраты показали «100,0% от выручки», потому
    # что выручка стала нулём. Поэтому сверяем новую сборку с прежней и, если
    # месяцев стало меньше или год просел больше чем на треть, ничего не пишем.
    old_months = [k for k in old if re.match(r'^\d{4}-\d{2}$', k)]
    old_year_rev = (old.get('year') or {}).get('total_rev', 0)
    if old_months and (len(months) < len(old_months) or
                       (old_year_rev and total_rev < old_year_rev * 0.67)):
        log("")
        log("  " + "=" * 56)
        log("  ОСТАНОВЛЕНО: новая сборка беднее прежней.")
        log(f"    было:  {len(old_months)} мес., год {old_year_rev:,} ₸")
        log(f"    стало: {len(months)} мес., год {total_rev:,} ₸")
        if not reports:
            log("    Выгрузок «I Отчет ПРОДАЖИ» не найдено вовсе —")
            log("    скорее всего, не отработал шаг выгрузки из iiko.")
        log("  Ни продажи_2026.html, ни sales_sum.js не тронуты.")
        log("  " + "=" * 56)
        sys.exit(1)

    bak = HTML_FILE + ".bak2"
    if not os.path.exists(bak):
        open(bak, 'w', encoding='utf-8').write(html)
    open(HTML_FILE, 'w', encoding='utf-8').write(inject_beacon(inject_bar_style(inject_footer(inject_opiu_columns(inject_returns_analytics(inject_returns_block(inject_ds(html, ds))))))))
    write_sales_sum(ds)
    log(f"  Записано: {HTML_FILE}")
    log("  Готово.")


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        import traceback
        log("  ОШИБКА:", e)
        traceback.print_exc()
        sys.exit(1)
